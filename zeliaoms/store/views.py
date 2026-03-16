
from django.shortcuts import render, redirect,get_object_or_404
from django.http import JsonResponse
from django.http import HttpResponse
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required, user_passes_test
from .forms import *
from .models import *
from django.db import transaction
from decimal import Decimal
import json
from django.utils import timezone
# import pandas as pd
from io import BytesIO
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter
from reportlab.lib.units import inch
from django.http import FileResponse
from datetime import timedelta
import datetime
from django.forms import formset_factory
from django.contrib import messages
from django.db.models.functions import TruncMonth
from django.db.models import Sum,F,ExpressionWrapper, FloatField, Count,Q,Value,Avg
from django.core.files.uploadedfile import InMemoryUploadedFile
from PIL import Image as PILImage

# Custom error handlers
from django.db.models import Sum, Count, Q, DecimalField

from decimal import Decimal
from django.db.models.functions import Coalesce

from django.forms import inlineformset_factory
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle,Image, PageBreak
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

import base64
from django.contrib.staticfiles.finders import find
# custom

from django.core.files.storage import default_storage
from django.conf import settings
import os
import csv
import io

from reportlab.lib import colors

from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus.flowables import Flowable
from reportlab.lib.colors import HexColor
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
# 


from django.http import HttpResponseServerError, HttpResponseForbidden, HttpResponseNotFound
from pathlib import Path


def custom_404(request, exception):
    return render(request, '404.html', status=404)

def custom_403(request, exception):
    return render(request, '403.html', status=403)

def custom_500(request):
    return render(request, '500.html', status=500)
def index(request):
    return render(request, "auth/login.html")
def accounts(request):
    return render(request, "auth/account.html")

from django.shortcuts import render, redirect
from django.contrib.auth import authenticate, login, get_user_model
from django.contrib import messages
from django.core.cache import cache
from django.core.mail import send_mail
from django.utils.crypto import get_random_string
from django.utils import timezone
from datetime import timedelta
from django.conf import settings

LOCKOUT_LIMIT = 5  # failed attempts before OTP
LOCKOUT_TIME = 300  # 5 minutes lockout
OTP_EXPIRE = 300    # 5 minutes OTP validity

User = get_user_model()


def _save_login_session_from_post(request, user):
    """
    Helper: save LoginSession record from data submitted with the login form
    (photo base64 + GPS coordinates captured by the frontend JS).
    """
    import base64 as _b64_inner
    from django.core.files.base import ContentFile as _CF

    photo_data  = request.POST.get('login_photo_b64', '')
    lat         = request.POST.get('login_lat') or None
    lng         = request.POST.get('login_lng') or None
    device_info = request.META.get('HTTP_USER_AGENT', '')[:500]

    session = LoginSession(
        user=user,
        latitude=lat or None,
        longitude=lng or None,
        ip_address=request.META.get('REMOTE_ADDR'),
        device_info=device_info,
    )
    if photo_data and photo_data.startswith('data:image'):
        try:
            _, encoded = photo_data.split(',', 1)
            image_bytes = _b64_inner.b64decode(encoded)
            fname = f"login_{user.id}_{timezone.now().strftime('%Y%m%d%H%M%S')}.jpg"
            session.login_photo.save(fname, _CF(image_bytes), save=False)
        except Exception:
            pass
    try:
        session.save()
    except Exception:
        pass  # Never block login due to session-save failure

def send_otp_to_user(user, otp):
    """
    Sends OTP to user's email.
    """
    subject = "Your Zelia OMS OTP"
    message = f"Hello {user.username},\n\nYour OTP is: {otp}\nIt is valid for 5 minutes."
    from_email = settings.DEFAULT_FROM_EMAIL
    recipient_list = [user.email]
    send_mail(subject, message, from_email, recipient_list)
    print(f"OTP sent to {user.email}: {otp}")  # for debugging only

def check_user_type(request):
    """Return whether this username requires camera/location verification.
    Admin/superuser accounts are exempt; all other users must verify.
    Returns {skip_verification: true/false}.
    """
    username = request.GET.get('u', '').strip()
    if not username:
        return JsonResponse({'skip_verification': False})
    try:
        u = User.objects.get(username=username)
        skip = u.is_staff or u.is_superuser
    except User.DoesNotExist:
        skip = False  # Unknown user — safe default: require verification
    return JsonResponse({'skip_verification': skip})


def login_user(request):
    show_otp = False
    lockout_seconds = None
    otp_seconds = None

    if request.method == "POST":
        username = request.POST.get('username')
        password = request.POST.get('password')
        otp_input = request.POST.get('otp')
        resend_otp = request.POST.get('resend_otp')

        # Retrieve User object
        try:
            user_obj = User.objects.get(username=username)
        except User.DoesNotExist:
            user_obj = None

        # Cache keys
        cache_key_attempts = f"login_attempts_{username}"
        cache_key_otp = f"otp_{username}"
        cache_key_otp_time = f"otp_time_{username}"
        cache_key_pw = f"pw_{username}"

        # Resend OTP
        if resend_otp and user_obj:
            otp = get_random_string(6, allowed_chars='0123456789')
            cache.set(cache_key_otp, otp, OTP_EXPIRE)
            cache.set(cache_key_otp_time, timezone.now(), OTP_EXPIRE)
            send_otp_to_user(user_obj, otp)
            show_otp = True
            otp_seconds = OTP_EXPIRE
            messages.success(request, "OTP resent successfully.")
            return render(request, "auth/login.html", {"show_otp": show_otp, "otp_seconds": otp_seconds})

        # OTP verification
        if otp_input and user_obj:
            otp_cached = cache.get(cache_key_otp)
            if otp_cached and otp_input == str(otp_cached):
                user = authenticate(request, username=username, password=cache.get(cache_key_pw))
                if user:
                    # Non-admin users must provide camera photo + GPS even via OTP
                    if not (user.is_staff or user.is_superuser):
                        login_photo = request.POST.get('login_photo_b64', '').strip()
                        login_lat   = request.POST.get('login_lat',       '').strip()
                        login_lng   = request.POST.get('login_lng',       '').strip()
                        if not login_photo or not login_lat or not login_lng:
                            messages.error(request, 'Camera photo and location are required to login. Please enable them and try again.')
                            return render(request, 'auth/login.html', {'show_otp': True, 'otp_seconds': OTP_EXPIRE})
                    login(request, user)
                    # Clear caches
                    cache.delete(cache_key_attempts)
                    cache.delete(cache_key_otp)
                    cache.delete(cache_key_pw)
                    _save_login_session_from_post(request, user)
                    messages.success(request, "Logged in successfully!")
                    return redirect("home")
            else:
                show_otp = True
                otp_time = cache.get(cache_key_otp_time)
                otp_seconds = max(0, OTP_EXPIRE - int((timezone.now() - otp_time).total_seconds())) if otp_time else OTP_EXPIRE
                messages.error(request, "Invalid OTP.")
                return render(request, "auth/login.html", {"show_otp": show_otp, "otp_seconds": otp_seconds})

        # Normal login
        user = authenticate(request, username=username, password=password)
        attempts = cache.get(cache_key_attempts, 0)

        if user:
            # Non-admin users must provide camera photo + GPS
            if not (user.is_staff or user.is_superuser):
                login_photo = request.POST.get('login_photo_b64', '').strip()
                login_lat   = request.POST.get('login_lat',       '').strip()
                login_lng   = request.POST.get('login_lng',       '').strip()
                if not login_photo or not login_lat or not login_lng:
                    messages.error(request, 'Camera photo and location are required to login. Please enable them and try again.')
                    return render(request, 'auth/login.html', {'show_otp': False, 'lockout_seconds': None, 'otp_seconds': None})
            login(request, user)
            cache.delete(cache_key_attempts)
            # Save live login session (photo + GPS) if provided by frontend JS
            _save_login_session_from_post(request, user)
            messages.success(request, "Logged in successfully!")
            return redirect("home")
        else:
            attempts += 1
            cache.set(cache_key_attempts, attempts, LOCKOUT_TIME)
            if attempts >= LOCKOUT_LIMIT and user_obj:
                otp = get_random_string(6, allowed_chars='0123456789')
                cache.set(cache_key_otp, otp, OTP_EXPIRE)
                cache.set(cache_key_otp_time, timezone.now(), OTP_EXPIRE)
                cache.set(cache_key_pw, password, OTP_EXPIRE)
                send_otp_to_user(user_obj, otp)
                show_otp = True
                otp_seconds = OTP_EXPIRE
                messages.warning(request, "Too many failed attempts. Enter OTP sent to your email.")
            else:
                messages.error(request, "Invalid credentials.")

    # GET request: show inactivity timeout message
    if request.method == "GET" and request.GET.get('to') == '1':
        messages.warning(request, 'You were logged out due to 30 minutes of inactivity.')

    # GET request: check if user is locked out
    if request.method == "GET" and 'username' in request.GET:
        username = request.GET.get('username')
        attempts = cache.get(f"login_attempts_{username}", 0)
        if attempts >= LOCKOUT_LIMIT:
            show_otp = True
            otp_seconds = OTP_EXPIRE

    return render(request, "auth/login.html", {
        "show_otp": show_otp,
        "lockout_seconds": lockout_seconds,
        "otp_seconds": otp_seconds
    })

@login_required
def logout_user(request):
    logout(request)
    messages.success(request, 'Logout successful!')
    return redirect('loginuser')
def registeruser(request):
    return render(request, 'auth/register.html')
@login_required
def update_user_profile(request):
    # Add your logic here
    if request.user.is_authenticated:
        current_user=User.objects.get(id=request.user.id)
        user_form = UpdateUserForm(request.POST or None, instance=current_user)
        if user_form.is_valid():
            user_form.save()
            login(request, current_user)
            messages.success(request, "User Has Been updated!!!")
            return redirect("home")
        return render(request, "auth/updateprofile.html", {"user_form": user_form})
    else:
        messages.success(request, "you must be logged in")
        return redirect("registeruser")
# update user profile
@login_required
def update_info_profile(request):
    try:
        current_user = UserProfile.objects.get(user=request.user)
    except UserProfile.DoesNotExist:
        messages.error(request, "User profile not found.")
        return redirect("home")
    if request.method == 'POST':
        form = UserInfoForm(request.POST, instance=current_user, user=request.user)
        if form.is_valid():
            form.save()
            messages.success(request, "Your Info Has Been Updated Successfully")
            return redirect("home")
        else:
            messages.error(request, "Please correct the errors below.")
    else:
        form = UserInfoForm(instance=current_user, user=request.user)
    return render(request, "auth/update_info.html", {"form": form})
# update password
@login_required
def update_password(request):
    if request.user.is_authenticated:
        current_user=request.user
        # did he/she fill the form
        if request.method == "POST":
            user_form=UpdatePasswordForm(current_user,request.POST)
            if user_form.is_valid():
                user_form.save()
                
                messages.success(request,"Password updated successfully")
                login(request, current_user)
                return redirect("home")
            else:
                return render(request, "auth/updatepassword.html", {"user_form": user_form})
                # for error in list(user_form.errors.values()):
                #     messages.error(request,error)
                #     return redirect("update_password")
        else:
            user_form=UpdatePasswordForm(current_user)
            return render(request,"auth/updatepassword.html",{"user_form":user_form})
    else:
        messages.success(request,"You must be logged in")
        return redirect("registeruser")
# analytics dashboard
# views.py
# views.py

def admin_required(view_func):
    decorated_view = user_passes_test(
        lambda u: hasattr(u, 'userprofile') and u.userprofile.is_admin(),
        login_url='loginuser/'
    )(view_func)
    return decorated_view
@login_required
@admin_required
def admin_dashboard(request):
    user = request.user
    try:
        user_profile = UserProfile.objects.get(user=user)
    except UserProfile.DoesNotExist:
        return render(request, '403.html', {'error': 'User profile not found'}, status=403)

    today = timezone.now().date()
    this_month_start = today.replace(day=1)
    last_month_start = (this_month_start - timedelta(days=1)).replace(day=1)
    last_month_end = this_month_start - timedelta(days=1)

    # Customer Metrics
    total_customers = Customer.objects.count()
    customers_this_month = Customer.objects.filter(created_at__gte=this_month_start).count()
    customers_last_month = Customer.objects.filter(
        created_at__gte=last_month_start, created_at__lte=last_month_end
    ).count()
    customer_percentage_change = (
        ((customers_this_month - customers_last_month) / customers_last_month * 100)
        if customers_last_month > 0 else 0
    )

    # Revenue Metrics
    total_revenue = Order.objects.filter(
        order_date__gte=this_month_start, paid_status='completed'
    ).aggregate(total=Sum('total_amount'))['total'] or 0
    revenue_last_month = Order.objects.filter(
        order_date__gte=last_month_start, order_date__lte=last_month_end, paid_status='completed'
    ).aggregate(total=Sum('total_amount'))['total'] or 0
    revenue_percentage_change = (
        ((total_revenue - revenue_last_month) / revenue_last_month * 100)
        if revenue_last_month > 0 else 0
    )

    total_revenue_this_month = Order.objects.filter(
        order_date__gte=this_month_start
    ).aggregate(total=Sum('total_amount'))['total'] or 0

    pending_revenue = Order.objects.filter(
        order_date__gte=this_month_start, paid_status='pending'
    ).aggregate(total=Sum('total_amount'))['total'] or 0

    # Order Metrics
    start_of_day = timezone.make_aware(timezone.datetime.combine(today, timezone.datetime.min.time()))
    end_of_day = start_of_day + timedelta(days=1)
    orders_today = Order.objects.filter(
        order_date__gte=start_of_day, order_date__lt=end_of_day
    ).count()
    orders_yesterday = Order.objects.filter(
        order_date__date=today - timedelta(days=1)
    ).count()
    orders_percentage_change = (
        ((orders_today - orders_yesterday) / orders_yesterday * 100)
        if orders_yesterday > 0 else 0
    )

    # Completed Orders (Deals)
    total_deals = Order.objects.filter(paid_status='completed').count()
    deals_this_month = Order.objects.filter(
        order_date__gte=this_month_start, paid_status='completed'
    ).count()
    deals_last_month = Order.objects.filter(
        order_date__gte=last_month_start, order_date__lte=last_month_end, paid_status='completed'
    ).count()
    deals_percentage_change = (
        ((deals_this_month - deals_last_month) / deals_last_month * 100)
        if deals_last_month > 0 else 0
    )

    # Recent Orders
    recent_orders = (
        Order.objects.select_related('customer')
        .prefetch_related('order_items__product')
        .order_by('-order_date')[:5]
    )

    # Top Products
    total_units_sold = OrderItem.objects.filter(
        order__order_date__gte=this_month_start
    ).aggregate(total=Sum('quantity'))['total'] or 1
    top_products = (
        OrderItem.objects.filter(order__order_date__gte=this_month_start)
        .values('product__name')
        .annotate(total_units=Sum('quantity'))
        .order_by('-total_units')[:5]
    )
    top_products = [
        {
            'product__name': item['product__name'],
            'total_units': item['total_units'],
            'percent': (item['total_units'] / total_units_sold) * 100
        }
        for item in top_products
    ]

    # Products Overview
    products = Product.objects.select_related('category').annotate(
        total_revenue=Sum('orderitem__line_total'),
        sales_count=Count('orderitem')
    )[:50]

    # Chart Data
    chart_data = list(
        Order.objects.filter(order_date__gte=last_month_start)
        .values('order_date__month')
        .annotate(total=Sum('total_amount'))
        .order_by('order_date__month')
        .values_list('total', flat=True)
    ) or [1000, 2000, 1500, 3000, 2500, 4000, 3500]

    context = {
        'total_customers': total_customers,
        'customer_percentage_change': customer_percentage_change,
        'total_revenue': total_revenue,
        'revenue_percentage_change': revenue_percentage_change,
        'total_revenue_this_month': total_revenue_this_month,
        'pending_revenue': pending_revenue,
        'orders_today': orders_today,
        'orders_percentage_change': orders_percentage_change,
        'total_deals': total_deals,
        'deals_percentage_change': deals_percentage_change,
        'recent_orders': recent_orders,
        'top_products': top_products,
        'products': products,
        'chart_data': chart_data,
    }

    return render(request, 'dashboard/admin.html', context)

@login_required
def salesperson_dashboard(request):
    user = request.user
    try:
        user_profile = UserProfile.objects.get(user=user)
    except UserProfile.DoesNotExist:
        return render(request, '403.html', {'error': 'User profile not found'}, status=403)

    if user_profile.is_admin():
        return redirect('admin_dashboard')

    if not user_profile.is_salesperson():
        return render(request, '403.html', {'error': 'You are not authorized to view this dashboard'}, status=403)

    today = timezone.now().date()
    this_month_start = today.replace(day=1)
    last_month_start = (this_month_start - timedelta(days=1)).replace(day=1)
    last_month_end = this_month_start - timedelta(days=1)

    # Customer Metrics
    total_customers = Customer.objects.filter(sales_person=user).count()
    customers_this_month = Customer.objects.filter(
        sales_person=user, created_at__gte=this_month_start
    ).count()
    customers_last_month = Customer.objects.filter(
        sales_person=user, created_at__gte=last_month_start, created_at__lte=last_month_end
    ).count()
    customer_percentage_change = (
        ((customers_this_month - customers_last_month) / customers_last_month * 100)
        if customers_last_month > 0 else 0
    )

    # Revenue Metrics
    total_revenue = Order.objects.filter(
        sales_person=user, order_date__gte=this_month_start, paid_status='completed'
    ).aggregate(total=Sum('total_amount'))['total'] or 0
    revenue_last_month = Order.objects.filter(
        sales_person=user, order_date__gte=last_month_start,
        order_date__lte=last_month_end, paid_status='completed'
    ).aggregate(total=Sum('total_amount'))['total'] or 0
    revenue_percentage_change = (
        ((total_revenue - revenue_last_month) / revenue_last_month * 100)
        if revenue_last_month > 0 else 0
    )

    total_revenue_month = Order.objects.filter(
        sales_person=user, order_date__gte=this_month_start
    ).aggregate(total=Sum('total_amount'))['total'] or 0

    pending_revenue = Order.objects.filter(
        sales_person=user, order_date__gte=this_month_start, paid_status='pending'
    ).aggregate(total=Sum('total_amount'))['total'] or 0

    # Order Metrics
    start_of_day = timezone.make_aware(timezone.datetime.combine(today, timezone.datetime.min.time()))
    end_of_day = start_of_day + timedelta(days=1)
    orders_today = Order.objects.filter(
        sales_person=user, order_date__gte=start_of_day, order_date__lt=end_of_day
    ).count()
    orders_yesterday = Order.objects.filter(
        sales_person=user, order_date__date=today - timedelta(days=1)
    ).count()
    orders_percentage_change = (
        ((orders_today - orders_yesterday) / orders_yesterday * 100)
        if orders_yesterday > 0 else 0
    )

    # Completed Orders (Deals)
    total_deals = Order.objects.filter(sales_person=user, paid_status='completed').count()
    deals_this_month = Order.objects.filter(
        sales_person=user, order_date__gte=this_month_start, paid_status='completed'
    ).count()
    deals_last_month = Order.objects.filter(
        sales_person=user, order_date__gte=last_month_start,
        order_date__lte=last_month_end, paid_status='completed'
    ).count()
    deals_percentage_change = (
        ((deals_this_month - deals_last_month) / deals_last_month * 100)
        if deals_last_month > 0 else 0
    )

    # Recent Orders
    recent_orders = (
        Order.objects.filter(sales_person=user)
        .select_related('customer')
        .prefetch_related('order_items__product')
        .order_by('-order_date')[:5]
    )

    # Top Products
    total_units_sold = (
        OrderItem.objects.filter(
            order__sales_person=user, order__order_date__gte=this_month_start
        ).aggregate(total=Sum('quantity'))['total'] or 1
    )
    top_products = (
        OrderItem.objects.filter(
            order__sales_person=user, order__order_date__gte=this_month_start
        )
        .values('product__name')
        .annotate(total_units=Sum('quantity'))
        .order_by('-total_units')[:5]
    )
    top_products = [
        {
            'product__name': item['product__name'],
            'total_units': item['total_units'],
            'percent': (item['total_units'] / total_units_sold) * 100
        }
        for item in top_products
    ]

    # Chart Data
    chart_data = list(
        Order.objects.filter(
            sales_person=user, order_date__gte=last_month_start
        )
        .values('order_date__month')
        .annotate(total=Sum('total_amount'))
        .order_by('order_date__month')
        .values_list('total', flat=True)
    ) or [1000, 2000, 1500, 3000, 2500, 4000, 3500]

    context = {
        'total_customers': total_customers,
        'customer_percentage_change': customer_percentage_change,
        'total_revenue': total_revenue,
        'revenue_percentage_change': revenue_percentage_change,
        'total_revenue_this_month': total_revenue_month,
        'pending_revenue': pending_revenue,
        'orders_today': orders_today,
        'orders_percentage_change': orders_percentage_change,
        'total_deals': total_deals,
        'deals_percentage_change': deals_percentage_change,
        'recent_orders': recent_orders,
        'top_products': top_products,
        'chart_data': chart_data,
    }

    return render(request, 'dashboard/sales_dashboard.html', context)



@login_required
def home(request):
    try:
        user_profile = UserProfile.objects.get(user=request.user)
    except UserProfile.DoesNotExist:
        return render(request, '403.html', {'error': 'User profile not found'}, status=403)

    if user_profile.is_admin():
        return redirect('admin_dashboard')
    elif user_profile.is_salesperson():
        return redirect('salesperson_dashboard')
    else:
        return render(request, '403.html', {'error': 'Invalid user role'}, status=403)

from django.db.models import Q
from reportlab.graphics.barcode import code128
from rest_framework.decorators import api_view
from rest_framework.response import Response

@login_required
def create_order(request):
    """Create a new order with items"""
    OrderItemFormSet = inlineformset_factory(
        Order,
        OrderItem,
        form=OrderItemForm,
        extra=0,
        can_delete=True,
        min_num=0,
        validate_min=False,
        max_num=30
    )

    VAT_RATE = Decimal('0.16')

    if request.method == 'POST':
        form = OrderForm(request.POST, user=request.user)
        formset = OrderItemFormSet(request.POST, instance=form.instance)

        if form.is_valid() and formset.is_valid():
            try:
                with transaction.atomic():
                    order = form.save(commit=False)
                    order.sales_person = request.user
                    
                    # PRIORITY 1: Use live location from form input (should be populated by GPS)
                    form_address = request.POST.get('address', '').strip()
                    
                    # PRIORITY 2: Get live location address if GPS captured it
                    live_location = request.POST.get('location_address', '').strip()
                    
                    # PRIORITY 3: Fallback to customer's address only if both above are empty
                    if form_address:
                        order.address = form_address[:500]
                    elif live_location:
                        order.address = live_location[:500]
                    elif order.customer and order.customer.address:
                        order.address = order.customer.address[:500]
                    
                    # Capture live GPS from hidden form fields
                    _lat = request.POST.get('latitude')
                    _lng = request.POST.get('longitude')
                    _loc = request.POST.get('location_address', '')
                    if _lat:
                        try:
                            order.latitude = Decimal(_lat)
                        except Exception:
                            pass
                    if _lng:
                        try:
                            order.longitude = Decimal(_lng)
                        except Exception:
                            pass
                    if _loc:
                        order.location_address = _loc[:500]
                    order.save()
                    formset.instance = order

                    for form in formset:
                        if form.cleaned_data.get('product') and not form.cleaned_data.get('DELETE'):
                            product = form.cleaned_data['product']
                            quantity = form.cleaned_data['quantity']
                            variance = form.cleaned_data.get('variance', Decimal('0.00'))
                            store = order.store
                            stock_field = f"{store}_stock"
                            current_stock = getattr(product, stock_field, 0)
                            if current_stock < quantity:
                                raise ValueError(f"Not enough stock for {product.name} in {store} store")

                            price_map = {
                                'factory': product.factory_price,
                                'distributor': product.distributor_price,
                                'wholesale': product.wholesale_price,
                                'Towns': product.offshore_price,
                                'Retail customer': product.retail_price,
                            }
                            base_price = price_map.get(order.customer_category, product.retail_price)
                            unit_price = base_price * (1 + VAT_RATE) if order.vat_variation == 'with_vat' else base_price
                            unit_price = Decimal(round(float(unit_price)))

                            item = form.instance
                            item.order = order
                            item.product = product
                            item.quantity = quantity
                            item.unit_price = unit_price
                            item.variance = variance
                            item.line_total = quantity * (unit_price + variance)
                            item.save()

                            setattr(product, stock_field, current_stock - quantity)
                            product.save()

                    for form in formset:
                        if form.cleaned_data.get('DELETE') and form.instance.pk:
                            form.instance.delete()

                    items_total = sum(item.line_total for item in order.order_items.all()) or Decimal('0.00')
                    order.total_amount = items_total + (order.delivery_fee or Decimal('0.00'))

                    # ── Handle payment at order creation ──────────────
                    pay_method = request.POST.get('pay_method_hidden', 'none')
                    pay_amount_str = request.POST.get('pay_amount', '0')
                    pay_phone = request.POST.get('pay_phone', '').strip()
                    try:
                        pay_amount = Decimal(pay_amount_str or '0')
                    except Exception:
                        pay_amount = Decimal('0')

                    if pay_method == 'cash' and pay_amount > 0:
                        order.amount_paid = min(pay_amount, order.total_amount)
                        if order.amount_paid >= order.total_amount:
                            order.paid_status = 'completed'
                        else:
                            order.paid_status = 'partially_paid'
                        Payment.objects.create(
                            order=order,
                            amount=order.amount_paid,
                            payment_method='cash',
                            recorded_by=request.user,
                        )

                    order.save()

                    # Generate PDF receipt
                    salesperson_name = request.user.get_full_name() or request.user.username
                    buffer = BytesIO()
                    doc = SimpleDocTemplate(buffer, pagesize=letter, leftMargin=50, rightMargin=50, topMargin=50, bottomMargin=50)
                    styles = getSampleStyleSheet()
                    elements = []

                    title_style = ParagraphStyle('CustomTitle', parent=styles['Title'], fontName='Helvetica-Bold', fontSize=18,
                                                textColor=colors.white, backColor=colors.HexColor('#1e3a8a'), alignment=1, spaceAfter=6, spaceBefore=6, leading=20)
                    tagline_style = ParagraphStyle('Tagline', parent=styles['Normal'], fontName='Helvetica-Oblique', fontSize=10,
                                                  textColor=colors.white, alignment=0, spaceAfter=8, spaceBefore=8, backColor=colors.HexColor('#3b82f6'),
                                                  borderPadding=5, borderWidth=0.5, borderColor=colors.HexColor('#1e3a8a'))
                    normal_style = ParagraphStyle('CustomNormal', parent=styles['Normal'], fontName='Helvetica', fontSize=10,
                                                 textColor=colors.black, spaceAfter=4)
                    italic_style = ParagraphStyle('CustomItalic', parent=styles['Italic'], fontName='Helvetica-Oblique', fontSize=9,
                                                 textColor=colors.grey, spaceBefore=8)
                    total_style = ParagraphStyle('Total', parent=styles['Heading2'], fontName='Helvetica-Bold', fontSize=14,
                                                textColor=colors.HexColor('#1e3a8a'), alignment=2, spaceBefore=12, spaceAfter=12,
                                                backColor=colors.HexColor('#e0f2fe'), borderPadding=5, borderWidth=1, borderColor=colors.HexColor('#1e3a8a'))
                    contact_style = ParagraphStyle('Contact', parent=styles['Normal'], fontName='Helvetica-Oblique', fontSize=8,
                                                  textColor=colors.HexColor('#4b5563'), alignment=1, spaceBefore=8)

                    logo_path = None
                    possible_paths = [
                        os.path.join(settings.STATIC_ROOT, 'assets/images/mcdave/Ant.jpg') if settings.STATIC_ROOT else None,
                        *[os.path.join(static_dir, 'assets/images/mcdave/Ant.jpg') for static_dir in getattr(settings, 'STATICFILES_DIRS', [])],
                        os.path.join(settings.BASE_DIR, 'static', 'assets/images/mcdave/Ant.jpg')
                    ]
                    for path in possible_paths:
                        if path and os.path.exists(path):
                            logo_path = path
                            break

                    logo_data = []
                    if logo_path:
                        try:
                            logo = Image(logo_path, width=80, height=45)
                            logo_data.append(logo)
                        except Exception:
                            logo_data.append("")
                    else:
                        logo_data.append("")

                    title_data = [
                        Paragraph("Antioch Africa Limited", title_style),
                        Paragraph("Reliable Excellence", tagline_style)
                    ]
                    header_table_data = [[logo_data[0], title_data]]
                    header_table = Table(header_table_data, colWidths=[100, 370])
                    header_table.setStyle([
                        ('VALIGN', (0, 0), (0, 0), 'TOP'),
                        ('VALIGN', (1, 0), (1, 0), 'MIDDLE'),
                        ('LEFTPADDING', (0, 0), (0, 0), 0),
                        ('RIGHTPADDING', (1, 0), (1, 0), 10),
                    ])
                    elements.append(header_table)
                    elements.append(Spacer(1, 12))

                    address_data = [[Paragraph("P.O. Box 12345-00100, Nairobi, Kenya", normal_style)],
                                    [Paragraph("Phone: +254 722 123456 / +254 733 789012", normal_style)],
                                    [Paragraph("Email: info@antioch.co.ke", normal_style)],
                                    [Paragraph("Website: www.antioch.co.ke", normal_style)]]
                    address_table = Table(address_data, colWidths=[512])
                    address_table.setStyle([('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#f8f9ff')),
                                           ('BOX', (0, 0), (-1, -1), 0.5, colors.HexColor('#e0e7ff')),
                                           ('LEFTPADDING', (0, 0), (-1, -1), 12), ('RIGHTPADDING', (0, 0), (-1, -1), 12),
                                           ('TOPPADDING', (0, 0), (-1, -1), 8), ('BOTTOMPADDING', (0, 0), (-1, -1), 8)])
                    elements.append(address_table)
                    elements.append(Spacer(1, 12))

                    barcode_value = f"ORDER-{order.id}"
                    barcode = code128.Code128(barcode_value, barHeight=20, barWidth=0.5)
                    payment_data = [[Paragraph("Paybill No: 522522", normal_style)], [Paragraph("A/c No: 5881754", normal_style)]]
                    payment_table_data = [[barcode, payment_data]]
                    payment_table = Table(payment_table_data, colWidths=[256, 256])
                    payment_table.setStyle([('VALIGN', (0, 0), (0, 0), 'TOP'), ('ALIGN', (0, 0), (0, 0), 'LEFT'),
                                           ('ALIGN', (1, 0), (1, 0), 'RIGHT'), ('LEFTPADDING', (0, 0), (0, 0), 0),
                                           ('RIGHTPADDING', (1, 0), (1, 0), 0)])
                    elements.append(payment_table)
                    elements.append(Spacer(1, 12))

                    divider = Table([[""]], colWidths=[512], rowHeights=[2])
                    divider.setStyle([('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#1e3a8a'))])
                    elements.append(divider)
                    elements.append(Spacer(1, 12))

                    date_str = order.order_date.strftime('%Y-%m-%d') if order.order_date else 'N/A'
                    elements.append(Paragraph(f"Store: {order.get_store_display()}", normal_style))
                    elements.append(Paragraph(f"Date: {date_str}", normal_style))
                    elements.append(Paragraph(f"M/s: {order.customer.first_name if order.customer else 'N/A'}", normal_style))
                    elements.append(Spacer(1, 8))

                    elements.append(Paragraph(f"Location: {order.address if hasattr(order, 'address') and order.address else 'N/A'}", contact_style))
                    elements.append(Spacer(1, 8))

                    table_data = [['Qty', 'Item Description', '@', 'Amount (Ksh)']]
                    for item in order.order_items.all():
                        quantity = item.quantity or 0
                        product_name = item.product.name if item.product else 'Unknown Product'
                        unit_price = item.unit_price or Decimal('0.00')
                        variance = item.variance or Decimal('0.00')
                        item_total = item.line_total or Decimal('0.00')
                        table_data.append([str(quantity), product_name, f"{(unit_price + variance):.0f}", f"{item_total:.2f}"])

                    if len(table_data) == 1:
                        table_data.append(["0", "No items", "0.00", "0.00"])

                    table = Table(table_data, colWidths=[40, 280, 50, 80])
                    table_style = [('GRID', (0, 0), (-1, -1), 0.5, colors.lightgrey),
                                   ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1e3a8a')),
                                   ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                                   ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                                   ('FONTSIZE', (0, 0), (-1, 0), 11),
                                   ('FONTSIZE', (0, 1), (-1, -1), 10),
                                   ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                                   ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                                   ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                                   ('BOX', (0, 0), (-1, -1), 1, colors.HexColor('#1e3a8a')),
                                   ('INNERGRID', (0, 0), (-1, -1), 0.25, colors.lightgrey),
                                   ('LEFTPADDING', (0, 0), (-1, -1), 8),
                                   ('RIGHTPADDING', (0, 0), (-1, -1), 8),
                                   ('TOPPADDING', (0, 0), (-1, -1), 6),
                                   ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
                                   ('ALIGN', (1, 1), (1, -1), 'LEFT')]
                    for i in range(1, len(table_data)):
                        if i % 2 == 0:
                            table_style.append(('BACKGROUND', (0, i), (-1, i), colors.HexColor('#f8f9ff')))
                    table.setStyle(table_style)
                    elements.append(table)

                    # Show subtotal, delivery fee (if > 0), and total
                    items_subtotal = sum(item.line_total for item in order.order_items.all()) or Decimal('0.00')
                    if order.delivery_fee and order.delivery_fee > 0:
                        subtotal_style = ParagraphStyle('Subtotal', parent=styles['Normal'], fontName='Helvetica', fontSize=11,
                                                       textColor=colors.HexColor('#1e3a8a'), alignment=2, spaceBefore=8)
                        elements.append(Paragraph(f"Subtotal: {items_subtotal:.2f}", subtotal_style))
                        elements.append(Paragraph(f"Delivery Fee: {order.delivery_fee:.2f}", subtotal_style))
                    elements.append(Paragraph(f"TOTAL: {order.total_amount or 0:.2f}", total_style))

                    footer_table = Table([[Paragraph(f"Receipt: #Mc{order.id}Z", italic_style), Paragraph("Goods once sold cannot be re-accepted", italic_style)]], colWidths=[235, 235])
                    footer_table.setStyle([('TOPPADDING', (0, 0), (-1, -1), 8),
                                          ('LINEABOVE', (0, 0), (-1, -1), 0.5, colors.HexColor('#e0e7ff'))])
                    elements.append(footer_table)
                    elements.append(Paragraph(f"Served by: {salesperson_name}", contact_style))
                    elements.append(Paragraph("Need assistance? Contact us at support@antioch.co.ke", contact_style))

                    # M-Pesa: return redirect signal instead of PDF
                    if pay_method == 'mpesa':
                        buffer.close()
                        return JsonResponse({
                            'success': True,
                            'mpesa_redirect': True,
                            'order_id': order.id,
                            'pay_phone': pay_phone,
                        })

                    doc.build(elements)
                    pdf_data = buffer.getvalue()
                    buffer.close()

                    pdf_base64 = base64.b64encode(pdf_data).decode('utf-8')

                    return JsonResponse({'success': True, 'receipt': pdf_base64})
            except Exception as e:
                return JsonResponse({'success': False, 'errors': str(e)})
        else:
            errors = {'form_errors': form.errors.get_json_data()}
            formset_errors = [form.errors.get_json_data() for form in formset if form.errors]
            errors['formset_errors'] = formset_errors
            return JsonResponse({'success': False, 'errors': json.dumps(errors)})
    else:
        form = OrderForm(user=request.user)
        formset = OrderItemFormSet(instance=Order())

    return render(request, 'addrecords/addorder.html', {'form': form, 'formset': formset})


# store/views.py
@login_required
def order_detail(request, order_id):
    """View order details + generate full, styled, downloadable receipt PDF"""
    order = get_object_or_404(Order, id=order_id)

    # Authorization
    user = request.user
    try:
        user_profile = UserProfile.objects.get(user=user)
    except UserProfile.DoesNotExist:
        return render(request, '403.html', {'error': 'User profile not found'}, status=403)

    if not user_profile.is_admin() and order.sales_person != user:
        return render(request, '403.html',
                     {'error': 'You are not authorized to view this order'}, status=403)

    payments = order.payments.all()
    remaining_balance = order.get_balance()

    # === ORDER ITEMS FOR DISPLAY ===
    items_with_details = []
    total_amount = Decimal('0.00')
    for item in order.order_items.all():
        base_price = item.unit_price or Decimal('0.00')
        variance = item.variance or Decimal('0.00')
        quantity = item.quantity or 0
        final_unit_price = base_price + variance
        line_total = quantity * final_unit_price
        total_amount += line_total

        stock_field = f"{order.store}_stock"
        store_stock = getattr(item.product, stock_field, 0) if hasattr(item.product, stock_field) else 0

        items_with_details.append({
            'item': item,
            'price_with_vat': final_unit_price,
            'store_stock': store_stock,
            'line_total': line_total
        })

    # === PDF RECEIPT GENERATION (FIXED: LOGO + NO CUTOFF) ===
    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=letter,
        leftMargin=40,      # Reduced
        rightMargin=40,     # Reduced
        topMargin=40,       # Reduced
        bottomMargin=40     # Reduced
    )
    styles = getSampleStyleSheet()
    elements = []

    # === STYLES ===
    title_style = ParagraphStyle(
        'CustomTitle', parent=styles['Title'], fontName='Helvetica-Bold', fontSize=18,
        textColor=colors.white, backColor=colors.HexColor('#1e3a8a'), alignment=1,
        spaceAfter=6, spaceBefore=6, leading=22
    )
    tagline_style = ParagraphStyle(
        'Tagline', parent=styles['Normal'], fontName='Helvetica-Oblique', fontSize=10,
        textColor=colors.white, alignment=0, spaceAfter=8, spaceBefore=8,
        backColor=colors.HexColor('#3b82f6'), borderPadding=5, borderWidth=0.5,
        borderColor=colors.HexColor('#1e3a8a')
    )
    normal_style = ParagraphStyle(
        'CustomNormal', parent=styles['Normal'], fontName='Helvetica', fontSize=10,
        textColor=colors.black, spaceAfter=4, leading=12
    )
    italic_style = ParagraphStyle(
        'CustomItalic', parent=styles['Italic'], fontName='Helvetica-Oblique', fontSize=9,
        textColor=colors.grey, spaceBefore=8, leading=11
    )
    total_style = ParagraphStyle(
        'Total', parent=styles['Heading2'], fontName='Helvetica-Bold', fontSize=14,
        textColor=colors.HexColor('#1e3a8a'), alignment=2, spaceBefore=12, spaceAfter=12,
        backColor=colors.HexColor('#e0f2fe'), borderPadding=6, borderWidth=1,
        borderColor=colors.HexColor('#1e3a8a')
    )
    large_italic_style = ParagraphStyle(
        name='LargeItalic', parent=italic_style, fontSize=14, leading=16
    )
    contact_style = ParagraphStyle(
        'Contact', parent=styles['Normal'], fontName='Helvetica-Oblique', fontSize=8,
        textColor=colors.HexColor('#4b5563'), alignment=1, spaceBefore=8, leading=10
    )

    # === LOGO (FIXED PATH + FALLBACK) ===
    logo_path = None
    possible_paths = [
        os.path.join(settings.STATIC_ROOT, 'assets/images/mcdave/Ant.jpg') if settings.STATIC_ROOT else None,
        *[os.path.join(d, 'assets/images/mcdave/Ant.jpg') for d in getattr(settings, 'STATICFILES_DIRS', [])],
        os.path.join(settings.BASE_DIR, 'static', 'assets/images/mcdave/Ant.jpg'),
        os.path.join(settings.BASE_DIR, 'zelia', 'static', 'assets/images/mcdave/Ant.jpg'),
    ]
    for path in possible_paths:
        if path and os.path.exists(path):
            logo_path = path
            break

    logo = None
    if logo_path:
        try:
            logo = Image(logo_path, width=90, height=50)  # Slightly larger
            logo.hAlign = 'LEFT'
        except Exception as e:
            print(f"Logo load error: {e}")
            logo = None

    # === HEADER ===
    company_name = Paragraph("Antioch Africa Limited", title_style)
    tagline = Paragraph("Reliable Excellence", tagline_style)
    header_data = [[logo or "", [company_name, tagline]]]
    header_table = Table(header_data, colWidths=[120, 400])
    header_table.setStyle(TableStyle([
        ('VALIGN', (0, 0), (0, 0), 'TOP'),
        ('VALIGN', (1, 0), (1, 0), 'MIDDLE'),
        ('LEFTPADDING', (0, 0), (0, 0), 0),
        ('RIGHTPADDING', (1, 0), (1, 0), 10),
    ]))
    elements.append(header_table)
    elements.append(Spacer(1, 12))

    # === ADDRESS ===
    address_lines = [
        "P.O. Box 12345-00100, Nairobi, Kenya",
        "Phone: +254 722 123456 / +254 733 789012",
        "Email: info@antioch.co.ke",
        "Website: www.antioch.co.ke"
    ]
    address_data = [[Paragraph(line, normal_style)] for line in address_lines]
    address_table = Table(address_data, colWidths=[520])
    address_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#f8f9ff')),
        ('BOX', (0, 0), (-1, -1), 0.5, colors.HexColor('#e0e7ff')),
        ('LEFTPADDING', (0, 0), (-1, -1), 12),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
    ]))
    elements.append(address_table)
    elements.append(Spacer(1, 12))

    # === BARCODE + PAYBILL ===
    barcode = code128.Code128(f"ORDER-{order.id}", barHeight=25, barWidth=0.55)
    paybill = [
        [Paragraph("Paybill No: 522522", normal_style)],
        [Paragraph("A/c No: 5881754", normal_style)]
    ]
    barcode_table = Table([[barcode, paybill]], colWidths=[260, 260])
    barcode_table.setStyle(TableStyle([
        ('VALIGN', (0, 0), (0, 0), 'TOP'),
        ('ALIGN', (0, 0), (0, 0), 'LEFT'),
        ('ALIGN', (1, 0), (1, 0), 'RIGHT'),
        ('LEFTPADDING', (0, 0), (0, 0), 0),
        ('RIGHTPADDING', (1, 0), (1, 0), 0),
    ]))
    elements.append(barcode_table)
    elements.append(Spacer(1, 12))

    # === DIVIDER ===
    elements.append(Table([[""]], colWidths=[520], rowHeights=[2],
                         style=[('BACKGROUND', (0,0),(-1,-1), colors.HexColor('#1e3a8a'))]))
    elements.append(Spacer(1, 12))

    # === CUSTOMER INFO ===
    customer_name = order.customer.get_full_name() if order.customer else "N/A"
    customer_phone = order.phone or (order.customer.phone_number if order.customer else "N/A")
    date_str = order.order_date.strftime('%Y-%m-%d') if order.order_date else "N/A"
    store_str = order.get_store_display() if order.store else "N/A"
    location = order.address or "N/A"

    info_left = [
        Paragraph(f"Date: {date_str}", normal_style),
        Paragraph(f"M/s: {customer_name}", normal_style),
        Paragraph(f"Phone: {customer_phone}", normal_style),
        Paragraph(f"Store: {store_str}", normal_style),
    ]
    info_right = [Paragraph(f"Location: {location}", italic_style)]
    info_table = Table([[info_left, info_right]], colWidths=[300, 220])
    info_table.setStyle(TableStyle([
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('LEFTPADDING', (0, 0), (-1, -1), 12),
        ('TOPPADDING', (0, 0), (-1, -1), 8),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
    ]))
    elements.append(info_table)
    elements.append(Spacer(1, 10))

    # === ITEMS TABLE (WIDER COLUMNS, NO CUTOFF) ===
    table_data = [['Qty', 'Item Description', 'Variance', '@', 'Amount (Ksh)']]
    for item in order.order_items.all():
        qty = item.quantity or 0
        name = item.product.name if item.product else "Unknown"
        var = item.variance or Decimal('0.00')
        price = item.unit_price or Decimal('0.00')
        final = price + var
        total = qty * final
        table_data.append([
            str(qty),
            name,
            f"{var:+.2f}" if var != 0 else "0.00",
            f"{final:.2f}",
            f"{total:.2f}"
        ])

    if len(table_data) == 1:
        table_data.append(["0", "No items", "0.00", "0.00", "0.00"])

    # WIDER COLUMNS: 45, 280, 65, 60, 90 → total 540
    table = Table(table_data, colWidths=[45, 280, 65, 60, 90])
    table.setStyle(TableStyle([
        ('GRID', (0, 0), (-1, -1), 0.5, colors.lightgrey),
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1e3a8a')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 11),
        ('FONTSIZE', (0, 1), (-1, -1), 10),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('BOX', (0, 0), (-1, -1), 1, colors.HexColor('#1e3a8a')),
        ('INNERGRID', (0, 0), (-1, -1), 0.25, colors.lightgrey),
        ('LEFTPADDING', (0, 0), (-1, -1), 8),
        ('RIGHTPADDING', (0, 0), (-1, -1), 8),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ('ALIGN', (1, 1), (1, -1), 'LEFT'),
        ('ALIGN', (4, 1), (4, -1), 'RIGHT'),
    ]))
    # Alternate row colors
    for i in range(1, len(table_data)):
        if i % 2 == 0:
            table.setStyle(TableStyle([('BACKGROUND', (0, i), (-1, i), colors.HexColor('#f8f9ff'))]))
    elements.append(table)

    # === TOTAL (with delivery fee if > 0) ===
    if order.delivery_fee and order.delivery_fee > 0:
        subtotal_style = ParagraphStyle('Subtotal', parent=styles['Normal'], fontName='Helvetica', fontSize=11,
                                       textColor=colors.HexColor('#1e3a8a'), alignment=2, spaceBefore=8)
        elements.append(Paragraph(f"Subtotal: {total_amount:.2f}", subtotal_style))
        elements.append(Paragraph(f"Delivery Fee: {order.delivery_fee:.2f}", subtotal_style))
        grand_total = total_amount + order.delivery_fee
        elements.append(Paragraph(f"TOTAL: {grand_total:.2f}", total_style))
    else:
        elements.append(Paragraph(f"TOTAL: {total_amount:.2f}", total_style))

    # === FOOTER ===
    salesperson = order.sales_person.get_full_name() if order.sales_person else "Not assigned"
    footer = Table([
        [Paragraph(f"Receipt: #Mc{order.id}Z", large_italic_style),
         Paragraph("Goods once sold cannot be re-accepted", italic_style)]
    ], colWidths=[260, 260])
    footer.setStyle(TableStyle([
        ('TOPPADDING', (0, 0), (-1, -1), 10),
        ('LINEABOVE', (0, 0), (-1, -1), 0.5, colors.HexColor('#e0e7ff')),
    ]))
    elements.append(footer)
    elements.append(Paragraph(f"Served by: {salesperson}", contact_style))
    elements.append(Paragraph("Need assistance? Contact us at support@antioch.co.ke", contact_style))

    # === BUILD PDF ===
    try:
        doc.build(elements)
        pdf_data = buffer.getvalue()
    except Exception as e:
        print(f"PDF Error: {e}")
        pdf_data = None
    buffer.close()

    pdf_base64 = base64.b64encode(pdf_data).decode('utf-8') if pdf_data else ''

    # === CONTEXT ===
    context = {
        'order': order,
        'payments': payments,
        'remaining_balance': remaining_balance,
        'is_admin': user_profile.is_admin(),
        'items_with_vat': items_with_details,
        'pdf_base64': pdf_base64,
        'total_amount': total_amount,
    }

    return render(request, 'view/orderdetail.html', context)




@login_required
def get_customer_details(request):
    """Get customer details for auto-filling form fields"""
    customer_id = request.GET.get('customer_id')
    try:
        customer = Customer.objects.get(id=customer_id)
        data = {
            'default_category': customer.default_category,
            'address': customer.address,
            'phone_number': customer.phone_number
        }
        return JsonResponse(data)
    except Customer.DoesNotExist:
        return JsonResponse({'error': 'Customer not found'}, status=404)


@login_required
def get_product_price(request):
    """Get product price based on category and VAT"""
    product_id = request.GET.get('product_id')
    customer_category = request.GET.get('customer_category')
    vat = request.GET.get('vat')
    store = request.GET.get('store')

    VAT_RATE = Decimal('0.16')

    if not all([product_id, customer_category, vat, store]):
        return JsonResponse({'error': 'Missing required parameters'}, status=400)

    try:
        product = Product.objects.get(id=product_id)

        price_map = {
            'factory': product.factory_price,
            'distributor': product.distributor_price,
            'wholesale': product.wholesale_price,
            'Towns': product.offshore_price,
            'Retail customer': product.retail_price,
        }

        base_price = price_map.get(customer_category, product.retail_price)

        if base_price is None or base_price <= 0:
            return JsonResponse({'error': f'Invalid price for category {customer_category}'}, status=400)

        final_price = base_price * (1 + VAT_RATE) if vat == 'with_vat' else base_price
        final_price = Decimal(round(float(final_price)))

        stock_field = f"{store}_stock"
        stock = getattr(product, stock_field, 0)

        return JsonResponse({
            'price': float(final_price),
            'stock': stock
        })

    except Product.DoesNotExist:
        return JsonResponse({'error': 'Product not found'}, status=404)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=400)


@login_required
def get_product_details_for_edit(request):
    """Get product details including current stock for editing"""
    product_id = request.GET.get('product_id')
    customer_category = request.GET.get('customer_category')
    vat = request.GET.get('vat')
    store = request.GET.get('store')
    item_id = request.GET.get('item_id')  # For existing items

    VAT_RATE = Decimal('0.16')

    if not all([product_id, customer_category, vat, store]):
        return JsonResponse({'error': 'Missing required parameters'}, status=400)

    try:
        product = Product.objects.get(id=product_id)

        price_map = {
            'factory': product.factory_price,
            'distributor': product.distributor_price,
            'wholesale': product.wholesale_price,
            'Towns': product.offshore_price,
            'Retail customer': product.retail_price,
        }

        base_price = price_map.get(customer_category, product.retail_price)
        final_price = base_price * (1 + VAT_RATE) if vat == 'with_vat' else base_price
        final_price = Decimal(round(float(final_price)))

        stock_field = f"{store}_stock"
        stock = getattr(product, stock_field, 0)
        
        # If editing existing item, add back its quantity to available stock
        if item_id:
            try:
                item = OrderItem.objects.get(id=item_id)
                stock += item.original_quantity
            except OrderItem.DoesNotExist:
                pass

        return JsonResponse({
            'price': float(final_price),
            'stock': stock
        })

    except Product.DoesNotExist:
        return JsonResponse({'error': 'Product not found'}, status=404)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=400)


@login_required
def search_customers(request):
    """Search customers for autocomplete"""
    query = request.GET.get('q', '')
    user = request.user
    try:
        user_profile = UserProfile.objects.get(user=user)
    except UserProfile.DoesNotExist:
        return JsonResponse({'error': 'User profile not found'}, status=403)

    if query:
        if user_profile.is_admin():
            customers = Customer.objects.filter(
                Q(first_name__icontains=query) | Q(last_name__icontains=query)
            )[:10]
        else:
            customers = Customer.objects.filter(
                Q(first_name__icontains=query) | Q(last_name__icontains=query)
            ).filter(Q(sales_person=user) | Q(sales_person__isnull=True))[:10]
        results = [
            {
                'id': customer.id,
                'text': f"{customer.get_full_name()}",
                'address': customer.address or '',
                'phone_number': customer.phone_number or '',
                'default_category': customer.default_category or ''
            }
            for customer in customers
        ]
    else:
        results = []
    return JsonResponse({'results': results})


@login_required
def search_products(request):
    """Search products for autocomplete"""
    query = request.GET.get('q', '')
    category = request.GET.get('category', '')
    vat = request.GET.get('vat', '') == 'with_vat'
    store = request.GET.get('store', '')
    
    if query:
        products = Product.objects.filter(Q(name__icontains=query))
        if store:
            stock_field = f"{store}_stock"
            products = products.filter(**{f"{stock_field}__gt": 0})
        products = products[:10]
        
        results = [
            {
                'id': product.id,
                'text': product.name,
                'price': float(round(float(getattr(product, f"{category}_price", product.retail_price) * (Decimal('1.16') if vat else 1))))
            }
            for product in products
        ]
    else:
        results = []
    return JsonResponse({'results': results})


@api_view(['GET'])
def get_product(request, pk):
    """API endpoint to get product details"""
    product = get_object_or_404(Product, pk=pk)
    return Response({
        'factory_price': str(product.factory_price),
        'distributor_price': str(product.distributor_price),
        'wholesale_price': str(product.wholesale_price),
        'offshore_price': str(product.offshore_price),
        'retail_price': str(product.retail_price),
    })


@login_required
def order_list(request):
    user = request.user
    try:
        user_profile = UserProfile.objects.get(user=user)
    except UserProfile.DoesNotExist:
        return render(request, '403.html', {'error': 'User profile not found'}, status=403)

    if user_profile.is_admin():
        orders = Order.objects.select_related('customer', 'sales_person').prefetch_related('order_items__product').order_by('-order_date')
    else:
        orders = Order.objects.filter(sales_person=user).select_related('customer', 'sales_person').prefetch_related('order_items__product').order_by('-order_date')

    orders_with_details = []
    for order in orders:
        items_with_details = []
        for item in order.order_items.all():
            # Pull base price and variance from database
            base_price = item.unit_price
            final_price = base_price + item.variance  # Only calculation: unit_price + variance
            stock_field = f"{order.store}_stock"
            store_stock = getattr(item.product, stock_field, 0)
            items_with_details.append({
                'item': item,
                'base_price': base_price,
                'price_with_vat': final_price,  # Represents unit_price + variance
                'variance': item.variance,
                'store_stock': store_stock
            })
        orders_with_details.append({
            'order': order,
            'items_with_vat': items_with_details  # Kept name for template compatibility
        })

    return render(request, 'dashboard/orders.html', {'orders_with_vat': orders_with_details})



@login_required
def edit_order(request, order_id):
    """Full order editing including products and quantities"""
    user = request.user
    try:
        user_profile = UserProfile.objects.get(user=user)
    except UserProfile.DoesNotExist:
        return render(request, '403.html', {'error': 'User profile not found'}, status=403)

    order = get_object_or_404(Order, id=order_id)
    
    # Authorization check
    if not user_profile.is_admin() and order.sales_person != user:
        return render(request, '403.html', 
                     {'error': 'You are not authorized to edit this order'}, status=403)

    VAT_RATE = Decimal('0.16')

    if request.method == 'POST':
        form = OrderEditForm(request.POST, instance=order, user=request.user)
        formset = OrderItemEditFormSet(request.POST, instance=order)
        
        # Set parent instance for each form in the formset for validation
        for item_form in formset:
            item_form.parent_instance = order

        # Debug: Print form data
        print("=== FORM VALIDATION DEBUG ===")
        print(f"Form valid: {form.is_valid()}")
        print(f"Formset valid: {formset.is_valid()}")
        
        if not form.is_valid():
            print("Form errors:", form.errors)
            for field, errors in form.errors.items():
                print(f"  {field}: {errors}")
        
        if not formset.is_valid():
            print("Formset errors:", formset.errors)
            for i, item_form in enumerate(formset):
                if item_form.errors:
                    print(f"  Item {i} errors:", item_form.errors)
            print("Non-form errors:", formset.non_form_errors())

        if form.is_valid() and formset.is_valid():
            try:
                with transaction.atomic():
                    order = form.save(commit=False)
                    order.save()

                    # Process each item in formset
                    for item_form in formset:
                        if item_form.cleaned_data.get('DELETE'):
                            # Restore stock for deleted items
                            if item_form.instance.pk:
                                product = item_form.instance.product
                                store = order.store
                                stock_field = f"{store}_stock"
                                current_stock = getattr(product, stock_field, 0)
                                setattr(product, stock_field, 
                                       current_stock + item_form.instance.quantity)
                                product.save()
                                item_form.instance.delete()
                        elif item_form.cleaned_data.get('product'):
                            product = item_form.cleaned_data['product']
                            new_quantity = item_form.cleaned_data['quantity']
                            variance = item_form.cleaned_data.get('variance', Decimal('0.00'))
                            store = order.store
                            stock_field = f"{store}_stock"
                            
                            # Calculate price
                            price_map = {
                                'factory': product.factory_price,
                                'distributor': product.distributor_price,
                                'wholesale': product.wholesale_price,
                                'Towns': product.offshore_price,
                                'Retail customer': product.retail_price,
                            }
                            base_price = price_map.get(order.customer_category, product.retail_price)
                            unit_price = base_price * (1 + VAT_RATE) if order.vat_variation == 'with_vat' else base_price
                            unit_price = Decimal(round(float(unit_price)))

                            # Handle stock adjustment
                            if item_form.instance.pk:
                                # Editing existing item
                                old_quantity = item_form.instance.original_quantity
                                quantity_diff = new_quantity - old_quantity
                                
                                current_stock = getattr(product, stock_field, 0)
                                new_stock = current_stock - quantity_diff
                                
                                if new_stock < 0:
                                    raise ValueError(
                                        f"Not enough stock for {product.name} in {store} store"
                                    )
                                
                                setattr(product, stock_field, new_stock)
                                product.save()
                                
                                # Update original quantity
                                item_form.instance.original_quantity = new_quantity
                            else:
                                # New item
                                current_stock = getattr(product, stock_field, 0)
                                if current_stock < new_quantity:
                                    raise ValueError(
                                        f"Not enough stock for {product.name} in {store} store"
                                    )
                                
                                setattr(product, stock_field, current_stock - new_quantity)
                                product.save()

                            # Save the item
                            item = item_form.save(commit=False)
                            item.order = order
                            item.product = product
                            item.quantity = new_quantity
                            item.unit_price = unit_price
                            item.variance = variance
                            item.line_total = new_quantity * (unit_price + variance)
                            item.save()

                    # Recalculate order total
                    items_total = sum(
                        item.line_total for item in order.order_items.all()
                    ) or Decimal('0.00')
                    order.total_amount = items_total + (order.delivery_fee or Decimal('0.00'))
                    order.update_paid_status()
                    order.save()

                    messages.success(request, 'Order updated successfully.')
                    return redirect('order_detail', order_id=order.id)
                    
            except Exception as e:
                messages.error(request, f'Error updating order: {str(e)}')
        else:
            messages.error(request, 'Please correct the errors below.')
    else:
        form = OrderEditForm(instance=order, user=request.user)
        formset = OrderItemEditFormSet(instance=order)

    context = {
        'form': form,
        'formset': formset,
        'order': order,
        'is_admin': user_profile.is_admin()
    }
    
    return render(request, 'view/editorder.html', context)





@login_required
def delete_order(request, order_id):
    order = get_object_or_404(Order, id=order_id)
    user = request.user
    try:
        user_profile = UserProfile.objects.get(user=user)
    except UserProfile.DoesNotExist:
        return render(request, '403.html', {'error': 'User profile not found'}, status=403)

    if not user_profile.is_admin() and order.sales_person != user:
        return render(request, '403.html', {'error': 'You are not authorized to delete this order'}, status=403)

    if request.method == 'POST':
        order.delete()
        messages.success(request, f'Order #{order_id} deleted successfully.')
        return redirect('order_list')
    return render(request, 'view/deleteorder.html', {'order': order})

from openpyxl import load_workbook
@login_required
def import_products(request):
    ProductFormSet = formset_factory(ProductForm, extra=0)
    upload_error = None

    if request.method == 'POST':
        if 'upload' in request.POST and request.FILES.get('excel_file'):
            excel_file = request.FILES['excel_file']
            try:
                wb = load_workbook(excel_file, read_only=True)
                ws = wb.active
                headers = [cell.value.lower().strip().replace(' ', '_') if cell.value else '' for cell in next(ws.rows)]
                column_mappings = {
                    'name': 'name',
                    'product_name': 'name',
                    'barcode': 'barcode',
                    'category': 'category',
                    'description': 'description',
                    'factory_price': 'factory_price',
                    'distributor_price': 'distributor_price',
                    'wholesale_price': 'wholesale_price',
                    'offshore_price': 'offshore_price',
                    'retail_price': 'retail_price',
                    'mcdave_stock': 'mcdave_stock',
                    'kisii_stock': 'kisii_stock',
                    'offshore_stock': 'offshore_stock',
                    'status': 'status',
                }
                mapped_headers = [column_mappings.get(h, h) for h in headers]
                expected_columns = [
                    'name', 'barcode', 'category', 'description', 'factory_price',
                    'distributor_price', 'wholesale_price', 'offshore_price', 'retail_price',
                    'mcdave_stock', 'kisii_stock', 'offshore_stock', 'status'
                ]
                missing_columns = [col for col in expected_columns if col not in mapped_headers]
                if missing_columns:
                    upload_error = f"Missing required columns: {', '.join(missing_columns)}"
                    return render(request, 'imports/products.html', {'upload_error': upload_error})

                initial_data = []
                for row in ws.iter_rows(min_row=2):
                    row_data = {mapped_headers[i]: cell.value for i, cell in enumerate(row)}
                    category_name = row_data.get('category')
                    try:
                        category = Category.objects.get(name__iexact=category_name)
                    except Category.DoesNotExist:
                        upload_error = f"Category '{category_name}' not found."
                        return render(request, 'imports/products.html', {'upload_error': upload_error})
                    mcdave_stock = row_data.get('mcdave_stock', 0) or 0
                    kisii_stock = row_data.get('kisii_stock', 0) or 0
                    offshore_stock = row_data.get('offshore_stock', 0) or 0
                    form_data = {
                        'name': row_data.get('name'),
                        'barcode': row_data.get('barcode'),
                        'category': category.pk,
                        'description': row_data.get('description', ''),
                        'factory_price': row_data.get('factory_price'),
                        'distributor_price': row_data.get('distributor_price'),
                        'wholesale_price': row_data.get('wholesale_price'),
                        'offshore_price': row_data.get('offshore_price'),
                        'retail_price': row_data.get('retail_price'),
                        'mcdave_stock': mcdave_stock,
                        'kisii_stock': kisii_stock,
                        'offshore_stock': offshore_stock,
                        'status': row_data.get('status', 'available'),
                    }
                    initial_data.append(form_data)
                request.session['imported_products'] = initial_data
                formset = ProductFormSet(initial=initial_data)
                return render(request, 'imports/products.html', {'formset': formset})
            except Exception as e:
                upload_error = f"Error processing Excel file: {str(e)}"
                return render(request, 'imports/products.html', {'upload_error': upload_error})
        
        elif request.session.get('imported_products'):
            formset = ProductFormSet(request.POST, request.FILES)
            if formset.is_valid():
                for form in formset:
                    if form.is_valid():
                        form.save()
                del request.session['imported_products']
                messages.success(request, 'Products imported successfully!')
                return redirect('products')
            return render(request, 'imports/products.html', {'formset': formset, 'upload_error': upload_error})

    if request.session.get('imported_products'):
        formset = ProductFormSet(initial=request.session['imported_products'])
        return render(request, 'imports/products.html', {'formset': formset})
    
    return render(request, 'imports/products.html', {'upload_error': upload_error})

@login_required
def add_payment(request, order_id):
    """Add a payment to an order"""
    order = get_object_or_404(Order, id=order_id)
    
    # Authorization check
    user = request.user
    try:
        user_profile = UserProfile.objects.get(user=user)
    except UserProfile.DoesNotExist:
        return render(request, '403.html', {'error': 'User profile not found'}, status=403)
    
    if not user_profile.is_admin() and order.sales_person != user:
        return render(request, '403.html', 
                     {'error': 'You are not authorized to add payments to this order'}, 
                     status=403)
    
    if request.method == 'POST':
        form = PaymentForm(request.POST, order=order)
        if form.is_valid():
            payment = form.save(commit=False)
            payment.order = order
            payment.recorded_by = request.user
            payment.save()
            
            # Update order's amount_paid
            order.amount_paid += payment.amount
            order.update_paid_status()
            order.save()
            
            messages.success(request, f'Payment of {payment.amount:.2f} recorded successfully.')
            return redirect('order_detail', order_id=order.id)
        else:
            messages.error(request, 'Please correct the errors below.')
    else:
        form = PaymentForm(order=order)
    
    context = {
        'form': form,
        'order': order,
        'remaining_balance': order.get_balance()
    }
    
    return render(request, 'addrecords/payment.html', context)

@login_required
def delete_payment(request, payment_id):
    """Delete a payment"""
    payment = get_object_or_404(Payment, id=payment_id)
    order = payment.order
    
    # Authorization check
    user = request.user
    try:
        user_profile = UserProfile.objects.get(user=user)
    except UserProfile.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'User profile not found'}, status=403)
    
    if not user_profile.is_admin():
        return JsonResponse({'success': False, 'error': 'Only admins can delete payments'}, 
                          status=403)
    
    if request.method == 'POST':
        try:
            with transaction.atomic():
                # Reduce order's amount_paid
                order.amount_paid -= payment.amount
                order.update_paid_status()
                order.save()
                
                # Delete payment
                payment.delete()
                
                messages.success(request, 'Payment deleted successfully.')
                return JsonResponse({'success': True})
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})
    
    return JsonResponse({'success': False, 'error': 'Invalid request method'})



@login_required
def productsline(request):
    user = request.user
    try:
        user_profile = UserProfile.objects.get(user=user)
    except UserProfile.DoesNotExist:
        return render(request, '403.html', {'error': 'User profile not found'}, status=403)

    products = Product.objects.annotate(
        sales=Sum('orderitem__quantity'),
        revenue=Sum('orderitem__line_total')
    ).order_by('category__name')  # Changed from 'name' to 'category__name'

    for product in products:
        product.sales = product.sales or 0
        product.revenue = product.revenue or Decimal('0.00')
        # Compute total stock for display
        product.total_stock = product.mcdave_stock + product.kisii_stock + product.offshore_stock

    context = {
        'products': products,
        'is_admin': user_profile.is_admin(),
    }
    return render(request, 'dashboard/products.html', context)
# views.py


@login_required
def add_product(request):
    user = request.user
    try:
        user_profile = UserProfile.objects.get(user=user)
    except UserProfile:
        return render(request, '403.html', {'error': 'User profile not found'}, status=403)

    if not user_profile.is_admin():
        return render(request, '403.html', {'error': 'You are not authorized to add products'}, status=403)

    if request.method == 'POST':
        form = ProductForm(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            messages.success(request, 'Product added successfully!')
            return redirect('products')
        else:
            messages.error(request, 'Please correct the errors below.')
    else:
        form = ProductForm()

    return render(request, 'addrecords/products.html', {'form': form})
@login_required
def product_detail(request, pk):
    user = request.user
    try:
        user_profile = UserProfile.objects.get(user=user)
    except UserProfile.DoesNotExist:
        return render(request, '403.html', {'error': 'User profile not found'}, status=403)

    product = get_object_or_404(Product, pk=pk)
    context = {
        'product': product,
        'is_admin': user_profile.is_admin()  # Pass to template to hide edit buttons
    }
    return render(request, 'view/product.html', context)
# views.py
@login_required
def product_update(request, pk):
    user = request.user
    try:
        user_profile = UserProfile.objects.get(user=user)
    except UserProfile.DoesNotExist:
        return render(request, '403.html', {'error': 'User profile not found'}, status=403)

    if not user_profile.is_admin():
        return render(request, '403.html', {'error': 'You are not authorized to update products'}, status=403)

    product = get_object_or_404(Product, pk=pk)
    if request.method == 'POST':
        form = ProductForm(request.POST, request.FILES, instance=product)
        if form.is_valid():
            form.save()
            return redirect('products')
    else:
        form = ProductForm(instance=product)
    return render(request, 'view/editproduct.html', {'form': form, 'product': product})
# views.py
@login_required
def delete_product(request, pk):
    user = request.user
    try:
        user_profile = UserProfile.objects.get(user=user)
    except UserProfile.DoesNotExist:
        return render(request, '403.html', {'error': 'User profile not found'}, status=403)

    if not user_profile.is_admin():
        return render(request, '403.html', {'error': 'You are not authorized to delete products'}, status=403)

    product = get_object_or_404(Product, pk=pk)
    if request.method == 'POST':
        product.delete()
        return redirect('products')
    return render(request, 'view/deleteproduct.html', {'product': product})
# views.py
@login_required
def add_category(request):
    user = request.user
    try:
        user_profile = UserProfile.objects.get(user=user)
    except UserProfile.DoesNotExist:
        return render(request, '403.html', {'error': 'User profile not found'}, status=403)

    if not user_profile.is_admin():
        return render(request, '403.html', {'error': 'You are not authorized to add categories'}, status=403)

    if request.method == 'POST':
        form = CategoryForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('products')
    else:
        form = CategoryForm()
    return render(request, 'addrecords/category.html', {'form': form})
@login_required
def customers_list(request):
    user = request.user
    try:
        user_profile = UserProfile.objects.get(user=user)
    except UserProfile.DoesNotExist:
        return render(request, '403.html', {'error': 'User profile not found'}, status=403)

    if user_profile.is_admin():
        customers = Customer.objects.all().order_by('first_name')
    else:
        # Salesperson sees their own customers plus universal customers (added by admin, where sales_person is None)
        customers = (Customer.objects.filter(sales_person=user) | Customer.objects.filter(sales_person__isnull=True)).order_by('first_name')
    
    return render(request, 'dashboard/customers.html', {'customers': customers})

@login_required
def add_customer(request):
    if request.method == 'POST':
        form = CustomerForm(request.POST)
        if form.is_valid():
            form.save(user=request.user)  # Pass the logged-in user to the form's save method
            messages.success(request, 'Customer added successfully!')
            return redirect('customers_list')
        else:
            messages.error(request, 'Please correct the errors below.')
    else:
        form = CustomerForm()
    return render(request, 'addrecords/customers.html', {'form': form})

@login_required
def delete_customer(request, pk):
    customer = get_object_or_404(Customer, pk=pk)
    user = request.user
    try:
        user_profile = UserProfile.objects.get(user=user)
    except UserProfile.DoesNotExist:
        return render(request, '403.html', {'error': 'User profile not found'}, status=403)

    # Admins can delete any customer, salespersons can only delete their own customers (not universal customers)
    if not user_profile.is_admin() and customer.sales_person != user:
        return render(request, '403.html', {'error': 'You are not authorized to delete this customer'}, status=403)

    if request.method == 'POST':
        customer.delete()
        return redirect('customers_list')
    return render(request, 'view/deletecustomer.html', {'customer': customer})




# ============================================================
# CUSTOMER DETAIL VIEW — drop-in replacement
# ============================================================
@login_required
def customer_detail_view(request, pk):
    customer = get_object_or_404(Customer, pk=pk)
    user = request.user

    try:
        user_profile = UserProfile.objects.get(user=user)
    except UserProfile.DoesNotExist:
        return render(request, '403.html', {'error': 'User profile not found'}, status=403)

    # Access control: admins can view any, salespersons only their own
    if not user_profile.is_admin() and customer.sales_person != user and customer.sales_person is not None:
        return render(request, '403.html', {'error': 'You are not authorized to view this customer'}, status=403)

    # ── Orders ──────────────────────────────────────────────
    orders = (
        customer.orders
        .select_related('sales_person')
        .prefetch_related('order_items')
        .order_by('-order_date')
    )

    # ── Financial totals ────────────────────────────────────
    ZERO = Decimal('0.00')
    agg = orders.aggregate(
        total_spent=Coalesce(Sum('total_amount'), ZERO, output_field=DecimalField()),
        total_paid=Coalesce(Sum('amount_paid'),   ZERO, output_field=DecimalField()),
    )
    total_spent   = agg['total_spent']
    total_paid    = agg['total_paid']
    balance       = total_spent - total_paid
    total_orders  = orders.count()
    paid_pct      = (total_paid / total_spent * 100) if total_spent else 0

    # ── All payments across all orders ──────────────────────
    from store.models import Payment  # adjust app name if needed
    payments = (
        Payment.objects
        .filter(order__customer=customer)
        .select_related('order', 'recorded_by')
        .order_by('-payment_date')
    )
    total_payment_count = payments.count()

    # ── Top products bought by this customer ────────────────
    from store.models import OrderItem  # adjust app name if needed
    top_products = (
        OrderItem.objects
        .filter(order__customer=customer)
        .values('product__name')
        .annotate(
            times_ordered=Count('id'),
            total_qty=Sum('quantity'),
            total_value=Coalesce(Sum('line_total'), ZERO, output_field=DecimalField()),
        )
        .order_by('-total_value')[:10]
    )

    return render(request, 'view/customer.html', {
        'customer':            customer,
        'orders':              orders,
        'payments':            payments,
        'top_products':        top_products,
        'total_orders':        total_orders,
        'total_spent':         total_spent,
        'total_paid':          total_paid,
        'balance':             balance,
        'paid_pct':            paid_pct,
        'total_payment_count': total_payment_count,
        'is_admin':            user_profile.is_admin(),
    })

@login_required
def edit_customer(request, pk):
    customer = get_object_or_404(Customer, pk=pk)
    user = request.user
    try:
        user_profile = UserProfile.objects.get(user=user)
    except UserProfile.DoesNotExist:
        return render(request, '403.html', {'error': 'User profile not found'}, status=403)

    # Admins can edit any customer, salespersons can only edit their own customers (not universal customers)
    if not user_profile.is_admin() and customer.sales_person != user:
        return render(request, '403.html', {'error': 'You are not authorized to edit this customer'}, status=403)

    form = CustomerForm(request.POST or None, instance=customer)
    if form.is_valid():
        form.save()
        return redirect('customers_list')
    return render(request, 'view/editcustomer.html', {'form': form, 'title': 'Edit Customer'})


from collections import defaultdict
from datetime import datetime as dt, timedelta
from django import forms
from django.shortcuts import render
from django.http import HttpResponse, FileResponse
from django.contrib.auth.decorators import login_required, user_passes_test
from django.db.models import Q, Sum, Count, Max, F
from django.utils import timezone
from io import BytesIO
import os
import json
from django.conf import settings
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib.units import inch, cm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak, Flowable
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors
from reportlab.lib.colors import HexColor
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
from reportlab.graphics.shapes import Drawing, Rect, Line
from reportlab.graphics.charts.barcharts import VerticalBarChart
from reportlab.graphics.charts.piecharts import Pie
from reportlab.graphics import renderPDF
from django.contrib.auth.models import User
from .models import Order, UserProfile, Customer, Product, OrderItem
import csv
import io
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment

@login_required
def analytics_view(request):
    user = request.user
    try:
        user_profile = UserProfile.objects.get(user=user)
    except UserProfile.DoesNotExist:
        return render(request, '403.html', {'error': 'User profile not found'}, status=403)

    end_date = timezone.now()
    start_date = end_date - timedelta(days=365)

    # ── Filter parameters ─────────────────────────────────────────────
    date_from_str = request.GET.get('date_from', '').strip()
    date_to_str   = request.GET.get('date_to', '').strip()
    store_filter  = request.GET.get('store', '').strip()
    sp_filter_id  = request.GET.get('salesperson', '').strip()

    if date_from_str:
        try:
            start_date = timezone.make_aware(datetime.datetime.strptime(date_from_str, '%Y-%m-%d'))
        except ValueError:
            pass
    if date_to_str:
        try:
            end_date = timezone.make_aware(
                datetime.datetime.strptime(date_to_str, '%Y-%m-%d').replace(hour=23, minute=59, second=59)
            )
        except ValueError:
            pass

    all_salespersons_qs = User.objects.filter(groups__name='Salespersons') if user_profile.is_admin() else None

    if user_profile.is_admin():
        orders = Order.objects.filter(order_date__range=[start_date, end_date], paid_status='completed')
        order_items = OrderItem.objects.filter(
            order__paid_status='completed', order__order_date__range=[start_date, end_date]
        )
        customers = Customer.objects.all()
        payments = Payment.objects.filter(order__paid_status='completed')
        salespersons = User.objects.filter(groups__name='Salespersons')

        if store_filter:
            orders      = orders.filter(store=store_filter)
            order_items = order_items.filter(order__store=store_filter)
            payments    = payments.filter(order__store=store_filter)
        if sp_filter_id:
            try:
                _sp_id      = int(sp_filter_id)
                orders      = orders.filter(sales_person_id=_sp_id)
                order_items = order_items.filter(order__sales_person_id=_sp_id)
                customers   = customers.filter(sales_person_id=_sp_id)
                payments    = payments.filter(order__sales_person_id=_sp_id)
                salespersons = salespersons.filter(id=_sp_id)
            except (ValueError, TypeError):
                pass
    else:
        orders = Order.objects.filter(
            sales_person=user, order_date__range=[start_date, end_date], paid_status='completed'
        )
        order_items = OrderItem.objects.filter(
            order__sales_person=user, order__paid_status='completed',
            order__order_date__range=[start_date, end_date]
        )
        customers   = Customer.objects.filter(sales_person=user)
        payments    = Payment.objects.filter(order__sales_person=user, order__paid_status='completed')
        salespersons = User.objects.filter(id=user.id)

        if store_filter:
            orders      = orders.filter(store=store_filter)
            order_items = order_items.filter(order__store=store_filter)
            payments    = payments.filter(order__store=store_filter)

    # Sales data: Group by month using Python
    sales_dict = defaultdict(float)
    for order in orders:
        month_key = order.order_date.strftime('%b %Y')
        sales_dict[month_key] += float(order.total_amount)
    
    sales_data = [{'month': k, 'total': v} for k, v in sorted(sales_dict.items(), key=lambda x: dt.strptime(x[0], '%b %Y'))]
    sales_labels = [d['month'] for d in sales_data]
    sales_values = [d['total'] for d in sales_data]

    category_sales = (
        orders.values('customer_category')
        .annotate(total=Sum('total_amount'))
    )
    category_labels = [d['customer_category'] for d in category_sales]
    category_values = [float(d['total']) for d in category_sales]

    top_products = (
        order_items.values('product__name')
        .annotate(total_quantity=Sum('quantity'))
        .order_by('-total_quantity')[:5]
    )
    product_labels = [p['product__name'] for p in top_products]
    product_values = [p['total_quantity'] for p in top_products]

    stock_data = (
        Product.objects.values('category__name')
        .annotate(total_stock=Sum(F('mcdave_stock') + F('kisii_stock') + F('offshore_stock')))
    )
    stock_labels = [s['category__name'] for s in stock_data]
    stock_values = [s['total_stock'] or 0 for s in stock_data]

    # Customer Activity Tracker
    thirty_days_ago = end_date - timedelta(days=30)
    active_customers = customers.filter(
        orders__order_date__gte=thirty_days_ago,
        orders__paid_status='completed'
    ).distinct().count()
    
    inactive_customers = customers.exclude(
        orders__order_date__gte=thirty_days_ago,
        orders__paid_status='completed'
    ).distinct().count()
    
    total_customers_count = customers.count()

    # Customer activity by category
    customer_activity_by_category = []
    for category in Customer.CATEGORY_CHOICES:
        cat_name = category[0]
        cat_customers = customers.filter(default_category=cat_name)
        active_in_cat = cat_customers.filter(
            orders__order_date__gte=thirty_days_ago,
            orders__paid_status='completed'
        ).distinct().count()
        inactive_in_cat = cat_customers.exclude(
            orders__order_date__gte=thirty_days_ago,
            orders__paid_status='completed'
        ).distinct().count()
        customer_activity_by_category.append({
            'category': cat_name,
            'active': active_in_cat,
            'inactive': inactive_in_cat
        })

    # Repeat Order Frequency
    repeat_orders = Order.objects.filter(
        customer__in=customers,
        paid_status='completed'
    ).values('customer').annotate(
        order_count=Count('id'),
        last_order=Max('order_date'),
        total_amount=Sum('total_amount')
    ).filter(order_count__gt=1).order_by('-order_count')[:10]

    # Average days between orders for repeat customers
    repeat_customers_data = []
    for cust_data in repeat_orders:
        customer_orders = Order.objects.filter(
            customer_id=cust_data['customer'],
            paid_status='completed'
        ).order_by('order_date')
        
        if customer_orders.count() > 1:
            dates = list(customer_orders.values_list('order_date', flat=True))
            intervals = [(dates[i+1] - dates[i]).days for i in range(len(dates)-1)]
            avg_interval = sum(intervals) / len(intervals) if intervals else 0
            repeat_customers_data.append({
                'customer': Customer.objects.get(id=cust_data['customer']).get_full_name(),
                'order_count': cust_data['order_count'],
                'total_amount': cust_data['total_amount'],
                'avg_days_between_orders': round(avg_interval, 1),
                'last_order': cust_data['last_order']
            })

    # Wholesalers Reorder Frequency
    wholesalers = customers.filter(default_category='wholesale')
    wholesaler_reorders = Order.objects.filter(
        customer__in=wholesalers,
        paid_status='completed'
    ).values('customer').annotate(
        order_count=Count('id'),
        last_order=Max('order_date'),
        total_amount=Sum('total_amount')
    ).filter(order_count__gt=1).order_by('-order_count')[:10]

    wholesaler_data = []
    for cust_data in wholesaler_reorders:
        customer_orders = Order.objects.filter(
            customer_id=cust_data['customer'],
            paid_status='completed'
        ).order_by('order_date')
        
        if customer_orders.count() > 1:
            dates = list(customer_orders.values_list('order_date', flat=True))
            intervals = [(dates[i+1] - dates[i]).days for i in range(len(dates)-1)]
            avg_interval = sum(intervals) / len(intervals) if intervals else 0
            wholesaler_data.append({
                'customer': Customer.objects.get(id=cust_data['customer']).get_full_name(),
                'order_count': cust_data['order_count'],
                'total_amount': cust_data['total_amount'],
                'avg_days_between_orders': round(avg_interval, 1),
                'last_order': cust_data['last_order']
            })

    # New Outlet Coverage — DAILY: new customers created TODAY per salesperson
    today = end_date.date()
    new_customers_today = customers.filter(created_at__date=today).count()
    new_customers_per_salesperson = []
    for sp in salespersons:
        new_cust_count = Customer.objects.filter(
            sales_person=sp,
            created_at__date=today
        ).count()
        new_customers_per_salesperson.append({
            'salesperson': sp.get_full_name(),
            'new_customers': new_cust_count,
            'target': 5,  # Target: 5 new clients per salesperson per day
            'achievement': min(new_cust_count / 5 * 100, 100)
        })

    # Daily Orders — BOTH pending and completed (excludes cancelled)
    daily_orders_per_salesperson = []
    for sp in salespersons:
        daily_qs = Order.objects.filter(
            sales_person=sp,
            order_date__date=today
        ).exclude(paid_status='cancelled')
        if store_filter:
            daily_qs = daily_qs.filter(store=store_filter)
        daily_total = daily_qs.aggregate(total=Sum('total_amount'))['total'] or 0
        daily_orders_per_salesperson.append({
            'salesperson': sp.get_full_name(),
            'daily_total': daily_total,
            'target': 50000,
            'achievement': min(daily_total / 50000 * 100, 100)
        })

    # Monthly Payments — completed orders only, last 30 days
    payments_per_salesperson = []
    for sp in salespersons:
        monthly_payments = Payment.objects.filter(
            order__sales_person=sp,
            order__paid_status='completed',
            payment_date__gte=thirty_days_ago
        ).aggregate(total=Sum('amount'))['total'] or 0
        payments_per_salesperson.append({
            'salesperson': sp.get_full_name(),
            'monthly_payments': monthly_payments,
            'target': 200000,
            'achievement': min(monthly_payments / 200000 * 100, 100)
        })

    context = {
        'sales_data': json.dumps({'labels': sales_labels, 'values': sales_values}),
        'category_data': json.dumps({'labels': category_labels, 'values': category_values}),
        'product_data': json.dumps({'labels': product_labels, 'values': product_values}),
        'stock_data': json.dumps({'labels': stock_labels, 'values': stock_values}),
        # Customer Tracker
        'active_customers': active_customers,
        'inactive_customers': inactive_customers,
        'total_customers': total_customers_count,
        'customer_activity_by_category': customer_activity_by_category,
        # Repeat Orders
        'repeat_customers_data': repeat_customers_data,
        'wholesaler_data': wholesaler_data,
        # Targets
        'new_customers_today': new_customers_today,
        'new_customers_per_salesperson': new_customers_per_salesperson,
        'daily_orders_per_salesperson': daily_orders_per_salesperson,
        'payments_per_salesperson': payments_per_salesperson,
        # Filters
        'filter_date_from': date_from_str or start_date.strftime('%Y-%m-%d'),
        'filter_date_to':   date_to_str   or end_date.strftime('%Y-%m-%d'),
        'filter_store':     store_filter,
        'filter_salesperson': sp_filter_id,
        'all_salespersons': all_salespersons_qs,
        'store_choices':    Order.STORE_CHOICES,
    }

    # Handle downloads
    download_format = request.GET.get('download')
    if download_format:
        if download_format == 'csv':
            return generate_analytics_csv(context, start_date, end_date)
        elif download_format == 'excel':
            return generate_analytics_excel(context, start_date, end_date)
        elif download_format == 'pdf':
            return generate_analytics_pdf(context, start_date, end_date)

    return render(request, 'reports/analytics.html', context)


def generate_analytics_csv(context, start_date, end_date):
    """Generate CSV download for analytics data"""
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="analytics_dashboard_{start_date.strftime("%Y%m%d")}_to_{end_date.strftime("%Y%m%d")}.csv"'

    writer = csv.writer(response)

    # Customer Activity Section
    writer.writerow(['Customer Activity Tracker'])
    writer.writerow(['Metric', 'Value'])
    writer.writerow(['Active Customers (Last 30 days)', context['active_customers']])
    writer.writerow(['Inactive Customers', context['inactive_customers']])
    writer.writerow(['Total Customers', context['total_customers']])
    writer.writerow([])

    # Customer Activity by Category
    writer.writerow(['Customer Activity by Category'])
    writer.writerow(['Category', 'Active', 'Inactive', 'Total'])
    for cat in context['customer_activity_by_category']:
        writer.writerow([cat['category'], cat['active'], cat['inactive'], cat['active'] + cat['inactive']])
    writer.writerow([])

    # Repeat Customers
    writer.writerow(['Top Repeat Customers'])
    writer.writerow(['Customer', 'Orders', 'Avg Days Between Orders', 'Total Amount', 'Last Order'])
    for cust in context['repeat_customers_data']:
        writer.writerow([
            cust['customer'],
            cust['order_count'],
            cust['avg_days_between_orders'],
            f"KSh {cust['total_amount']}",
            cust['last_order'].strftime('%Y-%m-%d')
        ])
    writer.writerow([])

    # Wholesaler Data
    writer.writerow(['Wholesaler Reorder Frequency'])
    writer.writerow(['Customer', 'Orders', 'Avg Days Between Orders', 'Total Amount', 'Last Order'])
    for cust in context['wholesaler_data']:
        writer.writerow([
            cust['customer'],
            cust['order_count'],
            cust['avg_days_between_orders'],
            f"KSh {cust['total_amount']}",
            cust['last_order'].strftime('%Y-%m-%d')
        ])
    writer.writerow([])

    # New Customers per Salesperson
    writer.writerow(['New Outlet Coverage (Target: 5 new clients per salesperson per day — today only)'])
    writer.writerow(['Salesperson', 'New Clients', 'Target', 'Achievement %'])
    for sp in context['new_customers_per_salesperson']:
        writer.writerow([sp['salesperson'], sp['new_customers'], sp['target'], f"{sp['achievement']:.1f}%"])
    writer.writerow([])

    # Daily Orders per Salesperson
    writer.writerow(['Daily Orders Target (Target: KSh 50,000 per salesperson per day — pending + completed)'])
    writer.writerow(['Salesperson', 'Today\'s Total', 'Target', 'Achievement %'])
    for sp in context['daily_orders_per_salesperson']:
        writer.writerow([sp['salesperson'], f"KSh {sp['daily_total']}", f"KSh {sp['target']}", f"{sp['achievement']:.1f}%"])
    writer.writerow([])

    # Monthly Payments per Salesperson
    writer.writerow(['Monthly Payments Target (Target: KSh 200,000 collected per salesperson)'])
    writer.writerow(['Salesperson', 'Collected This Month', 'Target', 'Achievement %'])
    for sp in context['payments_per_salesperson']:
        writer.writerow([sp['salesperson'], f"KSh {sp['monthly_payments']}", f"KSh {sp['target']}", f"{sp['achievement']:.1f}%"])

    return response


def generate_analytics_excel(context, start_date, end_date):
    """Generate Excel download for analytics data"""
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="analytics_dashboard_{start_date.strftime("%Y%m%d")}_to_{end_date.strftime("%Y%m%d")}.xlsx"'

    wb = Workbook()
    ws = wb.active
    ws.title = "Analytics Dashboard"

    # Customer Activity Section
    ws.append(['Customer Activity Tracker'])
    ws.append(['Metric', 'Value'])
    ws.append(['Active Customers (Last 30 days)', context['active_customers']])
    ws.append(['Inactive Customers', context['inactive_customers']])
    ws.append(['Total Customers', context['total_customers']])
    ws.append([])

    # Customer Activity by Category
    ws.append(['Customer Activity by Category'])
    ws.append(['Category', 'Active', 'Inactive', 'Total'])
    for cat in context['customer_activity_by_category']:
        ws.append([cat['category'], cat['active'], cat['inactive'], cat['active'] + cat['inactive']])
    ws.append([])

    # Repeat Customers
    ws.append(['Top Repeat Customers'])
    ws.append(['Customer', 'Orders', 'Avg Days Between Orders', 'Total Amount', 'Last Order'])
    for cust in context['repeat_customers_data']:
        ws.append([
            cust['customer'],
            cust['order_count'],
            cust['avg_days_between_orders'],
            f"KSh {cust['total_amount']}",
            cust['last_order'].strftime('%Y-%m-%d')
        ])
    ws.append([])

    # Wholesaler Data
    ws.append(['Wholesaler Reorder Frequency'])
    ws.append(['Customer', 'Orders', 'Avg Days Between Orders', 'Total Amount', 'Last Order'])
    for cust in context['wholesaler_data']:
        ws.append([
            cust['customer'],
            cust['order_count'],
            cust['avg_days_between_orders'],
            f"KSh {cust['total_amount']}",
            cust['last_order'].strftime('%Y-%m-%d')
        ])
    ws.append([])

    # New Customers per Salesperson
    ws.append(['New Outlet Coverage (Target: 5 new clients per salesperson per day — today only)'])
    ws.append(['Salesperson', 'New Clients', 'Target', 'Achievement %'])
    for sp in context['new_customers_per_salesperson']:
        ws.append([sp['salesperson'], sp['new_customers'], sp['target'], f"{sp['achievement']:.1f}%"])
    ws.append([])

    # Daily Orders per Salesperson
    ws.append(['Daily Orders Target (Target: KSh 50,000 per salesperson per day — pending + completed)'])
    ws.append(['Salesperson', 'Today\'s Total', 'Target', 'Achievement %'])
    for sp in context['daily_orders_per_salesperson']:
        ws.append([sp['salesperson'], f"KSh {sp['daily_total']}", f"KSh {sp['target']}", f"{sp['achievement']:.1f}%"])
    ws.append([])

    # Monthly Payments per Salesperson
    ws.append(['Monthly Payments Target (Target: KSh 200,000 collected per salesperson)'])
    ws.append(['Salesperson', 'Collected This Month', 'Target', 'Achievement %'])
    for sp in context['payments_per_salesperson']:
        ws.append([sp['salesperson'], f"KSh {sp['monthly_payments']}", f"KSh {sp['target']}", f"{sp['achievement']:.1f}%"])

    # Apply formatting
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.font = Font(name='Helvetica', size=10)
            cell.alignment = Alignment(horizontal='left', vertical='center')

    # Auto-adjust column widths
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width

    wb.save(response)
    return response


# --------------------------------------------------
# Stock export helper functions
# --------------------------------------------------

def generate_stock_csv(products):
    """Return HttpResponse with CSV file of product stock levels."""
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="stock_levels.csv"'
    writer = csv.writer(response)
    writer.writerow(['Product', 'McDave', 'Mombasa', 'Offshore', 'Total'])
    for p in products:
        total = (p.mcdave_stock or 0) + (p.kisii_stock or 0) + (p.offshore_stock or 0)
        writer.writerow([p.name, p.mcdave_stock, p.kisii_stock, p.offshore_stock, total])
    return response


def generate_stock_excel(products):
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="stock_levels.xlsx"'
    wb = Workbook()
    ws = wb.active
    ws.title = 'Stock Levels'
    ws.append(['Product', 'McDave', 'Mombasa', 'Offshore', 'Total'])
    for p in products:
        total = (p.mcdave_stock or 0) + (p.kisii_stock or 0) + (p.offshore_stock or 0)
        ws.append([p.name, p.mcdave_stock, p.kisii_stock, p.offshore_stock, total])
    # formatting
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=5):
        for cell in row:
            cell.font = Font(name='Helvetica', size=10)
            cell.alignment = Alignment(horizontal='left', vertical='center')
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        ws.column_dimensions[column].width = max_length + 2
    wb.save(response)
    return response


def generate_stock_pdf(products):
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="stock_levels.pdf"'
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter, leftMargin=40, rightMargin=40, topMargin=40, bottomMargin=40)
    styles = getSampleStyleSheet()
    elements = []
    elements.append(Paragraph('Stock Levels Report', styles['Title']))
    elements.append(Spacer(1, 12))
    data = [['Product', 'McDave', 'Mombasa', 'Offshore', 'Total']]
    for p in products:
        total = (p.mcdave_stock or 0) + (p.kisii_stock or 0) + (p.offshore_stock or 0)
        data.append([p.name, p.mcdave_stock, p.kisii_stock, p.offshore_stock, total])
    table = Table(data, colWidths=[180, 60, 60, 60, 60])
    table.setStyle(TableStyle([
        ('GRID', (0,0), (-1,-1), 0.5, colors.grey),
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#1e3a8a')),
        ('TEXTCOLOR', (0,0), (-1,0), colors.white),
        ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
    ]))
    elements.append(table)
    doc.build(elements)
    pdf = buffer.getvalue()
    buffer.close()
    response.write(pdf)
    return response


def generate_analytics_pdf(context, start_date, end_date):
    """Generate PDF download for analytics data"""
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="analytics_dashboard_{start_date.strftime("%Y%m%d")}_to_{end_date.strftime("%Y%m%d")}.pdf"'

    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter, leftMargin=50, rightMargin=50, topMargin=50, bottomMargin=50)
    styles = getSampleStyleSheet()
    elements = []

    # Title
    title_style = ParagraphStyle('CustomTitle', parent=styles['Title'], fontName='Helvetica-Bold', fontSize=18,
                                textColor=colors.white, backColor=colors.HexColor('#1e3a8a'), alignment=1,
                                spaceAfter=12, spaceBefore=12, leading=22)
    elements.append(Paragraph("Analytics Dashboard Report", title_style))
    elements.append(Spacer(1, 12))

    # Date range
    date_style = ParagraphStyle('DateStyle', parent=styles['Normal'], fontName='Helvetica', fontSize=12,
                               textColor=colors.grey, alignment=1)
    elements.append(Paragraph(f"Report Period: {start_date.strftime('%Y-%m-%d')} to {end_date.strftime('%Y-%m-%d')}", date_style))
    elements.append(Spacer(1, 20))

    # Customer Activity Section
    elements.append(Paragraph("Customer Activity Tracker", styles['Heading2']))
    elements.append(Spacer(1, 10))

    customer_data = [
        ['Metric', 'Value'],
        ['Active Customers (Last 30 days)', str(context['active_customers'])],
        ['Inactive Customers', str(context['inactive_customers'])],
        ['Total Customers', str(context['total_customers'])]
    ]
    customer_table = Table(customer_data, colWidths=[300, 150])
    customer_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1e3a8a')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 11),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.lightgrey),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('LEFTPADDING', (0, 0), (-1, -1), 8),
        ('RIGHTPADDING', (0, 0), (-1, -1), 8),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
    ]))
    elements.append(customer_table)
    elements.append(Spacer(1, 20))

    # Customer Activity by Category
    elements.append(Paragraph("Customer Activity by Category", styles['Heading3']))
    elements.append(Spacer(1, 10))

    category_data = [['Category', 'Active', 'Inactive', 'Total']]
    for cat in context['customer_activity_by_category']:
        category_data.append([cat['category'], str(cat['active']), str(cat['inactive']), str(cat['active'] + cat['inactive'])])

    category_table = Table(category_data, colWidths=[150, 80, 80, 80])
    category_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#3b82f6')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.lightgrey),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('LEFTPADDING', (0, 0), (-1, -1), 8),
        ('RIGHTPADDING', (0, 0), (-1, -1), 8),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
    ]))
    elements.append(category_table)
    elements.append(Spacer(1, 20))

    # Repeat Customers
    elements.append(Paragraph("Top Repeat Customers", styles['Heading3']))
    elements.append(Spacer(1, 10))

    repeat_data = [['Customer', 'Orders', 'Avg Days', 'Total Amount']]
    for cust in context['repeat_customers_data'][:10]:  # Limit to 10 for PDF
        repeat_data.append([
            cust['customer'],
            str(cust['order_count']),
            f"{cust['avg_days_between_orders']:.1f}",
            f"KSh {cust['total_amount']:,}"
        ])

    repeat_table = Table(repeat_data, colWidths=[150, 60, 70, 100])
    repeat_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#10b981')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.lightgrey),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('LEFTPADDING', (0, 0), (-1, -1), 8),
        ('RIGHTPADDING', (0, 0), (-1, -1), 8),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
    ]))
    elements.append(repeat_table)
    elements.append(Spacer(1, 20))

    # Salesperson Targets
    elements.append(Paragraph("Salesperson Performance Targets", styles['Heading2']))
    elements.append(Spacer(1, 10))

    # New Outlet Coverage
    elements.append(Paragraph("New Outlet Coverage (Target: 5 new clients per salesperson per day — today only)", styles['Heading4']))
    elements.append(Spacer(1, 5))

    outlet_data = [['Salesperson', 'New Clients', 'Target', 'Achievement']]
    for sp in context['new_customers_per_salesperson']:
        outlet_data.append([sp['salesperson'], str(sp['new_customers']), str(sp['target']), f"{sp['achievement']:.1f}%"])

    outlet_table = Table(outlet_data, colWidths=[150, 80, 60, 80])
    outlet_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#f59e0b')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.lightgrey),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('LEFTPADDING', (0, 0), (-1, -1), 8),
        ('RIGHTPADDING', (0, 0), (-1, -1), 8),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
    ]))
    elements.append(outlet_table)
    elements.append(Spacer(1, 15))

    # Daily Orders Target
    elements.append(Paragraph("Daily Orders Target (Target: KSh 50,000 per salesperson per day — pending + completed)", styles['Heading4']))
    elements.append(Spacer(1, 5))

    daily_data = [['Salesperson', 'Today\'s Total', 'Target', 'Achievement']]
    for sp in context['daily_orders_per_salesperson']:
        daily_data.append([sp['salesperson'], f"KSh {sp['daily_total']:,}", f"KSh {sp['target']:,}", f"{sp['achievement']:.1f}%"])

    daily_table = Table(daily_data, colWidths=[150, 100, 80, 80])
    daily_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#ef4444')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.lightgrey),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('LEFTPADDING', (0, 0), (-1, -1), 8),
        ('RIGHTPADDING', (0, 0), (-1, -1), 8),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
    ]))
    elements.append(daily_table)
    elements.append(Spacer(1, 15))

    # Monthly Payments Target
    elements.append(Paragraph("Monthly Payments Target (Target: KSh 200,000 collected per salesperson)", styles['Heading4']))
    elements.append(Spacer(1, 5))

    payment_data = [['Salesperson', 'Collected This Month', 'Target', 'Achievement']]
    for sp in context['payments_per_salesperson']:
        payment_data.append([sp['salesperson'], f"KSh {sp['monthly_payments']:,}", f"KSh {sp['target']:,}", f"{sp['achievement']:.1f}%"])

    payment_table = Table(payment_data, colWidths=[150, 120, 80, 80])
    payment_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#8b5cf6')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.lightgrey),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('LEFTPADDING', (0, 0), (-1, -1), 8),
        ('RIGHTPADDING', (0, 0), (-1, -1), 8),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
    ]))
    elements.append(payment_table)

    # Footer
    elements.append(Spacer(1, 30))
    footer_style = ParagraphStyle('Footer', parent=styles['Normal'], fontName='Helvetica-Oblique', fontSize=8,
                                 textColor=colors.grey, alignment=1)
    elements.append(Paragraph("Generated by Zelia OMS Analytics Dashboard", footer_style))
    elements.append(Paragraph(f"Report generated on {timezone.now().strftime('%Y-%m-%d %H:%M:%S')}", footer_style))

    doc.build(elements)
    pdf = buffer.getvalue()
    buffer.close()
    response.write(pdf)
    return response



# Custom Canvas for Header/Footer and Watermark
class ModernReportCanvas(canvas.Canvas):
    """Custom canvas for modern PDF styling with header, footer, and watermark"""
    
    def __init__(self, *args, **kwargs):
        canvas.Canvas.__init__(self, *args, **kwargs)
        self.pages = []
        
    def showPage(self):
        self.pages.append(dict(self.__dict__))
        self._startPage()
        
    def save(self):
        page_count = len(self.pages)
        for page_num, page in enumerate(self.pages, start=1):
            self.__dict__.update(page)
            self.draw_page_decorations(page_num, page_count)
            canvas.Canvas.showPage(self)
        canvas.Canvas.save(self)
        
    def draw_page_decorations(self, page_num, page_count):
        """Draw header, footer, and decorative elements"""
        page_width, page_height = letter
        
        # === HEADER ===
        # Header background with gradient effect (simulated with rectangles)
        self.setFillColor(HexColor('#2E86AB'))
        self.rect(0, page_height - 60, page_width, 60, fill=True, stroke=False)
        
        # Header accent line
        self.setFillColor(HexColor('#A23B72'))
        self.rect(0, page_height - 65, page_width, 5, fill=True, stroke=False)
        
        # Company logo placeholder (you can add actual logo if available)
        self.setFillColor(colors.white)
        self.setFont("Helvetica-Bold", 20)
        self.drawString(50, page_height - 40, "McDave Sales")
        
        # Header subtitle
        self.setFont("Helvetica", 10)
        self.drawString(50, page_height - 52, "Professional Sales Analytics")
        
        # === FOOTER ===
        # Footer background
        self.setFillColor(HexColor('#F8F9FA'))
        self.rect(0, 0, page_width, 50, fill=True, stroke=False)
        
        # Footer accent line
        self.setFillColor(HexColor('#2E86AB'))
        self.rect(0, 50, page_width, 2, fill=True, stroke=False)
        
        # Page number with modern styling
        self.setFillColor(HexColor('#2E86AB'))
        self.setFont("Helvetica-Bold", 9)
        page_text = f"Page {page_num} of {page_count}"
        self.drawRightString(page_width - 50, 20, page_text)
        
        # Footer left side - timestamp
        self.setFillColor(HexColor('#6C757D'))
        self.setFont("Helvetica", 8)
        timestamp = timezone.now().strftime('%B %d, %Y at %I:%M %p')
        self.drawString(50, 20, f"Generated: {timestamp}")
        
        # Footer center - confidential notice
        self.setFont("Helvetica-Oblique", 7)
        self.drawCentredString(page_width / 2, 30, "CONFIDENTIAL - For Internal Use Only")
        
        # Decorative corner elements
        self.setStrokeColor(HexColor('#A23B72'))
        self.setLineWidth(2)
        # Top left corner
        self.line(40, page_height - 70, 60, page_height - 70)
        self.line(40, page_height - 70, 40, page_height - 90)
        # Top right corner
        self.line(page_width - 40, page_height - 70, page_width - 60, page_height - 70)
        self.line(page_width - 40, page_height - 70, page_width - 40, page_height - 90)
        
        # Watermark (subtle)
        self.saveState()
        self.setFillColor(HexColor('#F0F0F0'))
        self.setFont("Helvetica-Bold", 60)
        self.translate(page_width / 2, page_height / 2)
        self.rotate(45)
        self.drawCentredString(0, 0, "McDave")
        self.restoreState()


@login_required
def sales_report_view(request):
    user = request.user
    try:
        user_profile = UserProfile.objects.get(user=user)
    except UserProfile.DoesNotExist:
        return render(request, '403.html', {'error': 'User profile not found'}, status=403)

    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    customer_category = request.GET.get('customer_category', '')
    paid_status = request.GET.get('paid_status', '')
    selected_user = request.GET.get('user', '') if user_profile.is_admin() else user.id
    selected_customer = request.GET.get('customer', '')

    # Base queryset
    if user_profile.is_admin():
        orders = Order.objects.filter(Q(paid_status='completed') | Q(paid_status='pending'))
        if selected_user:
            orders = orders.filter(sales_person_id=selected_user)
    else:
        orders = Order.objects.filter(
            Q(sales_person=user) & (Q(paid_status='completed') | Q(paid_status='pending'))
        )

    # Apply filters
    if start_date:
        try:
            start_date = dt.strptime(start_date, '%Y-%m-%d')
            start_date = timezone.make_aware(start_date)
            orders = orders.filter(order_date__gte=start_date)
        except ValueError:
            return HttpResponse("Invalid start date format. Please use YYYY-MM-DD.", status=400)

    if end_date:
        try:
            end_date = dt.strptime(end_date, '%Y-%m-%d')
            end_date = end_date.replace(hour=23, minute=59, second=59)
            end_date = timezone.make_aware(end_date)
            orders = orders.filter(order_date__lte=end_date)
        except ValueError:
            return HttpResponse("Invalid end date format. Please use YYYY-MM-DD.", status=400)

    if customer_category:
        orders = orders.filter(customer_category=customer_category)

    if paid_status:
        orders = orders.filter(paid_status=paid_status)

    if selected_customer:
        orders = orders.filter(customer_id=selected_customer)

    # Sales summary
    sales_summary = {
        'completed': orders.filter(paid_status='completed').aggregate(
            total=Sum('total_amount'), count=Count('id')
        ),
        'pending': orders.filter(paid_status='pending').aggregate(
            total=Sum('total_amount'), count=Count('id')
        ),
        'total': orders.aggregate(total=Sum('total_amount'), count=Count('id'))
    }

    # Customers summary
    customers_data = (
        orders.values(
            'customer__id',
            'customer__first_name',
            'customer__last_name',
            'customer__phone_number',
            'customer__default_category'
        )
        .annotate(
            total_spent=Sum('total_amount'),
            last_order_date=Max('order_date')
        )
        .order_by('-total_spent')
    )

    customers_summary = []
    for customer in customers_data:
        customer_orders = orders.filter(customer__id=customer['customer__id'])
        order_ids = ', '.join([f"#Mc{order.id}Z" for order in customer_orders])
        customers_summary.append({
            'customer__id': customer['customer__id'],
            'customer__first_name': customer['customer__first_name'],
            'customer__last_name': customer['customer__last_name'],
            'customer__phone_number': customer['customer__phone_number'],
            'customer__default_category': customer['customer__default_category'],
            'total_spent': customer['total_spent'],
            'last_order_date': customer['last_order_date'],
            'order_ids': order_ids
        })

    # PDF Download
    if 'download' in request.GET:
        download_type = request.GET.get('download')
        
        if download_type == 'pdf':
            return generate_beautiful_pdf(
                user=user,
                orders=orders,
                sales_summary=sales_summary,
                customers_summary=customers_summary,
                start_date=start_date,
                end_date=end_date,
                customer_category=customer_category,
                paid_status=paid_status,
                selected_customer=selected_customer
            )

    # Get all customers and users for dropdowns
    if user_profile.is_admin():
        all_customers = Customer.objects.filter(orders__isnull=False).distinct().order_by('first_name')
    else:
        all_customers = Customer.objects.filter(orders__sales_person=user).distinct().order_by('first_name')

    all_users = User.objects.filter(groups__name='Salespersons').order_by('first_name')

    context = {
        'orders': orders,
        'category_choices': Order.CATEGORY_CHOICES,
        'paid_status_choices': Order.PAID_STATUS_CHOICES,
        'sales_summary': sales_summary,
        'customers_summary': customers_summary,
        'users': all_users,
        'selected_user': selected_user,
        'all_customers': all_customers,
        'selected_customer': selected_customer,
    }
    return render(request, 'reports/sales.html', context)


def generate_beautiful_pdf(user, orders, sales_summary, customers_summary, 
                          start_date=None, end_date=None, customer_category=None, 
                          paid_status=None, selected_customer=None):
    """
    Generate a beautiful, modern PDF sales report
    """
    buffer = BytesIO()
    
    # Create document with custom canvas
    doc = SimpleDocTemplate(
        buffer,
        pagesize=letter,
        rightMargin=50,
        leftMargin=50,
        topMargin=80,
        bottomMargin=80,
        canvasmaker=ModernReportCanvas
    )
    
    # Container for PDF elements
    elements = []
    
    # Define modern styles
    styles = getSampleStyleSheet()
    
    # === CUSTOM STYLES ===
    
    # Ultra-modern title
    title_style = ParagraphStyle(
        'ModernTitle',
        parent=styles['Title'],
        fontSize=32,
        spaceAfter=10,
        alignment=TA_CENTER,
        textColor=HexColor('#2E86AB'),
        fontName='Helvetica-Bold',
        leading=38
    )
    
    # Subtitle style
    subtitle_style = ParagraphStyle(
        'Subtitle',
        parent=styles['Normal'],
        fontSize=14,
        spaceAfter=30,
        alignment=TA_CENTER,
        textColor=HexColor('#6C757D'),
        fontName='Helvetica',
        leading=18
    )
    
    # Section heading with background
    section_heading_style = ParagraphStyle(
        'SectionHeading',
        parent=styles['Heading1'],
        fontSize=18,
        spaceAfter=15,
        spaceBefore=25,
        textColor=colors.white,
        fontName='Helvetica-Bold',
        backColor=HexColor('#2E86AB'),
        borderPadding=12,
        leading=22
    )
    
    # Filter info box style
    filter_style = ParagraphStyle(
        'FilterInfo',
        parent=styles['Normal'],
        fontSize=10,
        spaceAfter=20,
        alignment=TA_LEFT,
        textColor=HexColor('#495057'),
        fontName='Helvetica',
        backColor=HexColor('#E3F2FD'),
        borderWidth=1,
        borderColor=HexColor('#2E86AB'),
        borderPadding=10,
        leftIndent=10,
        rightIndent=10
    )
    
    # Info box style
    info_box_style = ParagraphStyle(
        'InfoBox',
        parent=styles['Normal'],
        fontSize=9,
        textColor=HexColor('#6C757D'),
        fontName='Helvetica-Oblique',
        alignment=TA_CENTER,
        spaceAfter=10
    )
    
    # === TITLE PAGE ===
    
    elements.append(Spacer(1, 0.5 * inch))
    
    # Main title with icon
    elements.append(Paragraph("📊 SALES PERFORMANCE REPORT", title_style))
    
    # Subtitle
    report_for = user.get_full_name() or user.username
    elements.append(Paragraph(
        f"Comprehensive Sales Analysis for <b>{report_for}</b>", 
        subtitle_style
    ))
    
    # Report metadata box
    report_date = timezone.now().strftime('%B %d, %Y')
    report_time = timezone.now().strftime('%I:%M %p')
    
    elements.append(Paragraph(
        f"<b>Report Generated:</b> {report_date} at {report_time}",
        info_box_style
    ))
    
    elements.append(Spacer(1, 0.3 * inch))
    
    # === FILTER INFORMATION ===
    filter_info = []
    if start_date:
        filter_info.append(f"<b>Start Date:</b> {start_date.strftime('%B %d, %Y')}")
    if end_date:
        filter_info.append(f"<b>End Date:</b> {end_date.strftime('%B %d, %Y')}")
    if customer_category:
        filter_info.append(f"<b>Category:</b> {customer_category.title()}")
    if paid_status:
        filter_info.append(f"<b>Payment Status:</b> {paid_status.title()}")
    if selected_customer:
        try:
            customer_obj = Customer.objects.get(id=selected_customer)
            filter_info.append(f"<b>Customer:</b> {customer_obj.get_full_name()}")
        except Customer.DoesNotExist:
            pass
    
    if filter_info:
        filter_text = "<br/>".join(filter_info)
        elements.append(Paragraph(
            f"<b>🔍 Active Filters:</b><br/>{filter_text}",
            filter_style
        ))
    else:
        elements.append(Paragraph(
            "<b>🔍 Filters:</b> No filters applied - Showing all data",
            filter_style
        ))
    
    elements.append(Spacer(1, 0.4 * inch))
    
    # === EXECUTIVE SUMMARY ===
    elements.append(Paragraph("💼 EXECUTIVE SUMMARY", section_heading_style))
    elements.append(Spacer(1, 0.2 * inch))
    
    # Summary cards in a beautiful table
    completed_total = sales_summary['completed']['total'] or 0
    completed_count = sales_summary['completed']['count'] or 0
    pending_total = sales_summary['pending']['total'] or 0
    pending_count = sales_summary['pending']['count'] or 0
    total_revenue = sales_summary['total']['total'] or 0
    total_count = sales_summary['total']['count'] or 0
    
    # Calculate average order value
    avg_order_value = total_revenue / total_count if total_count > 0 else 0
    
    summary_data = [
        # Headers
        ['📈 PAID ORDERS', '⏳ PENDING ORDERS', '💰 TOTAL REVENUE', '📊 AVG ORDER'],
        # Values
        [
            f"{completed_count}\nKsh {completed_total:,.2f}",
            f"{pending_count}\nKsh {pending_total:,.2f}",
            f"{total_count}\nKsh {total_revenue:,.2f}",
            f"Ksh {avg_order_value:,.2f}"
        ]
    ]
    
    summary_table = Table(summary_data, colWidths=[2*inch, 2*inch, 2*inch, 1.5*inch])
    summary_table.setStyle(TableStyle([
        # Header styling
        ('BACKGROUND', (0, 0), (-1, 0), HexColor('#2E86AB')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 11),
        ('TOPPADDING', (0, 0), (-1, 0), 16),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 16),
        
        # Data styling
        ('BACKGROUND', (0, 1), (-1, 1), colors.white),
        ('FONTNAME', (0, 1), (-1, 1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 1), (-1, 1), 13),
        ('TEXTCOLOR', (0, 1), (-1, 1), HexColor('#2E86AB')),
        ('TOPPADDING', (0, 1), (-1, 1), 20),
        ('BOTTOMPADDING', (0, 1), (-1, 1), 20),
        
        # Borders
        ('BOX', (0, 0), (-1, -1), 2, HexColor('#2E86AB')),
        ('LINEBELOW', (0, 0), (-1, 0), 2, HexColor('#2E86AB')),
        ('INNERGRID', (0, 0), (-1, -1), 1, HexColor('#DEE2E6')),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        
        # Individual cell colors for visual interest
        ('BACKGROUND', (0, 1), (0, 1), HexColor('#D4EDDA')),  # Green tint for paid
        ('BACKGROUND', (1, 1), (1, 1), HexColor('#FFF3CD')),  # Yellow tint for pending
        ('BACKGROUND', (2, 1), (2, 1), HexColor('#D1ECF1')),  # Blue tint for total
        ('BACKGROUND', (3, 1), (3, 1), HexColor('#E2E3E5')),  # Gray tint for average
    ]))
    
    elements.append(summary_table)
    elements.append(Spacer(1, 0.5 * inch))
    
    # === KEY INSIGHTS ===
    insights_style = ParagraphStyle(
        'Insights',
        parent=styles['Normal'],
        fontSize=10,
        textColor=HexColor('#495057'),
        fontName='Helvetica',
        backColor=HexColor('#F8F9FA'),
        borderWidth=1,
        borderColor=HexColor('#DEE2E6'),
        borderPadding=12,
        leftIndent=15,
        bulletIndent=10,
        spaceAfter=5
    )
    
    completion_rate = (completed_total / total_revenue * 100) if total_revenue > 0 else 0
    
    elements.append(Paragraph("💡 KEY INSIGHTS", section_heading_style))
    elements.append(Spacer(1, 0.1 * inch))
    
    elements.append(Paragraph(
        f"• <b>Payment Completion Rate:</b> {completion_rate:.1f}% of total revenue has been received",
        insights_style
    ))
    elements.append(Paragraph(
        f"• <b>Outstanding Revenue:</b> Ksh {pending_total:,.2f} pending collection from {pending_count} orders",
        insights_style
    ))
    elements.append(Paragraph(
        f"• <b>Customer Base:</b> {len(customers_summary)} unique customers in this period",
        insights_style
    ))
    
    elements.append(Spacer(1, 0.4 * inch))
    
    # === TOP CUSTOMERS ===
    if customers_summary:
        elements.append(Paragraph("👥 TOP CUSTOMERS", section_heading_style))
        elements.append(Spacer(1, 0.15 * inch))
        
        # Take top 10 customers
        top_customers = customers_summary[:10]
        
        customer_table_data = [
            ['#', '👤 Customer', '📞 Phone', '🏷️ Category', '💵 Total Spent', '📦 Orders']
        ]
        
        for idx, customer in enumerate(top_customers, 1):
            name = f"{customer['customer__first_name']} {customer['customer__last_name'] or ''}".strip()
            if len(name) > 25:
                name = name[:22] + "..."
            
            phone = customer['customer__phone_number'] or 'N/A'
            if len(phone) > 15:
                phone = phone[:12] + "..."
            
            category = (customer['customer__default_category'] or 'N/A')[:12]
            
            # Count orders for this customer
            order_count = len([o for o in orders if o.customer_id == customer['customer__id']])
            
            customer_table_data.append([
                str(idx),
                name,
                phone,
                category,
                f"Ksh {customer['total_spent']:,.2f}",
                str(order_count)
            ])
        
        customer_table = Table(
            customer_table_data,
            colWidths=[0.4*inch, 1.8*inch, 1.2*inch, 1*inch, 1.3*inch, 0.7*inch]
        )
        
        # Gradient colors for ranking
        rank_colors = [
            HexColor('#FFD700'),  # Gold
            HexColor('#C0C0C0'),  # Silver
            HexColor('#CD7F32'),  # Bronze
        ]
        
        customer_table_style = [
            # Header
            ('BACKGROUND', (0, 0), (-1, 0), HexColor('#28A745')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('TOPPADDING', (0, 0), (-1, 0), 14),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 14),
            
            # Data
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 9),
            ('TOPPADDING', (0, 1), (-1, -1), 10),
            ('BOTTOMPADDING', (0, 1), (-1, -1), 10),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, HexColor('#F8F9FA')]),
            
            # Borders
            ('BOX', (0, 0), (-1, -1), 2, HexColor('#28A745')),
            ('LINEBELOW', (0, 0), (-1, 0), 2, HexColor('#28A745')),
            ('INNERGRID', (0, 0), (-1, -1), 0.5, HexColor('#DEE2E6')),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ]
        
        # Add special highlighting for top 3
        for i, color in enumerate(rank_colors[:min(3, len(top_customers))], 1):
            customer_table_style.append(('BACKGROUND', (0, i), (0, i), color))
            customer_table_style.append(('TEXTCOLOR', (0, i), (0, i), colors.black))
            customer_table_style.append(('FONTNAME', (0, i), (0, i), 'Helvetica-Bold'))
        
        customer_table.setStyle(TableStyle(customer_table_style))
        elements.append(customer_table)
        
        if len(customers_summary) > 10:
            elements.append(Spacer(1, 0.1 * inch))
            elements.append(Paragraph(
                f"<i>Showing top 10 of {len(customers_summary)} customers. "
                f"Remaining customers contributed Ksh {sum(c['total_spent'] for c in customers_summary[10:]):,.2f}</i>",
                info_box_style
            ))
        
        elements.append(PageBreak())
    
    # === DETAILED ORDER LISTING ===
    elements.append(Paragraph("📋 DETAILED ORDER HISTORY", section_heading_style))
    elements.append(Spacer(1, 0.15 * inch))
    
    # Show up to 100 orders
    orders_list = list(orders[:100])
    
    if orders_list:
        # Group orders into chunks for better pagination
        chunk_size = 25
        for chunk_idx in range(0, len(orders_list), chunk_size):
            chunk = orders_list[chunk_idx:chunk_idx + chunk_size]
            
            if chunk_idx > 0:
                elements.append(Spacer(1, 0.3 * inch))
            
            order_table_data = [
                ['🆔 Order ID', '👤 Customer', '📅 Date', '🏷️ Category', '💰 Amount', '✅ Status']
            ]
            
            for order in chunk:
                customer_name = order.customer.get_full_name() if order.customer else "Unknown"
                if len(customer_name) > 20:
                    customer_name = customer_name[:17] + "..."
                
                order_date = order.order_date.strftime('%m/%d/%Y')
                category = order.get_customer_category_display()[:12]
                amount = f"Ksh {order.total_amount:,.2f}"
                status = order.get_paid_status_display()
                
                order_table_data.append([
                    f"#Mc{order.id}Z",
                    customer_name,
                    order_date,
                    category,
                    amount,
                    status
                ])
            
            order_table = Table(
                order_table_data,
                colWidths=[1*inch, 1.5*inch, 0.9*inch, 1*inch, 1.1*inch, 1*inch]
            )
            
            order_table.setStyle(TableStyle([
                # Header
                ('BACKGROUND', (0, 0), (-1, 0), HexColor('#6610F2')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 9),
                ('TOPPADDING', (0, 0), (-1, 0), 12),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                
                # Data
                ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 1), (-1, -1), 8),
                ('TOPPADDING', (0, 1), (-1, -1), 8),
                ('BOTTOMPADDING', (0, 1), (-1, -1), 8),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, HexColor('#F8F9FA')]),
                
                # Borders
                ('BOX', (0, 0), (-1, -1), 1.5, HexColor('#6610F2')),
                ('LINEBELOW', (0, 0), (-1, 0), 2, HexColor('#6610F2')),
                ('INNERGRID', (0, 0), (-1, -1), 0.5, HexColor('#DEE2E6')),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ]))
            
            # Color code status column
            for i in range(1, len(order_table_data)):
                status = chunk[i-1].paid_status
                if status == 'completed':
                    order_table.setStyle(TableStyle([
                        ('BACKGROUND', (5, i), (5, i), HexColor('#D4EDDA')),
                        ('TEXTCOLOR', (5, i), (5, i), HexColor('#155724')),
                    ]))
                elif status == 'pending':
                    order_table.setStyle(TableStyle([
                        ('BACKGROUND', (5, i), (5, i), HexColor('#FFF3CD')),
                        ('TEXTCOLOR', (5, i), (5, i), HexColor('#856404')),
                    ]))
                elif status == 'partially_paid':
                    order_table.setStyle(TableStyle([
                        ('BACKGROUND', (5, i), (5, i), HexColor('#D1ECF1')),
                        ('TEXTCOLOR', (5, i), (5, i), HexColor('#0C5460')),
                    ]))
            
            elements.append(order_table)
        
        if len(orders) > 100:
            elements.append(Spacer(1, 0.2 * inch))
            elements.append(Paragraph(
                f"<i>Showing first 100 of {len(orders)} total orders. "
                f"Remaining orders total: Ksh {sum(o.total_amount for o in orders[100:]):,.2f}</i>",
                info_box_style
            ))
    else:
        elements.append(Paragraph(
            "No orders found matching the selected criteria.",
            info_box_style
        ))
    
    # === FINAL SUMMARY FOOTER ===
    elements.append(Spacer(1, 0.5 * inch))
    
    footer_summary_style = ParagraphStyle(
        'FooterSummary',
        parent=styles['Normal'],
        fontSize=10,
        textColor=HexColor('#495057'),
        fontName='Helvetica',
        alignment=TA_CENTER,
        backColor=HexColor('#E3F2FD'),
        borderWidth=2,
        borderColor=HexColor('#2E86AB'),
        borderPadding=15,
        leading=14
    )
    
    footer_text = f"""
    <b>REPORT SUMMARY</b><br/>
    This comprehensive report contains <b>{len(orders)}</b> orders with a total value of <b>Ksh {total_revenue:,.2f}</b><br/>
    Generated for <b>{user.get_full_name() or user.username}</b> on <b>{timezone.now().strftime('%B %d, %Y at %I:%M %p')}</b><br/>
    <br/>
    <i>For questions or clarifications, please contact your sales manager</i>
    """
    
    elements.append(Paragraph(footer_text, footer_summary_style))
    
    # Build the PDF
    doc.build(elements)
    buffer.seek(0)
    
    # Generate filename
    filename = f'sales_report_{timezone.now().strftime("%Y%m%d_%H%M%S")}.pdf'
    return FileResponse(buffer, as_attachment=True, filename=filename)


# # Custom gradient line for header (simulating a gradient effect with a colored bar)
class GradientLine(Flowable):
    def __init__(self, width, height, color_start, color_end):
        Flowable.__init__(self)
        self.width = width
        self.height = height
        self.color_start = color_start
        self.color_end = color_end

    def draw(self):
        self.canv.saveState()
        self.canv.setFillColor(self.color_start)
        self.canv.rect(0, 0, self.width, self.height, fill=1, stroke=0)
        self.canv.restoreState()

# Restrict access to superusers only
def is_admin(user):
    return user.is_superuser

@login_required
@user_passes_test(is_admin, login_url='/loginuser/')
def analytics_report(request):
    # Default date range: current week (Monday to Sunday)
    today = timezone.now().date()
    monday = today - timedelta(days=today.weekday())
    sunday = monday + timedelta(days=6)

    # Handle form submission for date range
    start_date = request.GET.get('start_date', monday.strftime('%Y-%m-%d'))
    end_date = request.GET.get('end_date', sunday.strftime('%Y-%m-%d'))

    try:
        start_date = dt.strptime(start_date, '%Y-%m-%d').date()
        end_date = dt.strptime(end_date, '%Y-%m-%d').date()
    except ValueError:
        start_date = monday
        end_date = sunday

    # Make dates timezone-aware
    start_of_period = timezone.make_aware(dt.combine(start_date, dt.min.time()))
    end_of_period = timezone.make_aware(dt.combine(end_date, dt.max.time()))

    # Analytics Data
    total_sales = Order.objects.filter(
        order_date__gte=start_of_period, order_date__lte=end_of_period, paid_status='completed'
    ).aggregate(total=Sum('total_amount'))['total'] or 0

    pending_revenue = Order.objects.filter(
        order_date__gte=start_of_period, order_date__lte=end_of_period, paid_status='pending'
    ).aggregate(total=Sum('total_amount'))['total'] or 0

    total_orders = Order.objects.filter(
        order_date__gte=start_of_period, order_date__lte=end_of_period
    ).count()

    completed_orders = Order.objects.filter(
        order_date__gte=start_of_period, order_date__lte=end_of_period, paid_status='completed'
    ).count()

    pending_orders = Order.objects.filter(
        order_date__gte=start_of_period, order_date__lte=end_of_period
    ).exclude(paid_status='completed').count()

    # Real-time stock levels with store-specific data
    stock_levels = Product.objects.annotate(
        total_stock=F('mcdave_stock') + F('kisii_stock') + F('offshore_stock')
    ).values('name', 'mcdave_stock', 'kisii_stock', 'offshore_stock', 'total_stock')

    top_products = OrderItem.objects.filter(
        order__order_date__gte=start_of_period, order__order_date__lte=end_of_period
    ).values('product__name').annotate(total_units=Sum('quantity')).order_by('-total_units')[:5]

    customers = Customer.objects.filter(
        orders__order_date__gte=start_of_period, orders__order_date__lte=end_of_period
    ).annotate(
        total_spent=Sum('orders__total_amount'),
        order_count=Count('orders')
    ).prefetch_related('orders')

    # Handle CSV download
    if 'download_csv' in request.GET:
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="analytics_report_{start_date}_to_{end_date}.csv"'
        writer = csv.writer(response)
        writer.writerow(['Summary Metrics'])
        writer.writerow(['Metric', 'Value'])
        writer.writerow(['Total Sales', f'${total_sales:.2f}'])
        writer.writerow(['Pending Revenue', f'${pending_revenue:.2f}'])
        writer.writerow(['Total Orders', total_orders])
        writer.writerow(['Completed Orders', completed_orders])
        writer.writerow(['Pending Orders', pending_orders])
        writer.writerow([])
        writer.writerow(['Stock Levels'])
        writer.writerow(['Product Name', 'McDave Stock', 'Kisii Stock', 'Offshore Stock', 'Total Stock'])
        for item in stock_levels:
            writer.writerow([item['name'], item['mcdave_stock'], item['kisii_stock'], item['offshore_stock'], item['total_stock']])
        writer.writerow([])
        writer.writerow(['Top Selling Products'])
        writer.writerow(['Rank', 'Product Name', 'Units Sold'])
        for i, item in enumerate(top_products, 1):
            writer.writerow([i, item['product__name'], item['total_units']])
        writer.writerow([])
        writer.writerow(['Customer Buying Metrics'])
        writer.writerow(['Customer Name', 'Total Spent', 'Order Count', 'Orders'])
        for customer in customers:
            orders = ', '.join(f'#Mc{order.id}Z' for order in customer.orders.all())
            writer.writerow([customer.get_full_name(), f'${customer.total_spent:.2f}', customer.order_count, orders])
        return response

    # Handle Excel download
    if 'download_excel' in request.GET:
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="analytics_report_{start_date}_to_{end_date}.xlsx"'
        wb = Workbook()
        ws = wb.active
        ws.title = "Analytics Report"
        ws.append(['Summary Metrics'])
        ws.append(['Metric', 'Value'])
        ws.append(['Total Sales', f'${total_sales:.2f}'])
        ws.append(['Pending Revenue', f'${pending_revenue:.2f}'])
        ws.append(['Total Orders', total_orders])
        ws.append(['Completed Orders', completed_orders])
        ws.append(['Pending Orders', pending_orders])
        ws.append([])
        ws.append(['Stock Levels'])
        ws.append(['Product Name', 'McDave Stock', 'Kisii Stock', 'Offshore Stock', 'Total Stock'])
        for item in stock_levels:
            ws.append([item['name'], item['mcdave_stock'], item['kisii_stock'], item['offshore_stock'], item['total_stock']])
        ws.append([])
        ws.append(['Top Selling Products'])
        ws.append(['Rank', 'Product Name', 'Units Sold'])
        for i, item in enumerate(top_products, 1):
            ws.append([i, item['product__name'], item['total_units']])
        ws.append([])
        ws.append(['Customer Buying Metrics'])
        ws.append(['Customer Name', 'Total Spent', 'Order Count', 'Order id'])
        for customer in customers:
            orders = ', '.join(f'#Mc{order.id}Z' for order in customer.orders.all())
            ws.append([customer.get_full_name(), f'${customer.total_spent:.2f}', customer.order_count, orders])
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                cell.font = Font(name='Helvetica', size=10)
                cell.alignment = Alignment(horizontal='left', vertical='center')
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = max_length + 2
            ws.column_dimensions[column].width = adjusted_width
        output = io.BytesIO()
        wb.save(output)
        response.write(output.getvalue())
        return response

    # Handle PDF download
    if 'download_pdf' in request.GET:
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="analytics_report_{start_date}_to_{end_date}.pdf"'
        doc = SimpleDocTemplate(response, pagesize=letter, leftMargin=0.5*inch, rightMargin=0.5*inch, topMargin=1*inch, bottomMargin=0.5*inch)
        styles = getSampleStyleSheet()
        styles.add(ParagraphStyle(name='Header', fontName='Helvetica-Bold', fontSize=20, textColor=HexColor('#2C3E50'), spaceAfter=6))
        styles.add(ParagraphStyle(name='SubHeader', fontName='Helvetica', fontSize=10, textColor=HexColor('#7F8C8D'), spaceAfter=12))
        styles.add(ParagraphStyle(name='SectionTitle', fontName='Helvetica-Bold', fontSize=14, textColor=HexColor('#34495E'), spaceBefore=12, spaceAfter=6))
        styles.add(ParagraphStyle(name='Footer', fontName='Helvetica', fontSize=8, textColor=HexColor('#7F8C8D'), alignment=1))
        elements = []
        def add_header_footer(canvas, doc):
            canvas.saveState()
            canvas.setFillColor(HexColor('#3498DB'))
            canvas.rect(0, doc.pagesize[1] - 0.3*inch, doc.pagesize[0], 0.3*inch, fill=1, stroke=0)
            canvas.setFont('Helvetica-Bold', 12)
            canvas.setFillColor(colors.white)
            canvas.drawString(0.5*inch, doc.pagesize[1] - 0.15*inch - 8, "Zelia Analytics Report")
            canvas.drawRightString(doc.pagesize[0] - 0.5*inch, doc.pagesize[1] - 0.15*inch - 8, f"{start_date.strftime('%Y-%m-%d')} to {end_date.strftime('%Y-%m-%d')}")
            canvas.restoreState()
            canvas.saveState()
            canvas.setFont('Helvetica', 8)
            canvas.setFillColor(HexColor('#7F8C8D'))
            page_num = f"Page {canvas.getPageNumber()}"
            canvas.drawCentredString(doc.pagesize[0]/2, 0.25*inch, page_num)
            canvas.drawString(0.5*inch, 0.25*inch, f"©McdaveZeliaoms All Rights Reserved")
            canvas.restoreState()
        elements.append(Paragraph("Analytics Report", styles['Header']))
        elements.append(Paragraph(f"Report Period: {start_date.strftime('%Y-%m-%d')} to {end_date.strftime('%Y-%m-%d')}", styles['SubHeader']))
        elements.append(Paragraph("Summary Metrics", styles['SectionTitle']))
        summary_data = [
            ['Total Sales', f'${total_sales:.2f}'],
            ['Pending Revenue', f'${pending_revenue:.2f}'],
            ['Total Orders', str(total_orders)],
            ['Completed Orders', str(completed_orders)],
            ['Pending Orders', str(pending_orders)],
        ]
        summary_table = Table(summary_data, colWidths=[2*inch, 2*inch])
        summary_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), HexColor('#ECF0F1')),
            ('TEXTCOLOR', (0, 0), (-1, 0), HexColor('#2C3E50')),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
            ('BACKGROUND', (0, 1), (-1, -1), colors.white),
            ('TEXTCOLOR', (0, 1), (-1, -1), HexColor('#34495E')),
            ('GRID', (0, 0), (-1, -1), 0.5, HexColor('#BDC3C7')),
            ('BOX', (0, 0), (-1, -1), 1, HexColor('#BDC3C7')),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ]))
        elements.append(summary_table)
        elements.append(Spacer(1, 0.25*inch))
        elements.append(Paragraph("Stock Levels", styles['SectionTitle']))
        stock_data = [['Product Name', 'McDave Stock', 'Kisii Stock', 'Offshore Stock', 'Total Stock']] + [[item['name'], str(item['mcdave_stock']), str(item['kisii_stock']), str(item['offshore_stock']), str(item['total_stock'])] for item in stock_levels]
        stock_table = Table(stock_data, colWidths=[3.5*inch, 1*inch, 1*inch, 1*inch, 1*inch])
        stock_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), HexColor('#ECF0F1')),
            ('TEXTCOLOR', (0, 0), (-1, 0), HexColor('#2C3E50')),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
            ('BACKGROUND', (0, 1), (-1, -1), colors.white),
            ('TEXTCOLOR', (0, 1), (-1, -1), HexColor('#34495E')),
            ('GRID', (0, 0), (-1, -1), 0.5, HexColor('#BDC3C7')),
            ('BOX', (0, 0), (-1, -1), 1, HexColor('#BDC3C7')),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ]))
        elements.append(stock_table)
        elements.append(Spacer(1, 0.25*inch))
        elements.append(Paragraph("Top Selling Products", styles['SectionTitle']))
        top_products_data = [['Rank', 'Product Name', 'Units Sold']] + [[str(i + 1), item['product__name'], str(item['total_units'])] for i, item in enumerate(top_products)]
        top_products_table = Table(top_products_data, colWidths=[1*inch, 3.5*inch, 1*inch])
        top_products_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), HexColor('#ECF0F1')),
            ('TEXTCOLOR', (0, 0), (-1, 0), HexColor('#2C3E50')),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
            ('BACKGROUND', (0, 1), (-1, -1), colors.white),
            ('TEXTCOLOR', (0, 1), (-1, -1), HexColor('#34495E')),
            ('GRID', (0, 0), (-1, -1), 0.5, HexColor('#BDC3C7')),
            ('BOX', (0, 0), (-1, -1), 1, HexColor('#BDC3C7')),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ]))
        elements.append(top_products_table)
        elements.append(Spacer(1, 0.25*inch))
        elements.append(Paragraph("Customer Buying Metrics", styles['SectionTitle']))
        customers_data = [['Customer Name', 'Total Spent', 'Orders', 'Order no']] + [[customer.get_full_name(), f'${customer.total_spent:.2f}', str(customer.order_count), ', '.join(f'#Mc{order.id}Z' for order in customer.orders.all())] for customer in customers]
        customers_table = Table(customers_data, colWidths=[2.5*inch, 1*inch, 0.75*inch, 2.5*inch])
        customers_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), HexColor('#ECF0F1')),
            ('TEXTCOLOR', (0, 0), (-1, 0), HexColor('#2C3E50')),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
            ('BACKGROUND', (0, 1), (-1, -1), colors.white),
            ('TEXTCOLOR', (0, 1), (-1, -1), HexColor('#34495E')),
            ('GRID', (0, 0), (-1, -1), 0.5, HexColor('#BDC3C7')),
            ('BOX', (0, 0), (-1, -1), 1, HexColor('#BDC3C7')),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ]))
        elements.append(customers_table)
        doc.build(elements, onFirstPage=add_header_footer, onLaterPages=add_header_footer)
        return response

    # Render the HTML page
    context = {
        'total_sales': total_sales,
        'pending_revenue': pending_revenue,
        'total_orders': total_orders,
        'completed_orders': completed_orders,
        'pending_orders': pending_orders,
        'stock_levels': stock_levels,
        'top_products': top_products,
        'customers': customers,
        'start_date': start_date.strftime('%Y-%m-%d'),
        'end_date': end_date.strftime('%Y-%m-%d'),
    }
    return render(request, 'reports/chart.html', context)

# Customer import form
class CustomerImportForm(forms.Form):
    excel_file = forms.FileField(
        label="Upload Excel File",
        widget=forms.FileInput(attrs={
            "class": "form-control",
            "accept": ".xlsx,.xls"
        })
    )

    def clean_excel_file(self):
        excel_file = self.cleaned_data['excel_file']
        max_size = 5 * 1024 * 1024  # 5MB
        if excel_file.size > max_size:
            raise forms.ValidationError("File size exceeds 5MB limit.")
        return excel_file
from openpyxl import load_workbook
@login_required
def import_customers(request):
    upload_error = None

    if request.method == 'POST':
        form = CustomerImportForm(request.POST, request.FILES)
        if form.is_valid():
            excel_file = request.FILES['excel_file']
            try:
                wb = load_workbook(excel_file, read_only=True)
                ws = wb.active
                headers = [cell.value.lower().strip().replace(' ', '_') if cell.value else '' for cell in next(ws.rows)]
                column_mappings = {
                    'shop_name': 'shop_name',
                    'contact_person': 'contact_person',
                    'email': 'email',
                    'phone_number': 'phone_number',
                    'address': 'address',
                    'default_category': 'default_category',
                }
                mapped_headers = [column_mappings.get(h, h) for h in headers]
                expected_columns = [
                    'shop_name', 'contact_person', 'email', 'phone_number', 'address', 'default_category'
                ]
                missing_columns = [col for col in expected_columns if col not in mapped_headers]
                if missing_columns:
                    upload_error = f"Missing required columns: {', '.join(missing_columns)}"
                    return render(request, 'imports/customer.html', {'form': form, 'upload_error': upload_error})

                valid_categories = [choice[0] for choice in Customer.CATEGORY_CHOICES if choice[0]]
                created_count = 0
                errors = []

                for index, row in enumerate(ws.iter_rows(min_row=2), start=2):
                    row_data = {mapped_headers[i]: cell.value for i, cell in enumerate(row)}
                    shop_name = str(row_data.get('shop_name', '')).strip()[:100]
                    contact_person = str(row_data.get('contact_person', '')).strip()[:100]
                    email = str(row_data.get('email', '')).strip()[:100] if row_data.get('email') else None
                    phone_number = str(row_data.get('phone_number', '')).strip()[:15] if row_data.get('phone_number') else None
                    address = str(row_data.get('address', '')).strip()[:100] if row_data.get('address') else ''
                    default_category = str(row_data.get('default_category', '')).strip()

                    if not shop_name or not contact_person:
                        errors.append(f"Row {index}: Missing shop_name or contact_person")
                        continue

                    if default_category and default_category not in valid_categories:
                        errors.append(f"Row {index}: Invalid category '{default_category}'")
                        continue

                    if email and Customer.objects.filter(email=email).exists():
                        errors.append(f"Row {index}: Email '{email}' already exists")
                        continue
                    if phone_number and Customer.objects.filter(phone_number=phone_number).exists():
                        errors.append(f"Row {index}: Phone number '{phone_number}' already exists")
                        continue

                    try:
                        customer = Customer(
                            first_name=shop_name,
                            last_name=contact_person,
                            email=email,
                            phone_number=phone_number,
                            address=address,
                            default_category=default_category or None,
                            sales_person=request.user if not request.user.userprofile.is_admin() else None,
                            created_at=timezone.now()
                        )
                        customer.save()
                        created_count += 1
                    except Exception as e:
                        errors.append(f"Row {index}: Error creating customer - {str(e)}")

                if created_count > 0:
                    messages.success(request, f"Successfully imported {created_count} customers")
                if errors:
                    for error in errors:
                        messages.error(request, error)

                return redirect('customers_list')
            except Exception as e:
                upload_error = f"Error processing Excel file: {str(e)}"
                return render(request, 'imports/customer.html', {'form': form, 'upload_error': upload_error})
        else:
            messages.error(request, "Invalid file format")
    else:
        form = CustomerImportForm()

    return render(request, 'imports/customer.html', {'form': form, 'upload_error': upload_error})
# ssss
@login_required
def activity_logs_view(request):
    user_profile = UserProfile.objects.get(user=request.user)
    if not user_profile.is_admin():
        return render(request, '403.html', {'error': 'Access denied'}, status=403)
    
    # Return all logs ordered by latest
    logs = ActivityLog.objects.all().order_by('-timestamp')
    
    context = {
        'logs': logs,
    }
    return render(request, 'reports/log.html', context)
from django.contrib.auth.decorators import login_required
from django.contrib.auth import authenticate
from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
from django.utils import timezone
import datetime

MAX_ATTEMPTS = 5          # lock after 5 wrong attempts
LOCKOUT_TIME = 3600        # 300 seconds (5 minutes)

@login_required
@csrf_exempt
def confirm_password(request):
    if request.method == "POST":
        user = request.user
        password = request.POST.get("password")

        # Initialize session tracking if not set
        if "failed_attempts" not in request.session:
            request.session["failed_attempts"] = 0
            request.session["lockout_until"] = None

        # Check if locked
        lockout_until = request.session.get("lockout_until")
        if lockout_until and timezone.now() < timezone.datetime.fromisoformat(lockout_until):
            remaining = (timezone.datetime.fromisoformat(lockout_until) - timezone.now()).seconds
            return JsonResponse({
                "success": False,
                "error": f"Too many failed attempts. Try again in {remaining} seconds."
            })

        # Authenticate
        check_user = authenticate(username=user.username, password=password)
        if check_user is not None:
            # Reset attempts
            request.session["failed_attempts"] = 0
            request.session["lockout_until"] = None
            return JsonResponse({"success": True, "redirect_url": "/secure-panel/shell/"})
        else:
            # Increment failed attempts
            request.session["failed_attempts"] += 1
            if request.session["failed_attempts"] >= MAX_ATTEMPTS:
                lock_until_time = timezone.now() + datetime.timedelta(seconds=LOCKOUT_TIME)
                request.session["lockout_until"] = lock_until_time.isoformat()
                return JsonResponse({
                    "success": False,
                    "error": f"Account locked. Try again in {LOCKOUT_TIME // 60} minutes."
                })

            return JsonResponse({
                "success": False,
                "error": f"Invalid password. {MAX_ATTEMPTS - request.session['failed_attempts']} attempts left."
            })

    return JsonResponse({"success": False, "error": "Invalid request"})






# =====================================================
# STOCK MANAGEMENT VIEWS
# =====================================================

@login_required
def stock_dashboard(request):
    """Main stock management dashboard with overview of all stores

    Also supports exporting the current stock list to CSV, Excel or PDF via
    ?download=csv|excel|pdf query parameter.
    """
    # handle export before doing the heavy context work
    download_format = request.GET.get('download')
    if download_format:
        prods = Product.objects.all()
        if download_format == 'csv':
            return generate_stock_csv(prods)
        elif download_format == 'excel':
            return generate_stock_excel(prods)
        elif download_format == 'pdf':
            return generate_stock_pdf(prods)

    user = request.user
    try:
        user_profile = UserProfile.objects.get(user=user)
    except UserProfile.DoesNotExist:
        return render(request, '403.html', {'error': 'User profile not found'}, status=403)

    # Get stock summary by store
    products = Product.objects.all()

    mcdave_total = sum(p.mcdave_stock for p in products)
    kisii_total = sum(p.kisii_stock for p in products)
    offshore_total = sum(p.offshore_stock for p in products)

    # Low stock alerts (threshold: 10)
    low_stock_threshold = 10
    low_stock_products = []
    for p in products:
        if p.mcdave_stock < low_stock_threshold:
            low_stock_products.append({'product': p, 'store': 'McDave', 'stock': p.mcdave_stock})
        if p.kisii_stock < low_stock_threshold:
            low_stock_products.append({'product': p, 'store': 'Mombasa', 'stock': p.kisii_stock})
        if p.offshore_stock < low_stock_threshold:
            low_stock_products.append({'product': p, 'store': 'Offshore', 'stock': p.offshore_stock})

    # Recent stock movements
    recent_movements = StockMovement.objects.select_related('product', 'recorded_by').order_by('-created_at')[:20]

    # Pending transfers
    pending_transfers = StockTransfer.objects.filter(status__in=['pending', 'in_transit']).order_by('-created_at')

    context = {
        'mcdave_total': mcdave_total,
        'kisii_total': kisii_total,
        'offshore_total': offshore_total,
        'total_stock': mcdave_total + kisii_total + offshore_total,
        'low_stock_products': low_stock_products[:10],
        'low_stock_count': len(low_stock_products),
        'recent_movements': recent_movements,
        'pending_transfers': pending_transfers,
        'is_admin': user_profile.is_admin(),
    }
    return render(request, 'stock/dashboard.html', context)


@login_required
def stock_list(request):
    """List all products with stock levels across all stores"""
    user = request.user
    try:
        user_profile = UserProfile.objects.get(user=user)
    except UserProfile.DoesNotExist:
        return render(request, '403.html', {'error': 'User profile not found'}, status=403)

    store_filter = request.GET.get('store', '')
    category_filter = request.GET.get('category', '')
    search = request.GET.get('search', '')

    products = Product.objects.select_related('category').all()

    if search:
        products = products.filter(Q(name__icontains=search) | Q(barcode__icontains=search))

    if category_filter:
        products = products.filter(category_id=category_filter)

    # Annotate with total stock
    products = products.annotate(
        total_stock=F('mcdave_stock') + F('kisii_stock') + F('offshore_stock')
    ).order_by('category__name', 'name')

    categories = Category.objects.all()

    context = {
        'products': products,
        'categories': categories,
        'store_filter': store_filter,
        'category_filter': category_filter,
        'search': search,
        'is_admin': user_profile.is_admin(),
    }
    return render(request, 'stock/list.html', context)


@login_required
def stock_adjustment(request):
    """Create a stock adjustment"""
    user = request.user
    try:
        user_profile = UserProfile.objects.get(user=user)
    except UserProfile.DoesNotExist:
        return render(request, '403.html', {'error': 'User profile not found'}, status=403)

    if not user_profile.is_admin():
        return render(request, '403.html', {'error': 'Only admins can make stock adjustments'}, status=403)

    if request.method == 'POST':
        product_id = request.POST.get('product')
        store = request.POST.get('store')
        new_quantity = int(request.POST.get('new_quantity', 0))
        reason = request.POST.get('reason')
        notes = request.POST.get('notes', '')

        try:
            product = Product.objects.get(id=product_id)

            # Get current stock
            stock_field = f"{store}_stock"
            previous_quantity = getattr(product, stock_field, 0)
            adjustment_quantity = new_quantity - previous_quantity

            with transaction.atomic():
                # Update product stock
                setattr(product, stock_field, new_quantity)
                product.save()

                # Create adjustment record
                adjustment = StockAdjustment.objects.create(
                    product=product,
                    store=store,
                    previous_quantity=previous_quantity,
                    new_quantity=new_quantity,
                    adjustment_quantity=adjustment_quantity,
                    reason=reason,
                    notes=notes,
                    adjusted_by=request.user
                )

                # Create stock movement record
                StockMovement.objects.create(
                    product=product,
                    store=store,
                    movement_type='adjustment',
                    quantity=adjustment_quantity,
                    previous_stock=previous_quantity,
                    new_stock=new_quantity,
                    notes=f"Adjustment: {dict(StockAdjustment.REASON_CHOICES).get(reason, reason)}. {notes}",
                    recorded_by=request.user
                )

                messages.success(request, f'Stock adjusted for {product.name}: {previous_quantity} → {new_quantity}')
                return redirect('stock_list')

        except Product.DoesNotExist:
            messages.error(request, 'Product not found')
        except Exception as e:
            messages.error(request, f'Error: {str(e)}')

    products = Product.objects.all().order_by('name')
    context = {
        'products': products,
        'stores': [('mcdave', 'McDave Store'), ('kisii', 'Mombasa Store'), ('offshore', 'Offshore Store')],
        'reasons': StockAdjustment.REASON_CHOICES,
    }
    return render(request, 'stock/adjustment.html', context)


@login_required
@admin_required
def stock_transfer_create(request):
    """Create a new stock transfer between stores"""
    user = request.user
    try:
        user_profile = UserProfile.objects.get(user=user)
    except UserProfile.DoesNotExist:
        return render(request, '403.html', {'error': 'User profile not found'}, status=403)

    if request.method == 'POST':
        from_store = request.POST.get('from_store')
        to_store = request.POST.get('to_store')
        notes = request.POST.get('notes', '')

        # Get products and quantities from form
        product_ids = request.POST.getlist('product_ids[]')
        quantities = request.POST.getlist('quantities[]')

        if from_store == to_store:
            messages.error(request, 'Cannot transfer to the same store')
            return redirect('stock_transfer_create')

        try:
            with transaction.atomic():
                # Create transfer
                transfer = StockTransfer.objects.create(
                    from_store=from_store,
                    to_store=to_store,
                    notes=notes,
                    initiated_by=request.user,
                    status='pending'
                )

                # Add items
                for product_id, qty in zip(product_ids, quantities):
                    if product_id and int(qty) > 0:
                        product = Product.objects.get(id=product_id)
                        StockTransferItem.objects.create(
                            transfer=transfer,
                            product=product,
                            quantity=int(qty)
                        )

                messages.success(request, f'Transfer #{transfer.id} created successfully')
                return redirect('stock_transfer_list')

        except Exception as e:
            messages.error(request, f'Error creating transfer: {str(e)}')

    products = Product.objects.all().order_by('name')
    context = {
        'products': products,
        'stores': StockTransfer.STORE_CHOICES,
    }
    return render(request, 'stock/transfer_create.html', context)


@login_required
def stock_transfer_list(request):
    """List all stock transfers"""
    user = request.user
    try:
        user_profile = UserProfile.objects.get(user=user)
    except UserProfile.DoesNotExist:
        return render(request, '403.html', {'error': 'User profile not found'}, status=403)

    status_filter = request.GET.get('status', '')

    transfers = StockTransfer.objects.select_related('initiated_by', 'received_by').prefetch_related('items__product')

    if status_filter:
        transfers = transfers.filter(status=status_filter)

    transfers = transfers.order_by('-created_at')

    context = {
        'transfers': transfers,
        'status_filter': status_filter,
        'status_choices': StockTransfer.STATUS_CHOICES,
        'is_admin': user_profile.is_admin(),
    }
    return render(request, 'stock/transfer_list.html', context)


@login_required
@admin_required
def stock_transfer_complete(request, transfer_id):
    """Complete a stock transfer and update stock levels"""
    user = request.user
    try:
        user_profile = UserProfile.objects.get(user=user)
    except UserProfile.DoesNotExist:
        return render(request, '403.html', {'error': 'User profile not found'}, status=403)

    transfer = get_object_or_404(StockTransfer, id=transfer_id)

    if transfer.status == 'completed':
        messages.warning(request, 'Transfer already completed')
        return redirect('stock_transfer_list')

    if transfer.status == 'cancelled':
        messages.error(request, 'Cannot complete a cancelled transfer')
        return redirect('stock_transfer_list')

    try:
        with transaction.atomic():
            from_stock_field = f"{transfer.from_store}_stock"
            to_stock_field = f"{transfer.to_store}_stock"

            for item in transfer.items.all():
                product = item.product

                # Get current stocks
                from_current = getattr(product, from_stock_field, 0)
                to_current = getattr(product, to_stock_field, 0)

                # Validate sufficient stock
                if from_current < item.quantity:
                    raise ValueError(f"Insufficient stock for {product.name} in {transfer.get_from_store_display()}")

                # Update stocks
                setattr(product, from_stock_field, from_current - item.quantity)
                setattr(product, to_stock_field, to_current + item.quantity)
                product.save()

                # Create movement records
                StockMovement.objects.create(
                    product=product,
                    store=transfer.from_store,
                    movement_type='transfer_out',
                    quantity=-item.quantity,
                    previous_stock=from_current,
                    new_stock=from_current - item.quantity,
                    transfer=transfer,
                    notes=f"Transfer to {transfer.get_to_store_display()}",
                    recorded_by=request.user
                )

                StockMovement.objects.create(
                    product=product,
                    store=transfer.to_store,
                    movement_type='transfer_in',
                    quantity=item.quantity,
                    previous_stock=to_current,
                    new_stock=to_current + item.quantity,
                    transfer=transfer,
                    notes=f"Transfer from {transfer.get_from_store_display()}",
                    recorded_by=request.user
                )

            # Update transfer status
            transfer.status = 'completed'
            transfer.completed_date = timezone.now()
            transfer.received_by = request.user
            transfer.save()

            messages.success(request, f'Transfer #{transfer.id} completed successfully')

    except ValueError as e:
        messages.error(request, str(e))
    except Exception as e:
        messages.error(request, f'Error completing transfer: {str(e)}')

    return redirect('stock_transfer_list')


@login_required
def stock_movement_history(request):
    """View stock movement history"""
    user = request.user
    try:
        user_profile = UserProfile.objects.get(user=user)
    except UserProfile.DoesNotExist:
        return render(request, '403.html', {'error': 'User profile not found'}, status=403)

    store_filter = request.GET.get('store', '')
    type_filter = request.GET.get('type', '')
    product_filter = request.GET.get('product', '')

    movements = StockMovement.objects.select_related('product', 'recorded_by', 'order').all()

    if store_filter:
        movements = movements.filter(store=store_filter)
    if type_filter:
        movements = movements.filter(movement_type=type_filter)
    if product_filter:
        movements = movements.filter(product_id=product_filter)

    movements = movements.order_by('-created_at')[:100]

    products = Product.objects.all().order_by('name')

    context = {
        'movements': movements,
        'products': products,
        'stores': StockMovement.STORE_CHOICES,
        'movement_types': StockMovement.MOVEMENT_TYPE_CHOICES,
        'store_filter': store_filter,
        'type_filter': type_filter,
        'product_filter': product_filter,
    }
    return render(request, 'stock/movement_history.html', context)


@login_required
@admin_required
def stock_receive(request):
    """Receive new stock (from supplier/purchase order)"""
    user = request.user
    try:
        user_profile = UserProfile.objects.get(user=user)
    except UserProfile.DoesNotExist:
        return render(request, '403.html', {'error': 'User profile not found'}, status=403)

    if not user_profile.is_admin():
        return render(request, '403.html', {'error': 'Only admins can receive stock'}, status=403)

    if request.method == 'POST':
        store = request.POST.get('store')
        reference = request.POST.get('reference', '')
        notes = request.POST.get('notes', '')
        product_ids = request.POST.getlist('product_ids[]')
        quantities = request.POST.getlist('quantities[]')

        try:
            with transaction.atomic():
                for product_id, qty in zip(product_ids, quantities):
                    if product_id and int(qty) > 0:
                        product = Product.objects.get(id=product_id)
                        stock_field = f"{store}_stock"
                        current_stock = getattr(product, stock_field, 0)
                        new_stock = current_stock + int(qty)

                        setattr(product, stock_field, new_stock)
                        product.save()

                        StockMovement.objects.create(
                            product=product,
                            store=store,
                            movement_type='in',
                            quantity=int(qty),
                            previous_stock=current_stock,
                            new_stock=new_stock,
                            reference_number=reference,
                            notes=notes,
                            recorded_by=request.user
                        )

                messages.success(request, 'Stock received successfully')
                return redirect('stock_dashboard')

        except Exception as e:
            messages.error(request, f'Error: {str(e)}')

    products = Product.objects.all().order_by('name')
    context = {
        'products': products,
        'stores': StockMovement.STORE_CHOICES,
    }
    return render(request, 'stock/receive.html', context)


@login_required
def get_product_stock(request):
    """API endpoint to get product stock for a specific store"""
    product_id = request.GET.get('product_id')
    store = request.GET.get('store')

    if not product_id or not store:
        return JsonResponse({'error': 'Missing parameters'}, status=400)

    try:
        product = Product.objects.get(id=product_id)
        stock_field = f"{store}_stock"
        stock = getattr(product, stock_field, 0)
        return JsonResponse({
            'product_id': product_id,
            'product_name': product.name,
            'store': store,
            'stock': stock
        })
    except Product.DoesNotExist:
        return JsonResponse({'error': 'Product not found'}, status=404)



# =====================================================
# CHATBOT VIEWS - UPDATED
# =====================================================

from datetime import datetime as dt, timedelta, time
from decimal import Decimal
from django.contrib.auth.decorators import login_required
from django.shortcuts import render
from django.http import JsonResponse
from django.db.models import Sum, Count, Q
from django.utils import timezone
import json
import re
from calendar import monthrange

from .models import (
    ChatMessage, ChatbotKnowledge, Order, OrderItem, 
    Product, Customer, User, UserProfile
)


# Helper functions for datetime handling
def make_aware_datetime(datetime_obj):
    """Safely make a datetime timezone-aware"""
    try:
        if timezone.is_aware(datetime_obj):
            return datetime_obj
        return timezone.make_aware(datetime_obj)
    except Exception:
        return datetime_obj


def get_date_range_datetimes(start_date, end_date):
    """Convert dates to timezone-aware datetime range"""
    start_dt = dt.combine(start_date, time.min)
    end_dt = dt.combine(end_date, time.max)
    return make_aware_datetime(start_dt), make_aware_datetime(end_dt)


@login_required
def chatbot_view(request):
    """Main chatbot interface"""
    # Get user's recent chat history
    chat_history = ChatMessage.objects.filter(user=request.user).order_by('-created_at')[:50]
    chat_history = reversed(list(chat_history))  # Show oldest first

    return render(request, 'chatbot/chat.html', {'chat_history': chat_history})


@login_required
def chatbot_send(request):
    """API endpoint to send a message to the chatbot"""
    if request.method != 'POST':
        return JsonResponse({'error': 'POST required'}, status=405)

    try:
        data = json.loads(request.body)
        user_message = data.get('message', '').strip()
    except:
        user_message = request.POST.get('message', '').strip()

    if not user_message:
        return JsonResponse({'error': 'Empty message'}, status=400)

    try:
        # Save user message
        ChatMessage.objects.create(
            user=request.user,
            sender='user',
            message=user_message
        )

        # Generate bot response
        bot_response = generate_chatbot_response(request.user, user_message)

        # Save bot response
        ChatMessage.objects.create(
            user=request.user,
            sender='bot',
            message=bot_response
        )

        return JsonResponse({
            'success': True,
            'response': bot_response
        })
    except Exception as e:
        import traceback
        error_msg = f"Error: {str(e)}\n{traceback.format_exc()}"
        print(error_msg)  # Log to console
        return JsonResponse({
            'success': False,
            'error': str(e),
            'response': f"Sorry, I encountered an error processing your request. Please try again."
        })


def parse_time_period(message_lower):
    """Parse time period from message and return start_date, end_date, period_name"""
    now = timezone.now()
    today = now.date()

    # Check for specific date patterns (e.g., "from 2024-01-01 to 2024-01-31" or "january 2024")
    date_range_match = re.search(r'from\s+(\d{4}-\d{2}-\d{2})\s+to\s+(\d{4}-\d{2}-\d{2})', message_lower)
    if date_range_match:
        try:
            start = dt.strptime(date_range_match.group(1), '%Y-%m-%d').date()
            end = dt.strptime(date_range_match.group(2), '%Y-%m-%d').date()
            return start, end, f"{start} to {end}"
        except:
            pass

    # Month names
    months = {
        'january': 1, 'february': 2, 'march': 3, 'april': 4, 'may': 5, 'june': 6,
        'july': 7, 'august': 8, 'september': 9, 'october': 10, 'november': 11, 'december': 12,
        'jan': 1, 'feb': 2, 'mar': 3, 'apr': 4, 'jun': 6, 'jul': 7, 'aug': 8, 'sep': 9, 'oct': 10, 'nov': 11, 'dec': 12
    }

    for month_name, month_num in months.items():
        if month_name in message_lower:
            year_match = re.search(r'20\d{2}', message_lower)
            year = int(year_match.group()) if year_match else today.year
            _, last_day = monthrange(year, month_num)
            start = dt(year, month_num, 1).date()
            end = dt(year, month_num, last_day).date()
            return start, end, f"{month_name.capitalize()} {year}"

    # Common period keywords
    if 'today' in message_lower:
        return today, today, "Today"
    elif 'yesterday' in message_lower:
        yesterday = today - timedelta(days=1)
        return yesterday, yesterday, "Yesterday"
    elif 'this week' in message_lower:
        start = today - timedelta(days=today.weekday())
        return start, today, "This Week"
    elif 'last week' in message_lower:
        start = today - timedelta(days=today.weekday() + 7)
        end = start + timedelta(days=6)
        return start, end, "Last Week"
    elif 'this month' in message_lower:
        start = today.replace(day=1)
        return start, today, "This Month"
    elif 'last month' in message_lower:
        first_of_month = today.replace(day=1)
        end = first_of_month - timedelta(days=1)
        start = end.replace(day=1)
        return start, end, "Last Month"
    elif 'this year' in message_lower:
        start = today.replace(month=1, day=1)
        return start, today, "This Year"
    elif 'last 7 days' in message_lower or 'past week' in message_lower:
        start = today - timedelta(days=7)
        return start, today, "Last 7 Days"
    elif 'last 30 days' in message_lower or 'past month' in message_lower:
        start = today - timedelta(days=30)
        return start, today, "Last 30 Days"
    elif 'last 90 days' in message_lower or 'past 3 months' in message_lower:
        start = today - timedelta(days=90)
        return start, today, "Last 90 Days"

    # Default to today
    return today, today, "Today"


def get_salesperson_from_message(message_lower, is_admin):
    """Extract salesperson name from message for admin queries"""
    if not is_admin:
        return None

    # Common patterns: "sales for john", "john's sales", "salesperson john", "by john"
    # Get all salespersons
    salespersons = User.objects.filter(
        Q(groups__name='Salespersons') | Q(is_superuser=True)
    ).distinct()

    for sp in salespersons:
        name_lower = sp.first_name.lower() if sp.first_name else sp.username.lower()
        username_lower = sp.username.lower()

        if name_lower in message_lower or username_lower in message_lower:
            return sp

    return None


def generate_chatbot_response(user, message):
    """Generate a response based on the user's message"""
    message_lower = message.lower()

    # Check knowledge base first
    knowledge_items = ChatbotKnowledge.objects.filter(is_active=True).order_by('-priority')
    for item in knowledge_items:
        keywords = [k.strip().lower() for k in item.keywords.split(',')]
        if any(keyword in message_lower for keyword in keywords):
            return item.answer

    # Built-in responses with dynamic data
    try:
        user_profile = UserProfile.objects.get(user=user)
        is_admin = user_profile.is_admin()
    except:
        is_admin = False

    # Greetings
    greetings = ['hello', 'hi', 'hey', 'good morning', 'good afternoon', 'good evening']
    if any(g in message_lower for g in greetings):
        role = "Admin" if is_admin else "Salesperson"
        return f"""Hello {user.first_name or user.username}! I'm Zelia Assistant.

**Your Role:** {role}

I can help you with:
- **Analytics** - "my sales this month", "analytics this week"
- **Orders** - "orders today", "pending orders"
- **Stock** - "stock levels", "low stock"
- **Reports** - "sales report", "top products"
{"- **Team Analytics** - 'sales by [name]', 'team performance'" if is_admin else ""}

Just ask me anything!"""

    # Help/Menu
    if any(word in message_lower for word in ['help', 'what can you do', 'menu', 'commands', 'options']):
        admin_help = """
**Admin Analytics**
- "team performance this month" - All salespersons summary
- "sales by [name]" - Specific salesperson analytics
- "overall sales this week" - Company-wide metrics
- "top salesperson this month" - Leaderboard
- "compare salespersons" - Performance comparison""" if is_admin else ""

        return f"""**Zelia Assistant - Full Command List**

**Sales & Analytics**
- "my sales today/this week/this month" - Your sales summary
- "my performance" - Your detailed metrics
- "analytics [period]" - Detailed analytics report
- "revenue this month" - Revenue breakdown
{admin_help}

**Orders**
- "orders today" - Today's orders with details
- "pending orders" - Unpaid orders
- "recent orders" - Latest 5 orders
- "orders this week" - Weekly order summary

**Products**
- "low stock" - Products running low
- "top selling products" - Best performers
- "product performance" - Product analytics

**Customers**
- "total customers" - Customer count
- "new customers this week" - Recent additions
- "top customers" - Highest value customers

**Stock**
- "stock levels" - Inventory overview
- "stock alerts" - Low stock warnings
- "stock by store" - Per-store inventory

**Time Periods You Can Use:**
today, yesterday, this week, last week, this month, last month, this year, last 7 days, last 30 days, January 2024, from 2024-01-01 to 2024-01-31"""

    # ============================================
    # ANALYTICS & SALES RECORDS
    # ============================================
    analytics_keywords = ['analytics', 'sales', 'performance', 'revenue', 'my record', 'my sales', 'report', 'summary']
    if any(kw in message_lower for kw in analytics_keywords):

        # Parse time period
        start_date, end_date, period_name = parse_time_period(message_lower)
        start_dt, end_dt = get_date_range_datetimes(start_date, end_date)

        # Check if admin is querying specific salesperson
        target_user = None
        querying_team = False
        querying_all = False

        if is_admin:
            target_user = get_salesperson_from_message(message_lower, is_admin)
            querying_team = any(kw in message_lower for kw in ['team', 'all salesperson', 'everyone', 'staff', 'compare'])
            querying_all = any(kw in message_lower for kw in ['overall', 'total', 'company', 'all sales'])

        # ---- TEAM PERFORMANCE (Admin only) ----
        if is_admin and (querying_team or 'leaderboard' in message_lower or 'top salesperson' in message_lower):
            salespersons = User.objects.filter(
                Q(groups__name='Salespersons') | Q(groups__name='Admins')
            ).distinct()

            team_stats = []
            for sp in salespersons:
                sp_orders = Order.objects.filter(
                    sales_person=sp,
                    order_date__range=[start_dt, end_dt]
                )
                order_count = sp_orders.count()
                total_sales = sp_orders.aggregate(t=Sum('total_amount'))['t'] or Decimal('0')
                paid_amount = sp_orders.aggregate(t=Sum('amount_paid'))['t'] or Decimal('0')

                if order_count > 0:
                    team_stats.append({
                        'name': sp.first_name or sp.username,
                        'orders': order_count,
                        'sales': total_sales,
                        'collected': paid_amount
                    })

            # Sort by sales
            team_stats.sort(key=lambda x: x['sales'], reverse=True)

            if not team_stats:
                return f"No sales data found for **{period_name}**."

            total_team_sales = sum(s['sales'] for s in team_stats)
            total_orders = sum(s['orders'] for s in team_stats)

            leaderboard = []
            for i, stat in enumerate(team_stats[:10], 1):
                medal = "🥇" if i == 1 else "🥈" if i == 2 else "🥉" if i == 3 else f"{i}."
                percentage = (stat['sales'] / total_team_sales * 100) if total_team_sales > 0 else 0
                leaderboard.append(
                    f"{medal} **{stat['name']}** - {stat['orders']} orders | KES {stat['sales']:,.2f} ({percentage:.1f}%)"
                )

            return f"""**Team Performance - {period_name}**

**Overall:**
- Total Orders: **{total_orders}**
- Total Sales: **KES {total_team_sales:,.2f}**
- Active Salespersons: **{len(team_stats)}**

**Leaderboard:**
{chr(10).join(leaderboard)}"""

        # ---- SPECIFIC SALESPERSON (Admin querying) ----
        if is_admin and target_user:
            orders = Order.objects.filter(
                sales_person=target_user,
                order_date__range=[start_dt, end_dt]
            ).order_by('-order_date')

            order_count = orders.count()
            total_sales = orders.aggregate(t=Sum('total_amount'))['t'] or Decimal('0')
            amount_paid = orders.aggregate(t=Sum('amount_paid'))['t'] or Decimal('0')
            pending_amount = total_sales - amount_paid

            # Payment status breakdown
            completed_orders = orders.filter(paid_status='completed').count()
            pending_orders = orders.filter(paid_status='pending').count()
            partial_orders = orders.filter(paid_status='partially_paid').count()

            # Store breakdown
            store_stats = orders.values('store').annotate(
                count=Count('id'),
                total=Sum('total_amount')
            )
            store_breakdown = []
            for s in store_stats:
                store_name = dict(Order.STORE_CHOICES).get(s['store'], s['store'])
                store_breakdown.append(f"- {store_name}: {s['count']} orders | KES {s['total']:,.2f}")

            # Recent orders
            recent = orders[:5]
            recent_list = []
            for o in recent:
                recent_list.append(
                    f"- #{o.id} | {o.customer.first_name} | KES {o.total_amount:,.2f} | {o.get_paid_status_display()}"
                )

            sp_name = target_user.first_name or target_user.username
            return f"""**Sales Report: {sp_name} - {period_name}**

**Summary:**
- Total Orders: **{order_count}**
- Total Sales: **KES {total_sales:,.2f}**
- Amount Collected: **KES {amount_paid:,.2f}**
- Pending Collection: **KES {pending_amount:,.2f}**

**Order Status:**
- Completed: {completed_orders} | Partial: {partial_orders} | Pending: {pending_orders}

**By Store:**
{chr(10).join(store_breakdown) if store_breakdown else "No orders"}

**Recent Orders:**
{chr(10).join(recent_list) if recent_list else "No recent orders"}"""

        # ---- OVERALL COMPANY SALES (Admin) ----
        if is_admin and querying_all:
            orders = Order.objects.filter(order_date__range=[start_dt, end_dt])

            order_count = orders.count()
            total_sales = orders.aggregate(t=Sum('total_amount'))['t'] or Decimal('0')
            amount_paid = orders.aggregate(t=Sum('amount_paid'))['t'] or Decimal('0')
            pending_amount = total_sales - amount_paid

            # Daily average
            days = (end_date - start_date).days + 1
            daily_avg = total_sales / days if days > 0 else total_sales

            # Store breakdown
            store_stats = orders.values('store').annotate(
                count=Count('id'),
                total=Sum('total_amount')
            ).order_by('-total')

            store_breakdown = []
            for s in store_stats:
                store_name = dict(Order.STORE_CHOICES).get(s['store'], s['store'])
                store_breakdown.append(f"- **{store_name}**: {s['count']} orders | KES {s['total']:,.2f}")

            # Category breakdown
            cat_stats = orders.values('customer_category').annotate(
                count=Count('id'),
                total=Sum('total_amount')
            ).order_by('-total')

            cat_breakdown = []
            for c in cat_stats:
                cat_name = dict(Order.CATEGORY_CHOICES).get(c['customer_category'], c['customer_category'])
                cat_breakdown.append(f"- {cat_name}: {c['count']} orders | KES {c['total']:,.2f}")

            # Top products
            top_products = OrderItem.objects.filter(
                order__order_date__range=[start_dt, end_dt]
            ).values('product__name').annotate(
                qty=Sum('quantity'),
                revenue=Sum('line_total')
            ).order_by('-revenue')[:5]

            product_list = []
            for p in top_products:
                product_list.append(f"- {p['product__name']}: {p['qty']} units | KES {p['revenue']:,.2f}")

            return f"""**Company Sales Overview - {period_name}**

**Key Metrics:**
- Total Orders: **{order_count}**
- Total Revenue: **KES {total_sales:,.2f}**
- Amount Collected: **KES {amount_paid:,.2f}**
- Outstanding: **KES {pending_amount:,.2f}**
- Daily Average: **KES {daily_avg:,.2f}**

**By Store:**
{chr(10).join(store_breakdown) if store_breakdown else "No data"}

**By Customer Type:**
{chr(10).join(cat_breakdown[:5]) if cat_breakdown else "No data"}

**Top Products:**
{chr(10).join(product_list) if product_list else "No data"}"""

        # ---- PERSONAL SALES (Default for salesperson, or admin viewing own) ----
        if is_admin:
            orders = Order.objects.filter(order_date__range=[start_dt, end_dt]).order_by('-order_date')
            scope = "All Orders"
        else:
            orders = Order.objects.filter(
                sales_person=user,
                order_date__range=[start_dt, end_dt]
            ).order_by('-order_date')
            scope = "My Sales"

        order_count = orders.count()
        total_sales = orders.aggregate(t=Sum('total_amount'))['t'] or Decimal('0')
        amount_paid = orders.aggregate(t=Sum('amount_paid'))['t'] or Decimal('0')
        pending_amount = total_sales - amount_paid

        if order_count == 0:
            return f"No sales recorded for **{period_name}**."

        # Payment breakdown
        completed = orders.filter(paid_status='completed').count()
        pending = orders.filter(paid_status='pending').count()
        partial = orders.filter(paid_status='partially_paid').count()

        # Store breakdown
        store_stats = orders.values('store').annotate(
            count=Count('id'),
            total=Sum('total_amount')
        )
        store_lines = []
        for s in store_stats:
            store_name = dict(Order.STORE_CHOICES).get(s['store'], s['store'])
            store_lines.append(f"- {store_name}: {s['count']} orders | KES {s['total']:,.2f}")

        # Recent orders with details
        recent = orders[:8]
        order_lines = []
        for o in recent:
            store_name = dict(Order.STORE_CHOICES).get(o.store, o.store)
            order_lines.append(
                f"- **#Mc{o.id}Z** | {o.customer.first_name} | {store_name} | KES {o.total_amount:,.2f} | {o.get_paid_status_display()}"
            )

        more_text = f"\n*...and {order_count - 8} more orders*" if order_count > 8 else ""

        return f"""**{scope} - {period_name}**

**Summary:**
- Total Orders: **{order_count}**
- Total Sales: **KES {total_sales:,.2f}**
- Collected: **KES {amount_paid:,.2f}**
- Pending: **KES {pending_amount:,.2f}**

**Payment Status:**
Completed: {completed} | Partial: {partial} | Pending: {pending}

**By Store:**
{chr(10).join(store_lines)}

**Order Details:**
{chr(10).join(order_lines)}{more_text}"""

    # ============================================
    # ORDERS (Enhanced)
    # ============================================
    if 'order' in message_lower:
        start_date, end_date, period_name = parse_time_period(message_lower)
        start_dt, end_dt = get_date_range_datetimes(start_date, end_date)

        if 'pending' in message_lower:
            if is_admin:
                pending_orders = Order.objects.filter(paid_status='pending').order_by('-order_date')[:10]
                total_pending = Order.objects.filter(paid_status='pending').aggregate(t=Sum('total_amount'))['t'] or 0
                count = Order.objects.filter(paid_status='pending').count()
            else:
                pending_orders = Order.objects.filter(sales_person=user, paid_status='pending').order_by('-order_date')[:10]
                total_pending = Order.objects.filter(sales_person=user, paid_status='pending').aggregate(t=Sum('total_amount'))['t'] or 0
                count = Order.objects.filter(sales_person=user, paid_status='pending').count()

            if count == 0:
                return "No pending orders found."

            order_lines = []
            for o in pending_orders:
                order_lines.append(f"- **#MC{o.id}Z** | {o.customer.first_name} | KES {o.total_amount:,.2f}")

            return f"""**Pending Orders**
Count: **{count}** | Total Value: **KES {total_pending:,.2f}**

{chr(10).join(order_lines)}"""

        if 'recent' in message_lower or 'latest' in message_lower:
            if is_admin:
                recent = Order.objects.order_by('-created_at')[:10]
            else:
                recent = Order.objects.filter(sales_person=user).order_by('-created_at')[:10]

            if recent:
                order_list = []
                for o in recent:
                    store_name = dict(Order.STORE_CHOICES).get(o.store, o.store)
                    order_list.append(f"- **#MC{o.id}Z** | {o.customer.first_name} | {store_name} | KES {o.total_amount:,.2f} | {o.get_paid_status_display()}")
                return f"**Recent Orders:**\n{chr(10).join(order_list)}"
            return "No recent orders found."

        if 'create' in message_lower or 'how to' in message_lower or 'make' in message_lower:
            return """**How to Create an Order:**

1. Go to **Orders** → **Create Order**
2. **Step 1 - Customer Details:**
   - Search and select a customer
   - Choose the store (McDave, Mombasa, Offshore)
   - Set customer category and VAT preference

3. **Step 2 - Add Products:**
   - Click "Add Product"
   - Search for products
   - Set quantities and variance if needed

4. **Step 3 - Review:**
   - Check order summary
   - Verify totals
   - Click "Create Order"

The system will automatically generate a receipt PDF!"""

        # Default: show orders for the period
        if is_admin:
            orders = Order.objects.filter(order_date__range=[start_dt, end_dt]).order_by('-order_date')
        else:
            orders = Order.objects.filter(sales_person=user, order_date__range=[start_dt, end_dt]).order_by('-order_date')

        count = orders.count()
        total = orders.aggregate(t=Sum('total_amount'))['t'] or Decimal('0')

        if count == 0:
            return f"No orders found for **{period_name}**."

        order_details = []
        for o in orders[:10]:
            store_display = dict(Order.STORE_CHOICES).get(o.store, o.store)
            order_details.append(
                f"- **#Mc{o.id}Z** | {o.customer.first_name} | {store_display} | KES {o.total_amount:,.2f} | {o.get_paid_status_display()}"
            )

        more_text = f"\n\n*...and {count - 10} more orders*" if count > 10 else ""

        return f"""**Orders - {period_name}**
Total: **{count}** orders | Value: **KES {total:,.2f}**

{chr(10).join(order_details)}{more_text}"""

    # ============================================
    # PRODUCTS
    # ============================================
    if 'product' in message_lower or 'top selling' in message_lower or 'best seller' in message_lower:
        if 'low stock' in message_lower or 'stock alert' in message_lower:
            low_stock = Product.objects.filter(
                Q(mcdave_stock__lt=10) | Q(kisii_stock__lt=10) | Q(offshore_stock__lt=10)
            )[:10]
            if low_stock:
                product_list = '\n'.join([f"- **{p.name}**: McDave={p.mcdave_stock}, Mombasa={p.kisii_stock}, Offshore={p.offshore_stock}" for p in low_stock])
                return f"**Low Stock Products (under 10 units):**\n{product_list}"
            return "Great news! No products are currently low on stock."

        if 'top' in message_lower or 'best' in message_lower or 'selling' in message_lower:
            start_date, end_date, period_name = parse_time_period(message_lower)
            start_dt, end_dt = get_date_range_datetimes(start_date, end_date)

            top = OrderItem.objects.filter(
                order__order_date__range=[start_dt, end_dt]
            ).values('product__name').annotate(
                total_qty=Sum('quantity'),
                total_revenue=Sum('line_total')
            ).order_by('-total_revenue')[:10]

            if top:
                product_list = []
                for i, p in enumerate(top, 1):
                    product_list.append(f"{i}. **{p['product__name']}** - {p['total_qty']} units | KES {p['total_revenue']:,.2f}")
                return f"**Top Selling Products - {period_name}:**\n{chr(10).join(product_list)}"
            return f"No sales data available for **{period_name}**."

    # ============================================
    # CUSTOMERS
    # ============================================
    if 'customer' in message_lower:
        if 'top' in message_lower or 'best' in message_lower or 'valuable' in message_lower:
            start_date, end_date, period_name = parse_time_period(message_lower)
            start_dt, end_dt = get_date_range_datetimes(start_date, end_date)

            top_customers = Order.objects.filter(
                order_date__range=[start_dt, end_dt]
            ).values('customer__first_name', 'customer__id').annotate(
                order_count=Count('id'),
                total_spent=Sum('total_amount')
            ).order_by('-total_spent')[:10]

            if top_customers:
                customer_list = []
                for i, c in enumerate(top_customers, 1):
                    customer_list.append(f"{i}. **{c['customer__first_name']}** - {c['order_count']} orders | KES {c['total_spent']:,.2f}")
                return f"**Top Customers - {period_name}:**\n{chr(10).join(customer_list)}"
            return f"No customer data for **{period_name}**."

        if 'total' in message_lower or 'count' in message_lower or 'how many' in message_lower:
            if is_admin:
                count = Customer.objects.count()
            else:
                count = Customer.objects.filter(Q(sales_person=user) | Q(sales_person__isnull=True)).count()
            return f"Total customers: **{count}**"

        if 'new' in message_lower or 'recent' in message_lower:
            start_date, end_date, period_name = parse_time_period(message_lower)
            start_dt, end_dt = get_date_range_datetimes(start_date, end_date)

            if is_admin:
                new_customers = Customer.objects.filter(created_at__range=[start_dt, end_dt]).order_by('-created_at')[:10]
                new_count = Customer.objects.filter(created_at__range=[start_dt, end_dt]).count()
            else:
                new_customers = Customer.objects.filter(
                    Q(sales_person=user) | Q(sales_person__isnull=True),
                    created_at__range=[start_dt, end_dt]
                ).order_by('-created_at')[:10]
                new_count = Customer.objects.filter(
                    Q(sales_person=user) | Q(sales_person__isnull=True),
                    created_at__range=[start_dt, end_dt]
                ).count()

            if new_count == 0:
                return f"No new customers added during **{period_name}**."

            customer_list = [f"- **{c.first_name}** | {c.address or 'N/A'} | {c.phone_number or 'N/A'}" for c in new_customers]
            return f"""**New Customers - {period_name}**
Total: **{new_count}**

{chr(10).join(customer_list)}"""

    # ============================================
    # STOCK
    # ============================================
    if 'stock' in message_lower:
        if 'level' in message_lower or 'overview' in message_lower or 'total' in message_lower:
            products = Product.objects.all()
            mcdave = sum(p.mcdave_stock for p in products)
            mombasa = sum(p.kisii_stock for p in products)
            offshore = sum(p.offshore_stock for p in products)
            total_products = products.count()
            total_stock = mcdave + mombasa + offshore

            return f"""**Stock Overview**

**By Store:**
- McDave Store: **{mcdave:,}** units
- Mombasa Store: **{mombasa:,}** units
- Offshore Store: **{offshore:,}** units

**Totals:**
- Total Products: **{total_products}**
- Total Stock: **{total_stock:,}** units
- Average per Product: **{total_stock // total_products if total_products > 0 else 0}** units"""

        if 'alert' in message_lower or 'low' in message_lower:
            low_stock = Product.objects.filter(
                Q(mcdave_stock__lt=10) | Q(kisii_stock__lt=10) | Q(offshore_stock__lt=10)
            )
            low_count = low_stock.count()

            if low_count == 0:
                return "No stock alerts. All products have adequate inventory."

            product_list = []
            for p in low_stock[:10]:
                product_list.append(f"- **{p.name}**: McDave={p.mcdave_stock}, Mombasa={p.kisii_stock}, Offshore={p.offshore_stock}")

            return f"""**Stock Alerts - {low_count} Products Low**

{chr(10).join(product_list)}

{"*...and more. Check Stock Management for full list.*" if low_count > 10 else ""}"""

    # ============================================
    # SYSTEM FEATURES
    # ============================================
    if 'feature' in message_lower or 'what does' in message_lower or 'system' in message_lower:
        return """**Zelia OMS Features:**

**Order Management**
- Create, edit, and track orders
- Multi-store support (McDave, Mombasa, Offshore)
- Payment tracking & partial payments

**Analytics & Reports**
- Real-time sales analytics
- Salesperson performance tracking
- Revenue reports by period
- Customer insights

**Product Management**
- Product catalog with categories
- Multi-tier pricing (Factory, Distributor, Wholesale, Retail)
- Barcode support

**Stock Management**
- Real-time inventory across stores
- Stock transfers between stores
- Low stock alerts
- Movement history

**Customer Management**
- Customer database
- Purchase history
- Category-based pricing"""

    # ============================================
    # DEFAULT RESPONSE
    # ============================================
    return f"""I'm not sure I understood that. Here are some things you can try:

**Quick Commands:**
- "my sales this month" - Your sales summary
- "orders today" - Today's orders
- "analytics this week" - Weekly analytics
- "top products" - Best selling products
- "stock levels" - Inventory overview
{"- 'team performance' - All salespersons" if is_admin else ""}
{"- 'sales by [name]' - Specific salesperson" if is_admin else ""}

**Time Periods:**
today, yesterday, this week, last week, this month, last month, January 2024, etc.

Type **"help"** for full command list!"""


@login_required
def chatbot_clear(request):
    """Clear chat history"""
    if request.method == 'POST':
        ChatMessage.objects.filter(user=request.user).delete()
        return JsonResponse({'success': True})
    return JsonResponse({'error': 'POST required'}, status=405)
    
    
# statements
# statements - Enhanced with Multi-Customer Support
# statements - Enhanced with Multi-Customer Support
from django.contrib.auth.decorators import login_required
from django.shortcuts import render
from django.http import HttpResponse
from django.db.models import Sum, Count, Q, Max, F
from django.utils import timezone
from datetime import datetime as dt
from decimal import Decimal

# ReportLab imports for PDF generation
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib import colors
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak, Image
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_LEFT, TA_CENTER, TA_RIGHT
from reportlab.pdfgen import canvas
from reportlab.lib.colors import HexColor
from io import BytesIO

# Import your models
from .models import Order, OrderItem, Customer, UserProfile, User


# Custom Canvas for Professional Header/Footer
class CustomerStatementCanvas(canvas.Canvas):
    """Compact custom canvas with minimal header/footer for maximum content space"""
    
    def __init__(self, *args, **kwargs):
        canvas.Canvas.__init__(self, *args, **kwargs)
        self.pages = []
        
    def showPage(self):
        self.pages.append(dict(self.__dict__))
        self._startPage()
        
    def save(self):
        page_count = len(self.pages)
        for page_num, page in enumerate(self.pages, start=1):
            self.__dict__.update(page)
            self.draw_page_decorations(page_num, page_count)
            canvas.Canvas.showPage(self)
        canvas.Canvas.save(self)
        
    def draw_page_decorations(self, page_num, page_count):
        """Draw compact header and footer"""
        page_width, page_height = letter
        
        # === ULTRA-COMPACT HEADER (30px instead of 40px) ===
        self.setFillColor(HexColor('#1a365d'))
        self.rect(0, page_height - 30, page_width, 30, fill=True, stroke=False)
        
        # Header accent line
        self.setFillColor(HexColor('#38a169'))
        self.rect(0, page_height - 32, page_width, 2, fill=True, stroke=False)
        
        # Company name - smaller font
        self.setFillColor(colors.white)
        self.setFont("Helvetica-Bold", 11)  # Reduced from 14
        self.drawString(30, page_height - 19, "McDave Sales - Customer Statement")
        
        # Page number in header (right side) - smaller
        self.setFont("Helvetica", 8)  # Reduced from 9
        page_text = f"Page {page_num}/{page_count}"
        self.drawRightString(page_width - 30, page_height - 19, page_text)
        
        # === ULTRA-COMPACT FOOTER (22px instead of 30px) ===
        self.setFillColor(HexColor('#f7fafc'))
        self.rect(0, 0, page_width, 22, fill=True, stroke=False)
        
        # Footer accent line
        self.setFillColor(HexColor('#1a365d'))
        self.rect(0, 22, page_width, 1, fill=True, stroke=False)
        
        # Footer info - smaller font, single line
        self.setFillColor(HexColor('#718096'))
        self.setFont("Helvetica", 6)  # Reduced from 7
        timestamp = timezone.now().strftime('%b %d, %Y %I:%M %p')
        self.drawString(30, 11, f"Generated: {timestamp}")
        
        self.setFont("Helvetica-Oblique", 6)  # Reduced from 7
        self.drawCentredString(page_width / 2, 11, "CONFIDENTIAL")
        
        self.setFont("Helvetica", 6)  # Reduced from 7
        self.drawRightString(page_width - 30, 11, "McDave Sales")


@login_required
def customer_statements_view(request):
    """
    Enhanced view for generating detailed customer account statements.
    Supports MULTIPLE customer selection with individual breakdowns.
    """
    user = request.user
    
    # Check user permissions
    try:
        user_profile = UserProfile.objects.get(user=user)
    except UserProfile.DoesNotExist:
        return render(request, '403.html', {'error': 'User profile not found'}, status=403)

    # Get filter parameters
    selected_customer_ids = request.GET.getlist('customers')  # Changed to getlist for multiple
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    paid_status = request.GET.get('paid_status', '')
    store_filter = request.GET.get('store', '')  # New: filter by store
    
    # Base queryset - filter based on user role
    if user_profile.is_admin():
        customers = Customer.objects.filter(orders__isnull=False).distinct().order_by('first_name')
        orders = Order.objects.all()
    else:
        customers = Customer.objects.filter(
            orders__sales_person=user
        ).distinct().order_by('first_name')
        orders = Order.objects.filter(sales_person=user)

    # Initialize variables
    selected_customers = []
    all_customer_data = []
    combined_summary = {
        'total_orders': 0,
        'total_spent': Decimal('0.00'),
        'total_paid': Decimal('0.00'),
        'total_balance': Decimal('0.00'),
        'completed_orders': 0,
        'pending_orders': 0,
        'partially_paid_orders': 0,
    }
    
    # If customers are selected, get their data
    if selected_customer_ids:
        try:
            selected_customers = Customer.objects.filter(id__in=selected_customer_ids)
            
            # Process each customer
            for customer in selected_customers:
                customer_orders = orders.filter(customer=customer)
                
                # Apply date filters
                if start_date:
                    try:
                        start_date_obj = dt.strptime(start_date, '%Y-%m-%d')
                        start_date_obj = timezone.make_aware(start_date_obj)
                        customer_orders = customer_orders.filter(order_date__gte=start_date_obj)
                    except ValueError:
                        pass
                
                if end_date:
                    try:
                        end_date_obj = dt.strptime(end_date, '%Y-%m-%d')
                        end_date_obj = end_date_obj.replace(hour=23, minute=59, second=59)
                        end_date_obj = timezone.make_aware(end_date_obj)
                        customer_orders = customer_orders.filter(order_date__lte=end_date_obj)
                    except ValueError:
                        pass
                
                # Apply paid status filter
                if paid_status:
                    customer_orders = customer_orders.filter(paid_status=paid_status)
                
                # Apply store filter
                if store_filter:
                    customer_orders = customer_orders.filter(store=store_filter)
                
                # Calculate customer summary
                total_orders = customer_orders.count()
                total_spent = customer_orders.aggregate(total=Sum('total_amount'))['total'] or Decimal('0.00')
                total_paid = customer_orders.aggregate(total=Sum('amount_paid'))['total'] or Decimal('0.00')
                total_balance = total_spent - total_paid
                
                completed_orders = customer_orders.filter(paid_status='completed').count()
                pending_orders = customer_orders.filter(paid_status='pending').count()
                partially_paid_orders = customer_orders.filter(paid_status='partially_paid').count()
                
                last_order = customer_orders.order_by('-order_date').first()
                first_order = customer_orders.order_by('order_date').first()
                
                # Store breakdown (orders by store)
                store_breakdown = []
                for store_choice in Order.STORE_CHOICES:
                    store_code = store_choice[0]
                    store_name = store_choice[1]
                    store_orders = customer_orders.filter(store=store_code)
                    store_count = store_orders.count()
                    
                    if store_count > 0:
                        store_total = store_orders.aggregate(total=Sum('total_amount'))['total'] or Decimal('0.00')
                        store_paid = store_orders.aggregate(total=Sum('amount_paid'))['total'] or Decimal('0.00')
                        store_breakdown.append({
                            'store_code': store_code,
                            'store_name': store_name,
                            'order_count': store_count,
                            'total_amount': store_total,
                            'amount_paid': store_paid,
                            'balance': store_total - store_paid,
                        })
                
                customer_summary = {
                    'total_orders': total_orders,
                    'total_spent': total_spent,
                    'total_paid': total_paid,
                    'total_balance': total_balance,
                    'completed_orders': completed_orders,
                    'pending_orders': pending_orders,
                    'partially_paid_orders': partially_paid_orders,
                    'last_order_date': last_order.order_date if last_order else None,
                    'first_order_date': first_order.order_date if first_order else None,
                    'customer_since': first_order.order_date if first_order else None,
                    'store_breakdown': store_breakdown,
                }
                
                # Get detailed order information with items
                order_details = []
                for order in customer_orders.order_by('-order_date'):
                    items = []
                    for item in order.order_items.all():
                        items.append({
                            'product_name': item.product.name,
                            'quantity': item.quantity,
                            'unit_price': item.unit_price,
                            'variance': item.variance,
                            'line_total': item.line_total,
                        })
                    
                    order_details.append({
                        'order': order,
                        'items': items,
                        'item_count': len(items),
                        'balance': order.total_amount - order.amount_paid,
                    })
                
                # Add to customer data list
                all_customer_data.append({
                    'customer': customer,
                    'summary': customer_summary,
                    'order_details': order_details,
                })
                
                # Update combined summary
                combined_summary['total_orders'] += total_orders
                combined_summary['total_spent'] += total_spent
                combined_summary['total_paid'] += total_paid
                combined_summary['total_balance'] += total_balance
                combined_summary['completed_orders'] += completed_orders
                combined_summary['pending_orders'] += pending_orders
                combined_summary['partially_paid_orders'] += partially_paid_orders
        
        except Customer.DoesNotExist:
            selected_customers = []
    
    # Handle PDF download
    if 'download' in request.GET and selected_customers:
        return generate_multi_customer_statement_pdf(
            all_customer_data=all_customer_data,
            combined_summary=combined_summary,
            start_date=start_date,
            end_date=end_date,
        )
    
    context = {
        'customers': customers,
        'selected_customers': selected_customers,
        'all_customer_data': all_customer_data,
        'combined_summary': combined_summary,
        'paid_status_choices': Order.PAID_STATUS_CHOICES,
        'store_choices': Order.STORE_CHOICES,
        'selected_customer_ids': selected_customer_ids,
        'has_multiple_customers': len(selected_customers) > 1,
    }
    
    return render(request, 'reports/customer_statements.html', context)


def generate_multi_customer_statement_pdf(all_customer_data, combined_summary, start_date, end_date):
    """
    Generate a comprehensive PDF for multiple customers with individual breakdowns.
    """
    buffer = BytesIO()
    
    # Create PDF with custom canvas - MINIMIZED MARGINS
    doc = SimpleDocTemplate(
        buffer,
        pagesize=letter,
        rightMargin=30,  # Reduced from 40
        leftMargin=30,   # Reduced from 40
        topMargin=35,    # Reduced from 45 (header is now 30px + 5px margin)
        bottomMargin=27, # Reduced from 35 (footer is now 22px + 5px margin)
    )
    
    elements = []
    styles = getSampleStyleSheet()
    
    # Custom styles
    title_style = ParagraphStyle(
        'CompactTitle',
        parent=styles['Heading1'],
        fontSize=14,  # Reduced from 20
        textColor=HexColor('#1a365d'),
        spaceAfter=4,  # Reduced from 10
        spaceBefore=0,
        alignment=TA_CENTER,
        fontName='Helvetica-Bold'
    )
    
    section_heading_style = ParagraphStyle(
        'CompactSectionHeading',
        parent=styles['Heading2'],
        fontSize=9,  # Reduced from 12
        textColor=HexColor('#2d3748'),
        spaceAfter=3,  # Reduced from 6
        spaceBefore=5,  # Reduced from 10
        fontName='Helvetica-Bold',
        borderPadding=(0, 0, 2, 0),  # Reduced from 3
        borderColor=HexColor('#38a169'),
        borderWidth=1.5,  # Reduced from 2
    )
    
    customer_heading_style = ParagraphStyle(
        'CustomerHeading',
        parent=styles['Heading2'],
        fontSize=10,  # Reduced from 14
        textColor=HexColor('#1a365d'),
        fontName='Helvetica-Bold',
        spaceAfter=4,  # Reduced from 8
        spaceBefore=6,  # Reduced from 12
        leftIndent=0,
        backColor=HexColor('#edf2f7'),
        borderPadding=4,  # Reduced from 8
    )
    
    order_heading_style = ParagraphStyle(
        'CompactOrderHeading',
        parent=styles['Heading3'],
        fontSize=8,  # Reduced from 10
        textColor=HexColor('#1a365d'),
        fontName='Helvetica-Bold',
        spaceAfter=2,  # Reduced from 4
        spaceBefore=4,  # Reduced from 8
    )
    
    # === TITLE ===
    if len(all_customer_data) > 1:
        title = Paragraph("MULTI-CUSTOMER ACCOUNT STATEMENT", title_style)
    else:
        title = Paragraph("CUSTOMER ACCOUNT STATEMENT", title_style)
    elements.append(title)
    elements.append(Spacer(1, 3))  # Reduced from 8
    
    # === COMBINED SUMMARY (if multiple customers) ===
    if len(all_customer_data) > 1:
        elements.append(Paragraph("Combined Summary - All Customers", section_heading_style))
        elements.append(Spacer(1, 2))  # Reduced from 4
        
        summary_data = [
            ['Total Customers:', str(len(all_customer_data))],
            ['Total Orders:', str(combined_summary['total_orders'])],
            ['Completed:', str(combined_summary['completed_orders'])],
            ['Pending:', str(combined_summary['pending_orders'])],
            ['Partial:', str(combined_summary['partially_paid_orders'])],
            ['Total Amount:', f"Ksh {combined_summary['total_spent']:,.2f}"],
            ['Total Paid:', f"Ksh {combined_summary['total_paid']:,.2f}"],
            ['Outstanding Balance:', f"Ksh {combined_summary['total_balance']:,.2f}"],
        ]
        
        if start_date or end_date:
            period = f"{start_date or 'Start'} to {end_date or 'Present'}"
            summary_data.insert(0, ['Period:', period])
        
        summary_table = Table(summary_data, colWidths=[2*inch, 3.5*inch])
        summary_table.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ('FONTNAME', (1, 0), (1, -2), 'Helvetica'),
            ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 7),  # Reduced from 9
            ('TEXTCOLOR', (0, 0), (-1, -1), HexColor('#2d3748')),
            ('BACKGROUND', (0, 0), (0, -1), HexColor('#f7fafc')),
            ('BACKGROUND', (0, -1), (-1, -1), 
             HexColor('#feb2b2') if combined_summary['total_balance'] > 0 else HexColor('#c6f6d5')),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
            ('LEFTPADDING', (0, 0), (-1, -1), 4),  # Reduced from 8
            ('RIGHTPADDING', (0, 0), (-1, -1), 4),  # Reduced from 8
            ('TOPPADDING', (0, 0), (-1, -1), 2),  # Reduced from 5
            ('BOTTOMPADDING', (0, 0), (-1, -1), 2),  # Reduced from 5
            ('GRID', (0, 0), (-1, -1), 0.5, HexColor('#e2e8f0')),
        ]))
        
        elements.append(summary_table)
        elements.append(Spacer(1, 6))  # Reduced from 15
        elements.append(PageBreak())
    
    # === INDIVIDUAL CUSTOMER SECTIONS ===
    for idx, customer_data in enumerate(all_customer_data, 1):
        customer = customer_data['customer']
        summary = customer_data['summary']
        order_details = customer_data['order_details']
        
        # Customer Header
        customer_title = f"Customer {idx}: {customer.first_name} {customer.last_name or ''}"
        elements.append(Paragraph(customer_title, customer_heading_style))
        elements.append(Spacer(1, 3))  # Reduced from 6
        
        # Customer Info & Summary side by side
        customer_info = [
            ['Phone:', customer.phone_number or 'N/A'],
            ['Email:', customer.email or 'N/A'],
            ['Category:', customer.get_default_category_display()],
            ['Since:', summary.get('customer_since').strftime('%b %d, %Y') if summary.get('customer_since') else 'N/A'],
        ]
        
        customer_table = Table(customer_info, colWidths=[1*inch, 2*inch])
        customer_table.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 7),  # Reduced from 8
            ('BACKGROUND', (0, 0), (0, -1), HexColor('#f7fafc')),
            ('GRID', (0, 0), (-1, -1), 0.5, HexColor('#e2e8f0')),
            ('LEFTPADDING', (0, 0), (-1, -1), 3),  # Reduced from 5
            ('TOPPADDING', (0, 0), (-1, -1), 2),  # Reduced from 3
            ('BOTTOMPADDING', (0, 0), (-1, -1), 2),  # Reduced from 3
        ]))
        
        summary_info = [
            ['Total Orders:', str(summary['total_orders'])],
            ['Completed:', str(summary['completed_orders'])],
            ['Pending:', str(summary['pending_orders'])],
            ['Partial:', str(summary['partially_paid_orders'])],
            ['Total:', f"Ksh {summary['total_spent']:,.2f}"],
            ['Paid:', f"Ksh {summary['total_paid']:,.2f}"],
            ['Balance:', f"Ksh {summary['total_balance']:,.2f}"],
        ]
        
        summary_table = Table(summary_info, colWidths=[0.9*inch, 1.4*inch])
        summary_table.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 7),  # Reduced from 8
            ('BACKGROUND', (0, 0), (0, -1), HexColor('#f7fafc')),
            ('BACKGROUND', (0, -1), (-1, -1), 
             HexColor('#feb2b2') if summary['total_balance'] > 0 else HexColor('#c6f6d5')),
            ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
            ('GRID', (0, 0), (-1, -1), 0.5, HexColor('#e2e8f0')),
            ('LEFTPADDING', (0, 0), (-1, -1), 3),  # Reduced from 5
            ('TOPPADDING', (0, 0), (-1, -1), 2),  # Reduced from 3
            ('BOTTOMPADDING', (0, 0), (-1, -1), 2),  # Reduced from 3
        ]))
        
        combined_table = Table([[customer_table, summary_table]], colWidths=[3*inch, 2.3*inch])
        combined_table.setStyle(TableStyle([
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ('LEFTPADDING', (0, 0), (-1, -1), 0),
            ('RIGHTPADDING', (0, 0), (-1, -1), 10),
        ]))
        
        elements.append(combined_table)
        elements.append(Spacer(1, 4))  # Reduced from 10
        
        # Store Breakdown (if multiple stores)
        if summary['store_breakdown'] and len(summary['store_breakdown']) > 1:
            elements.append(Paragraph("Orders by Store", ParagraphStyle(
                'StoreBreakdown',
                parent=styles['Heading3'],
                fontSize=10,
                textColor=HexColor('#2d3748'),
                fontName='Helvetica-Bold',
                spaceAfter=4,
            )))
            
            store_data = [['Store', 'Orders', 'Total', 'Paid', 'Balance']]
            for store in summary['store_breakdown']:
                store_data.append([
                    store['store_name'],
                    str(store['order_count']),
                    f"Ksh {store['total_amount']:,.2f}",
                    f"Ksh {store['amount_paid']:,.2f}",
                    f"Ksh {store['balance']:,.2f}",
                ])
            
            store_table = Table(store_data, colWidths=[1.5*inch, 0.6*inch, 1*inch, 1*inch, 1*inch])
            store_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), HexColor('#1a365d')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 7),  # Reduced from 8
                ('ALIGN', (1, 0), (-1, -1), 'RIGHT'),
                ('GRID', (0, 0), (-1, -1), 0.5, HexColor('#e2e8f0')),
                ('LEFTPADDING', (0, 0), (-1, -1), 3),  # Reduced from 5
                ('TOPPADDING', (0, 0), (-1, -1), 2),  # Reduced from 3
                ('BOTTOMPADDING', (0, 0), (-1, -1), 2),  # Reduced from 3
            ]))
            
            elements.append(store_table)
            elements.append(Spacer(1, 4))  # Reduced from 10
        
        # Order History
        elements.append(Paragraph("Order History", section_heading_style))
        elements.append(Spacer(1, 2))  # Reduced from 4
        
        for order_idx, detail in enumerate(order_details, 1):
            order = detail['order']
            items = detail['items']
            
            # Order header
            order_header_text = (
                f"<b>Order #Mc{order.id}Z</b> | "
                f"{order.order_date.strftime('%b %d, %Y %I:%M %p')} | "
                f"{order.get_paid_status_display()} | "
                f"{order.get_store_display()}"
            )
            
            order_header = Paragraph(order_header_text, order_heading_style)
            elements.append(order_header)
            
            # Items table
            items_data = [['#', 'Product', 'Qty', 'Price', 'Var', 'Total']]
            
            for i, item in enumerate(items, 1):
                items_data.append([
                    str(i),
                    item['product_name'][:35] + '...' if len(item['product_name']) > 35 else item['product_name'],
                    str(item['quantity']),
                    f"{item['unit_price']:,.0f}",
                    f"{item['variance']:+,.0f}" if item['variance'] != 0 else '-',
                    f"{item['line_total']:,.2f}",
                ])
            
            # Totals
            items_data.append(['', '', '', '', 'Total:', f"{order.total_amount:,.2f}"])
            items_data.append(['', '', '', '', 'Paid:', f"{order.amount_paid:,.2f}"])
            items_data.append(['', '', '', '', 'Balance:', f"{detail['balance']:,.2f}"])
            
            items_table = Table(
                items_data, 
                colWidths=[0.3*inch, 2.8*inch, 0.4*inch, 0.7*inch, 0.6*inch, 0.8*inch]
            )
            items_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), HexColor('#1a365d')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 7),  # Reduced from 8
                ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
                ('FONTNAME', (0, 1), (-1, -4), 'Helvetica'),
                ('FONTSIZE', (0, 1), (-1, -4), 6),  # Reduced from 7
                ('ALIGN', (0, 1), (0, -4), 'CENTER'),
                ('ALIGN', (2, 1), (2, -4), 'CENTER'),
                ('ALIGN', (3, 1), (-1, -1), 'RIGHT'),
                ('FONTNAME', (4, -3), (-1, -1), 'Helvetica-Bold'),
                ('FONTSIZE', (4, -3), (-1, -1), 7),  # Reduced from 8
                ('BACKGROUND', (4, -3), (-1, -3), HexColor('#f7fafc')),
                ('BACKGROUND', (4, -2), (-1, -2), HexColor('#f7fafc')),
                ('BACKGROUND', (4, -1), (-1, -1), 
                 HexColor('#fed7d7') if detail['balance'] > 0 else HexColor('#c6f6d5')),
                ('GRID', (0, 0), (-1, -4), 0.5, HexColor('#e2e8f0')),
                ('LINEABOVE', (4, -3), (-1, -3), 1, HexColor('#1a365d')),
                ('LINEABOVE', (4, -1), (-1, -1), 1.5, HexColor('#1a365d')),
                ('LEFTPADDING', (0, 0), (-1, -1), 2),  # Reduced from 4
                ('RIGHTPADDING', (0, 0), (-1, -1), 2),  # Reduced from 4
                ('TOPPADDING', (0, 0), (-1, -1), 2),  # Reduced from 3
                ('BOTTOMPADDING', (0, 0), (-1, -1), 2),  # Reduced from 3
            ]))
            
            elements.append(items_table)
            elements.append(Spacer(1, 3))  # Reduced from 6
            
            # Page break after every 4 orders (increased from 3 to fit more)
            if order_idx % 4 == 0 and order_idx < len(order_details):
                elements.append(PageBreak())
        
        # Page break between customers (except last one)
        if idx < len(all_customer_data):
            elements.append(PageBreak())
    
    # Footer note
    elements.append(Spacer(1, 4))  # Reduced from 10
    footer_note = Paragraph(
        "<b>Note:</b> This statement is a comprehensive record of all transactions. "
        "Please verify amounts and report discrepancies within 7 days.",
        ParagraphStyle(
            'FooterNote',
            parent=styles['Normal'],
            fontSize=6,  # Reduced from 7
            textColor=HexColor('#718096'),
        )
    )
    elements.append(footer_note)
    
    # Build PDF
    doc.build(elements, canvasmaker=CustomerStatementCanvas)
    
    # Return response
    buffer.seek(0)
    response = HttpResponse(buffer, content_type='application/pdf')
    
    if len(all_customer_data) > 1:
        filename = f"Multi_Customer_Statement_{timezone.now().strftime('%Y%m%d')}.pdf"
    else:
        customer = all_customer_data[0]['customer']
        filename = f"Statement_{customer.first_name}_{customer.last_name or ''}_{timezone.now().strftime('%Y%m%d')}.pdf"
    
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    return response

# =====================================================================
# ===  CUSTOMER FEEDBACK VIEWS  =======================================
# =====================================================================

import base64 as _b64
from django.core.files.base import ContentFile as _ContentFile


@login_required
def feedback_list(request):
    """List all customer feedbacks with filter, ordering and pagination."""
    from django.core.paginator import Paginator
    from django.contrib.auth import get_user_model as _gum
    user = request.user
    is_admin = user.is_superuser or user.groups.filter(name='Admins').exists()

    if is_admin:
        qs = CustomerFeedback.objects.select_related('customer', 'salesperson').all()
    else:
        qs = CustomerFeedback.objects.select_related('customer', 'salesperson').filter(salesperson=user)

    # --- Filters ---
    fb_type   = request.GET.get('type', '')
    rating    = request.GET.get('rating', '')
    sp_id     = request.GET.get('sp', '')
    search    = request.GET.get('q', '')
    date_from = request.GET.get('date_from', '')
    date_to   = request.GET.get('date_to', '')
    ordering  = request.GET.get('order', '-created_at')

    if fb_type:
        qs = qs.filter(feedback_type=fb_type)
    if rating:
        qs = qs.filter(rating=rating)
    if is_admin and sp_id:
        qs = qs.filter(salesperson_id=sp_id)
    if search:
        qs = qs.filter(
            Q(shop_name__icontains=search) | Q(customer__first_name__icontains=search) |
            Q(customer__last_name__icontains=search) | Q(comment__icontains=search)
        )
    if date_from:
        try:
            from datetime import datetime as _dt
            qs = qs.filter(created_at__date__gte=_dt.strptime(date_from, '%Y-%m-%d').date())
        except ValueError:
            pass
    if date_to:
        try:
            from datetime import datetime as _dt
            qs = qs.filter(created_at__date__lte=_dt.strptime(date_to, '%Y-%m-%d').date())
        except ValueError:
            pass

    valid_orders = {'-created_at', 'created_at', 'rating', '-rating', 'shop_name', 'feedback_type'}
    if ordering not in valid_orders:
        ordering = '-created_at'
    qs = qs.order_by(ordering)

    total = qs.count()
    paginator = Paginator(qs, 20)
    page_obj  = paginator.get_page(request.GET.get('page'))
    salespersons = _gum().objects.filter(is_active=True).order_by('first_name') if is_admin else None

    context = {
        'feedbacks': page_obj,
        'page_obj': page_obj,
        'total': total,
        'is_admin': is_admin,
        'salespersons': salespersons,
        # filter values
        'fb_type': fb_type,
        'rating': rating,
        'sp_id': sp_id,
        'search': search,
        'date_from': date_from,
        'date_to': date_to,
        'ordering': ordering,
        'feedback_types': CustomerFeedback.FEEDBACK_TYPE_CHOICES,
    }
    return render(request, 'feedback/list.html', context)


@login_required
def add_feedback(request):
    """Create a new customer feedback entry with camera photo."""
    if request.method == 'POST':
        form = CustomerFeedbackForm(request.POST, request.FILES, user=request.user)
        if form.is_valid():
            feedback = form.save(commit=False)
            feedback.salesperson = request.user

            # Auto-fill shop details from the selected customer
            customer = feedback.customer
            if not feedback.shop_name:
                feedback.shop_name = customer.first_name
            if not feedback.contact_person:
                feedback.contact_person = customer.last_name
            if not feedback.phone_number:
                feedback.phone_number = customer.phone_number

            # Handle base64 camera photo (sent from JS watermark canvas)
            photo_data = request.POST.get('photo_base64', '')
            if photo_data and photo_data.startswith('data:image'):
                try:
                    header, encoded = photo_data.split(',', 1)
                    image_bytes = _b64.b64decode(encoded)
                    from django.utils import timezone as _tz
                    fname = f"feedback_{request.user.id}_{_tz.now().strftime('%Y%m%d%H%M%S')}.jpg"
                    feedback.photo.save(fname, _ContentFile(image_bytes), save=False)
                except Exception:
                    pass  # photo is optional; don't block save

            feedback.save()
            messages.success(request, 'Feedback submitted successfully!')
            return redirect('feedback_list')
        else:
            messages.error(request, 'Please fix the errors below.')
    else:
        # Pre-populate customer info if customer_id provided in query string
        initial = {}
        cid = request.GET.get('customer_id')
        if cid:
            try:
                cust = Customer.objects.get(pk=cid)
                initial = {
                    'customer': cust,
                    'shop_name': cust.first_name,
                    'contact_person': cust.last_name,
                    'phone_number': cust.phone_number,
                }
            except Customer.DoesNotExist:
                pass
        form = CustomerFeedbackForm(initial=initial, user=request.user)

    return render(request, 'feedback/add_feedback.html', {'form': form})


@login_required
def feedback_detail(request, pk):
    """View a single feedback entry."""
    feedback = get_object_or_404(CustomerFeedback, pk=pk)
    if not (request.user.is_superuser or
            request.user.groups.filter(name='Admins').exists() or
            feedback.salesperson == request.user):
        messages.error(request, 'You do not have permission to view this feedback.')
        return redirect('feedback_list')
    return render(request, 'feedback/detail.html', {'feedback': feedback})


@login_required
def get_customer_for_feedback(request):
    """AJAX: return customer details for feedback form auto-fill."""
    customer_id = request.GET.get('customer_id')
    if not customer_id:
        return JsonResponse({'error': 'No customer_id'}, status=400)
    try:
        cust = Customer.objects.get(pk=customer_id)
        return JsonResponse({
            'shop_name': cust.first_name or '',
            'contact_person': cust.last_name or '',
            'phone_number': cust.phone_number or '',
            'address': cust.address or '',
        })
    except Customer.DoesNotExist:
        return JsonResponse({'error': 'Not found'}, status=404)


@login_required
def salesperson_feedback_metrics(request):
    """Admin view: feedback metrics per salesperson."""
    if not (request.user.is_superuser or request.user.groups.filter(name='Admins').exists()):
        messages.error(request, 'Admins only.')
        return redirect('feedback_list')

    from django.db.models import Avg, Count, Q
    from django.contrib.auth import get_user_model
    _User = get_user_model()

    # All salespersons who have collected at least one feedback
    salespersons = _User.objects.filter(
        feedbacks_collected__isnull=False
    ).distinct().order_by('first_name', 'username')

    rows = []
    for sp in salespersons:
        qs = CustomerFeedback.objects.filter(salesperson=sp)
        agg = qs.aggregate(
            total=Count('id'),
            avg_rating=Avg('rating'),
            quality=Count('id', filter=Q(feedback_type='quality')),
            pricing=Count('id', filter=Q(feedback_type='pricing')),
            payments=Count('id', filter=Q(feedback_type='payments')),
            delivery_time=Count('id', filter=Q(feedback_type='delivery_time')),
            rating_5=Count('id', filter=Q(rating=5)),
            rating_4=Count('id', filter=Q(rating=4)),
            rating_3=Count('id', filter=Q(rating=3)),
            rating_2=Count('id', filter=Q(rating=2)),
            rating_1=Count('id', filter=Q(rating=1)),
        )
        recent = qs.select_related('customer').order_by('-created_at')[:3]
        rows.append({
            'salesperson': sp,
            'total': agg['total'],
            'avg_rating': round(agg['avg_rating'] or 0, 1),
            'quality': agg['quality'],
            'pricing': agg['pricing'],
            'payments': agg['payments'],
            'delivery_time': agg['delivery_time'],
            'rating_5': agg['rating_5'],
            'rating_4': agg['rating_4'],
            'rating_3': agg['rating_3'],
            'rating_2': agg['rating_2'],
            'rating_1': agg['rating_1'],
            'recent': recent,
        })

    # Overall totals
    overall = CustomerFeedback.objects.aggregate(
        total=Count('id'),
        avg_rating=Avg('rating'),
        quality=Count('id', filter=Q(feedback_type='quality')),
        pricing=Count('id', filter=Q(feedback_type='pricing')),
        payments=Count('id', filter=Q(feedback_type='payments')),
        delivery_time=Count('id', filter=Q(feedback_type='delivery_time')),
    )
    overall['avg_rating'] = round(overall['avg_rating'] or 0, 1)

    context = {
        'rows': rows,
        'overall': overall,
    }
    return render(request, 'feedback/salesperson_metrics.html', context)


# =====================================================================
# ===  INTERNAL MINI-CHAT VIEWS  =====================================
# =====================================================================

@login_required
def internal_chat(request):
    """Main chat interface — loads last 50 messages visible to this user."""
    user = request.user
    # Visible messages: sent by anyone to this user (or broadcast) or sent by this user
    messages_qs = InternalMessage.objects.select_related('sender', 'recipient', 'feedback').filter(
        Q(recipient=user) | Q(recipient__isnull=True) | Q(sender=user)
    ).order_by('-created_at')[:50]

    messages_list = list(reversed(list(messages_qs)))

    # Mark all unread as read for this user
    InternalMessage.objects.filter(
        Q(recipient=user) | Q(recipient__isnull=True),
        is_read=False
    ).exclude(sender=user).update(is_read=True)

    users = User.objects.filter(is_active=True).exclude(pk=user.pk).order_by('first_name', 'username')
    form = InternalMessageForm()

    context = {
        'messages_list': messages_list,
        'form': form,
        'users': users,
        'current_user': user,
    }
    return render(request, 'chat/internal_chat.html', context)


@login_required
def send_internal_message(request):
    """AJAX POST: save a new message (text, file, image, location, contact) and return it as JSON."""
    if request.method != 'POST':
        return JsonResponse({'error': 'POST required'}, status=405)

    from django.contrib.auth import get_user_model as _gum
    user = request.user
    recipient_id = request.POST.get('recipient') or None
    recipient = None
    if recipient_id:
        try:
            recipient = _gum().objects.get(pk=recipient_id)
        except Exception:
            pass

    attach_type = request.POST.get('attach_type', 'text')
    msg = InternalMessage(sender=user, recipient=recipient, message_type='user', attach_type=attach_type)

    if attach_type == 'text':
        text = request.POST.get('message', '').strip()
        if not text:
            return JsonResponse({'error': 'Empty message'}, status=400)
        msg.message = text

    elif attach_type in ('file', 'image'):
        f = request.FILES.get('attachment')
        if not f:
            return JsonResponse({'error': 'No file uploaded'}, status=400)
        msg.attachment = f
        msg.attachment_name = f.name
        msg.message = request.POST.get('message', '').strip()

    elif attach_type == 'location':
        lat = request.POST.get('latitude', '').strip()
        lng = request.POST.get('longitude', '').strip()
        label = request.POST.get('location_label', '').strip()
        if not lat or not lng:
            return JsonResponse({'error': 'Location coordinates required'}, status=400)
        from decimal import Decimal as _D, InvalidOperation
        try:
            msg.latitude  = _D(lat)
            msg.longitude = _D(lng)
        except InvalidOperation:
            return JsonResponse({'error': 'Invalid coordinates'}, status=400)
        msg.location_label = label
        msg.message = label or f'{lat},{lng}'

    elif attach_type == 'contact':
        msg.contact_name  = request.POST.get('contact_name', '').strip()
        msg.contact_phone = request.POST.get('contact_phone', '').strip()
        if not msg.contact_name or not msg.contact_phone:
            return JsonResponse({'error': 'Contact name and phone required'}, status=400)
        msg.message = f'{msg.contact_name} — {msg.contact_phone}'

    else:
        return JsonResponse({'error': 'Unknown attach_type'}, status=400)

    msg.save()

    attachment_url = msg.attachment.url if msg.attachment else None
    return JsonResponse({
        'id': msg.id,
        'sender': user.get_full_name() or user.username,
        'sender_id': user.id,
        'recipient': recipient.get_full_name() or recipient.username if recipient else None,
        'message': msg.message,
        'message_type': msg.message_type,
        'attach_type': msg.attach_type,
        'attachment_url': attachment_url,
        'attachment_name': msg.attachment_name,
        'latitude': str(msg.latitude) if msg.latitude else None,
        'longitude': str(msg.longitude) if msg.longitude else None,
        'location_label': msg.location_label,
        'contact_name': msg.contact_name,
        'contact_phone': msg.contact_phone,
        'created_at': msg.created_at.strftime('%H:%M'),
        'created_ts': msg.created_at.timestamp(),
        'is_broadcast': recipient is None,
    })


@login_required
def poll_messages(request):
    """
    AJAX GET: return messages newer than `since` (Unix timestamp).
    Frontend polls this every 5 seconds.
    """
    since_ts = request.GET.get('since', 0)
    try:
        since_ts = float(since_ts)
    except (ValueError, TypeError):
        since_ts = 0

    from datetime import datetime as _dt
    import datetime
    _utc = datetime.timezone.utc
    since_dt = _dt.fromtimestamp(since_ts, tz=_utc) if since_ts else None

    user = request.user
    qs = InternalMessage.objects.select_related('sender', 'recipient', 'feedback').filter(
        Q(recipient=user) | Q(recipient__isnull=True) | Q(sender=user)
    )
    if since_dt:
        qs = qs.filter(created_at__gt=since_dt)

    qs = qs.order_by('created_at')[:50]

    data = []
    for msg in qs:
        data.append({
            'id': msg.id,
            'sender': msg.sender.get_full_name() or msg.sender.username if msg.sender else 'System',
            'sender_id': msg.sender.id if msg.sender else None,
            'recipient': msg.recipient.get_full_name() or msg.recipient.username if msg.recipient else None,
            'message': msg.message,
            'message_type': msg.message_type,
            'attach_type': msg.attach_type,
            'attachment_url': msg.attachment.url if msg.attachment else None,
            'attachment_name': msg.attachment_name,
            'latitude': str(msg.latitude) if msg.latitude else None,
            'longitude': str(msg.longitude) if msg.longitude else None,
            'location_label': msg.location_label,
            'contact_name': msg.contact_name,
            'contact_phone': msg.contact_phone,
            'created_at': msg.created_at.strftime('%H:%M'),
            'created_ts': msg.created_at.timestamp(),
            'is_broadcast': msg.recipient is None,
            'feedback_id': msg.feedback.id if msg.feedback else None,
        })

    # Unread count for badge
    unread_count = InternalMessage.objects.filter(
        Q(recipient=user) | Q(recipient__isnull=True),
        is_read=False
    ).exclude(sender=user).count()

    return JsonResponse({'messages': data, 'unread_count': unread_count})


@login_required
def unread_count_api(request):
    """AJAX: return just the unread message count for the navbar badge."""
    user = request.user
    count = InternalMessage.objects.filter(
        Q(recipient=user) | Q(recipient__isnull=True),
        is_read=False
    ).exclude(sender=user).count()
    return JsonResponse({'unread_count': count})


# =====================================================================
# ===  NOTIFICATION VIEWS  ============================================
# =====================================================================

from .models import Notification

@login_required
def notifications_json(request):
    """AJAX: return recent unread notifications as JSON for the bell dropdown."""
    user = request.user
    notifs = Notification.objects.filter(user=user).order_by('-created_at')[:20]
    unread_count = Notification.objects.filter(user=user, is_read=False).count()
    data = [
        {
            'id': n.id,
            'event_type': n.event_type,
            'icon': n.icon,
            'title': n.title,
            'body': n.body,
            'url': n.url,
            'is_read': n.is_read,
            'created_at': n.created_at.strftime('%d %b %H:%M'),
        }
        for n in notifs
    ]
    return JsonResponse({'notifications': data, 'unread_count': unread_count})


@login_required
def mark_notifications_read(request):
    """AJAX POST: mark all or specific notification(s) as read."""
    if request.method != 'POST':
        return JsonResponse({'error': 'POST required'}, status=405)
    user  = request.user
    n_id  = request.POST.get('id')
    if n_id:
        Notification.objects.filter(user=user, pk=n_id).update(is_read=True)
    else:
        Notification.objects.filter(user=user, is_read=False).update(is_read=True)
    unread = Notification.objects.filter(user=user, is_read=False).count()
    return JsonResponse({'ok': True, 'unread_count': unread})


@login_required
def notification_list(request):
    """Full notifications page with filter and pagination."""
    from django.core.paginator import Paginator
    user = request.user
    qs   = Notification.objects.filter(user=user).order_by('-created_at')

    event_filter = request.GET.get('event', '')
    read_filter  = request.GET.get('read', '')
    if event_filter:
        qs = qs.filter(event_type=event_filter)
    if read_filter == '0':
        qs = qs.filter(is_read=False)
    elif read_filter == '1':
        qs = qs.filter(is_read=True)

    # Mark as read on open (only unread ones)
    qs.filter(is_read=False).update(is_read=True)

    paginator = Paginator(qs, 30)
    page_obj  = paginator.get_page(request.GET.get('page'))
    unread    = Notification.objects.filter(user=user, is_read=False).count()

    context = {
        'notifications': page_obj,
        'page_obj': page_obj,
        'event_choices': Notification.EVENT_TYPE_CHOICES,
        'event_filter': event_filter,
        'read_filter': read_filter,
        'unread': unread,
    }
    return render(request, 'notifications/list.html', context)


# =====================================================================
# ===  M-PESA STK PUSH VIEWS  =========================================
# =====================================================================

from .mpesa import initiate_stk_push, query_stk_status
from .forms import MPesaSTKForm


@login_required
def mpesa_stk_push(request, order_id):
    """
    Show STK Push form and initiate payment when submitted.
    Supports partial payments — salesperson enters the amount.
    """
    order = get_object_or_404(Order, pk=order_id)
    remaining = order.get_balance()

    if request.method == 'POST':
        form = MPesaSTKForm(request.POST)
        if form.is_valid():
            phone = form.cleaned_data['phone_number']
            amount = form.cleaned_data['amount']

            if amount > remaining:
                messages.error(request, f'Amount (KSh {amount}) exceeds remaining balance (KSh {remaining}).')
                return render(request, 'payments/mpesa_stk.html', {
                    'form': form, 'order': order, 'remaining': remaining,
                })

            try:
                response = initiate_stk_push(phone, amount, order.id)
                if response.get('ResponseCode') == '0':
                    txn = MPesaTransaction.objects.create(
                        order=order,
                        amount=amount,
                        phone_number=phone,
                        checkout_request_id=response.get('CheckoutRequestID'),
                        merchant_request_id=response.get('MerchantRequestID'),
                        status='pending',
                        initiated_by=request.user,
                    )
                    messages.success(
                        request,
                        f'STK Push sent to {phone}. Ask the customer to enter their M-Pesa PIN.'
                    )
                    return render(request, 'payments/mpesa_pending.html', {
                        'txn': txn,
                        'order': order,
                        'phone': phone,
                        'amount': amount,
                    })
                else:
                    messages.error(request, f"M-Pesa error: {response.get('ResponseDescription', 'Unknown error')}")
            except Exception as e:
                messages.error(request, f'Failed to initiate M-Pesa payment: {e}')
        else:
            messages.error(request, 'Please correct the errors below.')
    else:
        initial_phone = order.phone or (order.customer.phone_number or '')
        form = MPesaSTKForm(initial={'phone_number': initial_phone, 'amount': remaining})

    return render(request, 'payments/mpesa_stk.html', {
        'form': form,
        'order': order,
        'remaining': remaining,
    })


@login_required
def mpesa_check_status(request, txn_id):
    """AJAX: poll transaction status (used by the pending payment page)."""
    txn = get_object_or_404(MPesaTransaction, pk=txn_id)
    if txn.status == 'pending':
        try:
            result = query_stk_status(txn.checkout_request_id)
            result_code = str(result.get('ResultCode', ''))
            if result_code == '0':
                txn.status = 'success'
                txn.result_code = result_code
                txn.result_description = result.get('ResultDesc', '')
                txn.save()
                Payment.objects.create(
                    order=txn.order,
                    amount=txn.amount,
                    payment_method='mpesa',
                    reference_number=txn.mpesa_receipt_number,
                    notes=f'M-Pesa STK Push — TxnID #{txn.id}',
                    recorded_by=txn.initiated_by,
                )
                txn.order.amount_paid = (txn.order.amount_paid or 0) + txn.amount
                txn.order.update_paid_status()
            elif result_code != '':
                txn.status = 'failed'
                txn.result_code = result_code
                txn.result_description = result.get('ResultDesc', 'Payment failed or cancelled')
                txn.save()
        except Exception:
            pass

    return JsonResponse({
        'status': txn.status,
        'receipt': txn.mpesa_receipt_number or '',
        'description': txn.result_description or '',
    })


from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_POST


@csrf_exempt
@require_POST
def mpesa_callback(request):
    """
    Daraja sends payment results here via POST (JSON body).
    No login required — secured by Daraja's IP only.
    """
    try:
        body = json.loads(request.body)
        callback = body.get('Body', {}).get('stkCallback', {})
        checkout_request_id = callback.get('CheckoutRequestID')
        result_code = str(callback.get('ResultCode', ''))
        result_desc = callback.get('ResultDesc', '')

        txn = MPesaTransaction.objects.filter(checkout_request_id=checkout_request_id).first()
        if not txn:
            return JsonResponse({'ResultCode': 0, 'ResultDesc': 'Accepted'})

        if result_code == '0':
            metadata = callback.get('CallbackMetadata', {}).get('Item', [])
            receipt = next(
                (item['Value'] for item in metadata if item.get('Name') == 'MpesaReceiptNumber'), ''
            )
            txn.status = 'success'
            txn.mpesa_receipt_number = receipt
            txn.result_code = result_code
            txn.result_description = result_desc
            txn.save()

            Payment.objects.create(
                order=txn.order,
                amount=txn.amount,
                payment_method='mpesa',
                reference_number=receipt,
                notes=f'M-Pesa callback — {checkout_request_id}',
                recorded_by=txn.initiated_by,
            )
            txn.order.amount_paid = (txn.order.amount_paid or 0) + txn.amount
            txn.order.update_paid_status()
        else:
            txn.status = 'failed' if result_code != '1032' else 'cancelled'
            txn.result_code = result_code
            txn.result_description = result_desc
            txn.save()

    except Exception as e:
        import logging
        logging.getLogger(__name__).error(f"M-Pesa callback error: {e}")

    return JsonResponse({'ResultCode': 0, 'ResultDesc': 'Accepted'})


# =====================================================================
# ===  LOGIN SESSION (LIVE PHOTO + GEOLOCATION)  =====================
# =====================================================================

@login_required
def login_history(request):
    """Show login session history — admins see all, salespersons see own."""
    from django.core.paginator import Paginator
    from django.contrib.auth import get_user_model
    user = request.user
    is_admin = user.is_superuser or user.groups.filter(name='Admins').exists()
    if is_admin:
        qs = LoginSession.objects.select_related('user').order_by('-login_at')
    else:
        qs = LoginSession.objects.filter(user=user).order_by('-login_at')

    # --- Filters ---
    sp_filter  = request.GET.get('user', '')
    date_from  = request.GET.get('date_from', '')
    date_to    = request.GET.get('date_to', '')
    ordering   = request.GET.get('order', '-login_at')

    if is_admin and sp_filter:
        qs = qs.filter(user_id=sp_filter)
    if date_from:
        try:
            from datetime import datetime as _dt
            qs = qs.filter(login_at__date__gte=_dt.strptime(date_from, '%Y-%m-%d').date())
        except ValueError:
            pass
    if date_to:
        try:
            from datetime import datetime as _dt
            qs = qs.filter(login_at__date__lte=_dt.strptime(date_to, '%Y-%m-%d').date())
        except ValueError:
            pass

    valid_orders = {'-login_at', 'login_at', 'user__first_name', '-user__first_name'}
    if ordering not in valid_orders:
        ordering = '-login_at'
    qs = qs.order_by(ordering)

    _User = get_user_model()
    salespersons = _User.objects.filter(is_active=True).order_by('first_name') if is_admin else None
    paginator = Paginator(qs, 25)
    page_obj  = paginator.get_page(request.GET.get('page'))

    context = {
        'sessions': page_obj,
        'page_obj': page_obj,
        'is_admin': is_admin,
        'salespersons': salespersons,
        'sp_filter': sp_filter,
        'date_from': date_from,
        'date_to': date_to,
        'ordering': ordering,
    }
    return render(request, 'auth/login_history.html', context)


@csrf_exempt
@login_required
def save_login_session(request):
    """
    Called via AJAX immediately after login to save live photo + GPS.
    """
    if request.method != 'POST':
        return JsonResponse({'error': 'POST required'}, status=405)

    latitude = request.POST.get('latitude')
    longitude = request.POST.get('longitude')
    photo_data = request.POST.get('photo_base64', '')
    device_info = request.POST.get('device_info', '')

    session = LoginSession(
        user=request.user,
        latitude=latitude or None,
        longitude=longitude or None,
        ip_address=request.META.get('REMOTE_ADDR'),
        device_info=device_info[:500] if device_info else '',
    )

    if photo_data and photo_data.startswith('data:image'):
        try:
            header, encoded = photo_data.split(',', 1)
            image_bytes = _b64.b64decode(encoded)
            from django.utils import timezone as _tz
            fname = f"login_{request.user.id}_{_tz.now().strftime('%Y%m%d%H%M%S')}.jpg"
            session.login_photo.save(fname, _ContentFile(image_bytes), save=False)
        except Exception:
            pass

    session.save()
    return JsonResponse({'status': 'ok'})


# =====================================================================
# ===  DAILY BEAT VIEWS  =============================================
# =====================================================================

@login_required
def beat_today(request):
    """Salesperson's today beat — list of planned customers + log visits."""
    from datetime import date as _date
    today_name = _date.today().strftime('%A')  # e.g. 'Monday'
    user = request.user
    is_admin = user.is_superuser or user.groups.filter(name='Admins').exists()

    # For admins viewing a specific salesperson
    sp_id = request.GET.get('sp')
    if is_admin and sp_id:
        from django.contrib.auth import get_user_model as _gum
        try:
            view_user = _gum().objects.get(pk=sp_id)
        except Exception:
            view_user = user
    else:
        view_user = user

    plans = BeatPlan.objects.filter(
        salesperson=view_user, day_of_week=today_name, is_active=True
    ).select_related('customer')

    # Already visited today
    visited_ids = set(BeatVisit.objects.filter(
        salesperson=view_user, visit_date=_date.today()
    ).values_list('plan_id', flat=True))

    from django.contrib.auth import get_user_model as _gum2
    salespersons = _gum2().objects.filter(is_active=True).order_by('first_name') if is_admin else None

    context = {
        'plans': plans,
        'visited_ids': visited_ids,
        'today': today_name,
        'today_date': _date.today(),
        'is_admin': is_admin,
        'salespersons': salespersons,
        'view_user': view_user,
    }
    return render(request, 'beat/today.html', context)


@login_required
def beat_log_visit(request, plan_id):
    """Log a visit against a beat plan entry."""
    from datetime import date as _date
    plan = get_object_or_404(BeatPlan, pk=plan_id)
    if plan.salesperson != request.user and not (request.user.is_superuser or request.user.groups.filter(name='Admins').exists()):
        messages.error(request, 'Not authorized.')
        return redirect('beat_today')

    if request.method == 'POST':
        outcome = request.POST.get('outcome', 'info_gathered')
        notes   = request.POST.get('notes', '')
        lat     = request.POST.get('latitude') or None
        lng     = request.POST.get('longitude') or None

        visit = BeatVisit(
            plan=plan,
            salesperson=plan.salesperson,
            customer=plan.customer,
            visit_date=_date.today(),
            outcome=outcome,
            notes=notes,
        )
        if lat:
            try: visit.latitude = Decimal(lat)
            except Exception: pass
        if lng:
            try: visit.longitude = Decimal(lng)
            except Exception: pass
        visit.save()
        messages.success(request, f'Visit to {plan.customer} logged!')
        return redirect('beat_today')

    return render(request, 'beat/log_visit.html', {'plan': plan, 'outcomes': BeatVisit.OUTCOME_CHOICES})


@login_required
def beat_plans(request):
    """List & manage beat plans — with filter, ordering and pagination."""
    from django.core.paginator import Paginator
    from django.contrib.auth import get_user_model as _gum
    user = request.user
    is_admin = user.is_superuser or user.groups.filter(name='Admins').exists()

    if is_admin:
        qs = BeatPlan.objects.select_related('salesperson', 'customer').all()
    else:
        qs = BeatPlan.objects.select_related('customer').filter(salesperson=user)

    # --- Filters ---
    day_filter    = request.GET.get('day', '')
    sp_filter     = request.GET.get('sp', '')
    active_filter = request.GET.get('active', '')
    ordering      = request.GET.get('order', 'day_of_week')

    if day_filter:
        qs = qs.filter(day_of_week=day_filter)
    if is_admin and sp_filter:
        qs = qs.filter(salesperson_id=sp_filter)
    if active_filter == '1':
        qs = qs.filter(is_active=True)
    elif active_filter == '0':
        qs = qs.filter(is_active=False)

    valid_orders = {'day_of_week', '-day_of_week', 'customer__first_name', '-customer__first_name',
                    'salesperson__first_name', '-salesperson__first_name'}
    if ordering not in valid_orders:
        ordering = 'day_of_week'
    qs = qs.order_by(ordering)

    salespersons = _gum().objects.filter(is_active=True).order_by('first_name') if is_admin else None
    paginator = Paginator(qs, 25)
    page_obj  = paginator.get_page(request.GET.get('page'))

    context = {
        'plans': page_obj,
        'page_obj': page_obj,
        'is_admin': is_admin,
        'salespersons': salespersons,
        'day_choices': BeatPlan.DAY_CHOICES,
        'day_filter': day_filter,
        'sp_filter': sp_filter,
        'active_filter': active_filter,
        'ordering': ordering,
    }
    return render(request, 'beat/plans.html', context)


@login_required
def beat_plan_create(request):
    """Create a new beat plan entry."""
    from django.contrib.auth import get_user_model as _gum
    user = request.user
    is_admin = user.is_superuser or user.groups.filter(name='Admins').exists()

    if request.method == 'POST':
        sp_id        = request.POST.get('salesperson') if is_admin else user.id
        customer_ids = request.POST.getlist('customer')
        day          = request.POST.get('day_of_week')
        notes        = request.POST.get('notes', '')
        added = skipped = 0
        try:
            sp = _gum().objects.get(pk=sp_id)
            for cid in customer_ids:
                try:
                    cust = Customer.objects.get(pk=cid)
                    _, created = BeatPlan.objects.get_or_create(
                        salesperson=sp, customer=cust, day_of_week=day,
                        defaults={'notes': notes, 'is_active': True}
                    )
                    if created:
                        added += 1
                    else:
                        skipped += 1
                except Customer.DoesNotExist:
                    pass
            if added:
                messages.success(request, f'{added} customer(s) added to {sp.get_full_name() or sp.username}\'s {day} beat.' + (f' {skipped} already existed.' if skipped else ''))
            elif skipped:
                messages.info(request, f'All selected customers are already in the {day} beat.')
        except Exception as e:
            messages.error(request, str(e))
        return redirect('beat_plans')

    salespersons = _gum().objects.filter(is_active=True).order_by('first_name') if is_admin else None
    customers    = Customer.objects.all().order_by('first_name')
    context = {
        'salespersons': salespersons,
        'customers': customers,
        'days': BeatPlan.DAY_CHOICES,
        'is_admin': is_admin,
        'current_user': user,
    }
    return render(request, 'beat/plan_create.html', context)


@login_required
def beat_overview(request):
    """Admin overview — all salespersons' beat visit logs with filter + pagination."""
    if not (request.user.is_superuser or request.user.groups.filter(name='Admins').exists()):
        return redirect('beat_today')
    from datetime import date as _date, datetime as _dt
    from django.contrib.auth import get_user_model as _gum
    from django.db.models import Count
    from django.core.paginator import Paginator

    # --- Filters ---
    date_filter    = request.GET.get('date', str(_date.today()))
    sp_filter      = request.GET.get('sp', '')
    outcome_filter = request.GET.get('outcome', '')
    date_from      = request.GET.get('date_from', '')
    date_to        = request.GET.get('date_to', '')

    try:
        filter_date = _dt.strptime(date_filter, '%Y-%m-%d').date()
    except Exception:
        filter_date = _date.today()

    visits = BeatVisit.objects.select_related('salesperson', 'customer', 'plan').order_by('-visit_date', 'salesperson__first_name')

    # Apply date filter only if no date_from/date_to range is set
    if date_from or date_to:
        if date_from:
            try:
                visits = visits.filter(visit_date__gte=_dt.strptime(date_from, '%Y-%m-%d').date())
            except ValueError:
                pass
        if date_to:
            try:
                visits = visits.filter(visit_date__lte=_dt.strptime(date_to, '%Y-%m-%d').date())
            except ValueError:
                pass
    else:
        visits = visits.filter(visit_date=filter_date)

    if sp_filter:
        visits = visits.filter(salesperson_id=sp_filter)
    if outcome_filter:
        visits = visits.filter(outcome=outcome_filter)

    # Per-salesperson summary (from filtered queryset)
    summary = {}
    for v in visits:
        key = v.salesperson_id
        if key not in summary:
            summary[key] = {'salesperson': v.salesperson, 'total': 0, 'order_placed': 0, 'no_contact': 0, 'other': 0}
        summary[key]['total'] += 1
        if v.outcome == 'order_placed':
            summary[key]['order_placed'] += 1
        elif v.outcome == 'no_contact':
            summary[key]['no_contact'] += 1
        else:
            summary[key]['other'] += 1

    paginator = Paginator(visits, 30)
    page_obj  = paginator.get_page(request.GET.get('page'))

    context = {
        'visits': page_obj,
        'page_obj': page_obj,
        'summary': list(summary.values()),
        'filter_date': filter_date,
        'salespersons': _gum().objects.filter(is_active=True).order_by('first_name'),
        'sp_filter': sp_filter,
        'outcome_filter': outcome_filter,
        'outcome_choices': BeatVisit.OUTCOME_CHOICES,
        'date_from': date_from,
        'date_to': date_to,
    }
    return render(request, 'beat/overview.html', context)


# ─────────────────────────────────────────────────────────────────────────────
# CUSTOMER TRACKER
# ─────────────────────────────────────────────────────────────────────────────
@login_required
def customer_tracker(request):
    """
    Customer tracker: active vs inactive customers by order amount, count & category.
    Active = placed at least one order within the last `inactive_days` days.
    """
    from django.core.paginator import Paginator
    from django.contrib.auth import get_user_model as _gum
    from django.db.models import Max, Min

    user = request.user
    is_admin = user.is_superuser or user.groups.filter(name='Admins').exists()

    # ── Filters ────────────────────────────────────────────────────────────
    category_filter = request.GET.get('category', '')
    sp_filter       = request.GET.get('sp', '')
    status_filter   = request.GET.get('status', '')   # active | inactive | ''
    search          = request.GET.get('q', '')
    ordering        = request.GET.get('order', '-last_order_date')
    try:
        inactive_days = max(int(request.GET.get('inactive_days', 30)), 1)
    except ValueError:
        inactive_days = 30

    cutoff = timezone.now().date() - timedelta(days=inactive_days)

    # ── Base queryset ───────────────────────────────────────────────────────
    qs = Customer.objects.select_related('sales_person')
    if not is_admin:
        qs = qs.filter(sales_person=user)

    if category_filter:
        qs = qs.filter(default_category=category_filter)
    if is_admin and sp_filter:
        qs = qs.filter(sales_person_id=sp_filter)
    if search:
        qs = qs.filter(
            Q(first_name__icontains=search) | Q(last_name__icontains=search) |
            Q(phone_number__icontains=search) | Q(address__icontains=search)
        )

    # ── Annotate order stats ────────────────────────────────────────────────
    qs = qs.annotate(
        total_orders=Count('order', distinct=True),
        total_value=Coalesce(Sum('order__total_amount'), Decimal('0')),
        last_order_date=Max('order__order_date'),
        first_order_date=Min('order__order_date'),
    )

    # Status filter (applied after annotation)
    if status_filter == 'active':
        qs = qs.filter(last_order_date__gte=cutoff)
    elif status_filter == 'inactive':
        qs = qs.filter(Q(last_order_date__lt=cutoff) | Q(last_order_date__isnull=True))

    # ── Ordering ───────────────────────────────────────────────────────────
    valid_orders = {
        'last_order_date', '-last_order_date',
        'total_orders', '-total_orders',
        'total_value', '-total_value',
        'first_name', '-first_name',
        'default_category',
    }
    if ordering not in valid_orders:
        ordering = '-last_order_date'
    qs = qs.order_by(ordering)

    # ── Summary stats ──────────────────────────────────────────────────────
    all_annotated = Customer.objects.select_related('sales_person')
    if not is_admin:
        all_annotated = all_annotated.filter(sales_person=user)
    all_annotated = all_annotated.annotate(last_order_date=Max('order__order_date'))

    total_customers = all_annotated.count()
    active_count    = all_annotated.filter(last_order_date__gte=cutoff).count()
    inactive_count  = all_annotated.filter(
        Q(last_order_date__lt=cutoff) | Q(last_order_date__isnull=True)
    ).count()
    no_orders_count = all_annotated.filter(last_order_date__isnull=True).count()

    # Category distribution
    from django.db.models.functions import Coalesce as _Coalesce
    category_stats = (
        Customer.objects
        .filter(sales_person=user) if not is_admin
        else Customer.objects.all()
    )
    category_stats = (
        category_stats
        .values('default_category')
        .annotate(count=Count('id'))
        .order_by('-count')
    )

    # Category choices for filter dropdown
    category_choices = (
        Customer.objects.values_list('default_category', flat=True)
        .distinct().exclude(default_category='').order_by('default_category')
    )

    paginator = Paginator(qs, 25)
    page_obj  = paginator.get_page(request.GET.get('page'))

    context = {
        'page_obj': page_obj,
        'customers': page_obj,
        'total_customers': total_customers,
        'active_count': active_count,
        'inactive_count': inactive_count,
        'no_orders_count': no_orders_count,
        'cutoff': cutoff,
        'inactive_days': inactive_days,
        'is_admin': is_admin,
        'salespersons': _gum().objects.filter(is_active=True).order_by('first_name') if is_admin else None,
        'category_filter': category_filter,
        'sp_filter': sp_filter,
        'status_filter': status_filter,
        'search': search,
        'ordering': ordering,
        'category_choices': list(category_choices),
        'category_stats': list(category_stats),
    }
    return render(request, 'customers/tracker.html', context)


# ─────────────────────────────────────────────────────────────────────────────
# SALES ANALYTICS DASHBOARD
# ─────────────────────────────────────────────────────────────────────────────
@login_required
def sales_analytics(request):
    """
    Rich analytics dashboard:
    - Per-rep daily targets (50K orders, 50K payments, 5 new outlets/month)
    - Repeat order frequency
    - Wholesaler reorder frequency
    - Daily order trend
    - Active vs inactive outlet breakdown
    - Top SKUs
    - Distribution channels
    - Overdue payments / problem identification
    """
    from django.contrib.auth import get_user_model as _gum
    from django.db.models import Max, Min
    from django.db.models.functions import TruncMonth

    user = request.user
    is_admin = user.is_superuser or user.groups.filter(name='Admins').exists()

    today         = timezone.now().date()
    date_from_str = request.GET.get('date_from', '')
    date_to_str   = request.GET.get('date_to', '')
    sp_filter     = request.GET.get('sp', '')

    try:
        date_from = _dt.strptime(date_from_str, '%Y-%m-%d').date() if date_from_str else (today - timedelta(days=30))
    except ValueError:
        date_from = today - timedelta(days=30)
    try:
        date_to = _dt.strptime(date_to_str, '%Y-%m-%d').date() if date_to_str else today
    except ValueError:
        date_to = today

    period_days = max((date_to - date_from).days + 1, 1)

    # ── Targets ────────────────────────────────────────────────────────────
    ORDERS_TARGET      = 50000   # KSh avg per day per rep
    PAYMENTS_TARGET    = 50000   # KSh avg per day per rep
    NEW_OUTLETS_TARGET = 5       # new customers per period per rep

    # ── Salespersons to evaluate ───────────────────────────────────────────
    UModel = _gum()
    if is_admin:
        reps_qs = UModel.objects.filter(is_active=True).exclude(is_superuser=True).order_by('first_name')
        if sp_filter:
            reps_qs = reps_qs.filter(pk=sp_filter)
    else:
        reps_qs = UModel.objects.filter(pk=user.pk)

    def _pct_color(p):
        if p >= 100: return 'success'
        if p >= 70:  return 'warning'
        return 'danger'

    rep_stats = []
    for rep in reps_qs:
        orders_qs   = Order.objects.filter(sales_person=rep, order_date__range=[date_from, date_to])
        order_total = orders_qs.aggregate(s=Coalesce(Sum('total_amount'), Decimal('0')))['s']
        order_count = orders_qs.count()
        daily_avg_orders = float(order_total) / period_days

        payments_qs      = Payment.objects.filter(recorded_by=rep, payment_date__range=[date_from, date_to])
        payment_total    = payments_qs.aggregate(s=Coalesce(Sum('amount'), Decimal('0')))['s']
        daily_avg_pay    = float(payment_total) / period_days

        new_outlets = Customer.objects.filter(
            sales_person=rep,
            created_at__date__range=[date_from, date_to]
        ).count()

        # Today's numbers
        todays_orders   = Order.objects.filter(sales_person=rep, order_date=today).aggregate(s=Coalesce(Sum('total_amount'), Decimal('0')))['s']
        todays_payments = Payment.objects.filter(recorded_by=rep, payment_date=today).aggregate(s=Coalesce(Sum('amount'), Decimal('0')))['s']

        orders_pct   = min(int(daily_avg_orders / ORDERS_TARGET * 100), 100)
        payments_pct = min(int(daily_avg_pay / PAYMENTS_TARGET * 100), 100)
        outlets_pct  = min(int(new_outlets / NEW_OUTLETS_TARGET * 100), 100)

        rep_stats.append({
            'rep': rep,
            'order_total': order_total,
            'order_count': order_count,
            'daily_avg_orders': round(daily_avg_orders, 0),
            'orders_pct': orders_pct,
            'orders_color': _pct_color(orders_pct),
            'payment_total': payment_total,
            'daily_avg_payments': round(daily_avg_pay, 0),
            'payments_pct': payments_pct,
            'payments_color': _pct_color(payments_pct),
            'new_outlets': new_outlets,
            'outlets_pct': outlets_pct,
            'outlets_color': _pct_color(outlets_pct),
            'todays_orders': todays_orders,
            'todays_payments': todays_payments,
        })

    # ── Repeat order frequency (customers who ordered >1 time) ─────────────
    base_order_qs = Order.objects.filter(order_date__range=[date_from, date_to])
    if not is_admin:
        base_order_qs = base_order_qs.filter(sales_person=user)

    repeat_orders = (
        base_order_qs
        .values('customer__id', 'customer__first_name', 'customer__last_name',
                'customer__default_category', 'customer__phone_number')
        .annotate(order_count=Count('id'), total_value=Sum('total_amount'))
        .filter(order_count__gt=1)
        .order_by('-order_count')[:15]
    )

    # ── Wholesaler reorder frequency ───────────────────────────────────────
    wholesaler_reorders = (
        base_order_qs
        .filter(customer_category__in=['wholesale', 'distributor', 'factory'])
        .values('customer__id', 'customer__first_name', 'customer__last_name',
                'customer__default_category', 'customer__phone_number')
        .annotate(
            order_count=Count('id'),
            total_value=Sum('total_amount'),
            avg_value=Avg('total_amount'),
        )
        .order_by('-order_count')[:15]
    )

    # ── Daily order trend ──────────────────────────────────────────────────
    daily_orders = (
        base_order_qs
        .values('order_date')
        .annotate(count=Count('id'), total=Coalesce(Sum('total_amount'), Decimal('0')))
        .order_by('order_date')
    )

    # ── Active vs Inactive outlets (all time) ──────────────────────────────
    cutoff_30 = today - timedelta(days=30)
    cutoff_60 = today - timedelta(days=60)
    cutoff_90 = today - timedelta(days=90)

    all_cust = Customer.objects.annotate(last_order=Max('order__order_date'))
    if not is_admin:
        all_cust = all_cust.filter(sales_person=user)

    active_30   = all_cust.filter(last_order__gte=cutoff_30).count()
    active_60   = all_cust.filter(last_order__lt=cutoff_30, last_order__gte=cutoff_60).count()
    active_90   = all_cust.filter(last_order__lt=cutoff_60, last_order__gte=cutoff_90).count()
    inactive_90 = all_cust.filter(Q(last_order__lt=cutoff_90) | Q(last_order__isnull=True)).count()

    # ── Top SKUs by revenue ────────────────────────────────────────────────
    item_qs = OrderItem.objects.filter(order__order_date__range=[date_from, date_to])
    if not is_admin:
        item_qs = item_qs.filter(order__sales_person=user)

    top_products = (
        item_qs
        .values('product__id', 'product__name')
        .annotate(
            qty_sold=Sum('quantity'),
            revenue=Coalesce(Sum('line_total'), Decimal('0')),
            times_ordered=Count('order', distinct=True),
        )
        .order_by('-revenue')[:15]
    )

    # ── Distribution channels (by customer category) ───────────────────────
    channel_breakdown = (
        base_order_qs
        .values('customer_category')
        .annotate(
            order_count=Count('id'),
            total_value=Coalesce(Sum('total_amount'), Decimal('0')),
        )
        .order_by('-total_value')
    )

    # ── New outlets per month trend ────────────────────────────────────────
    new_outlet_trend = (
        Customer.objects
        .filter(created_at__date__range=[date_from, date_to])
        .annotate(month=TruncMonth('created_at'))
        .values('month')
        .annotate(count=Count('id'))
        .order_by('month')
    )
    if not is_admin:
        new_outlet_trend = new_outlet_trend.filter(sales_person=user)

    # ── Identify problems: overdue unpaid orders ───────────────────────────
    overdue_qs = Order.objects.filter(
        order_date__lt=cutoff_30,
        paid_status=False
    ).select_related('customer', 'sales_person').annotate(
        outstanding=ExpressionWrapper(
            F('total_amount') - F('amount_paid'),
            output_field=DecimalField()
        )
    ).order_by('-outstanding')
    if not is_admin:
        overdue_qs = overdue_qs.filter(sales_person=user)
    overdue_orders = overdue_qs[:10]

    # ── JSON data for charts ───────────────────────────────────────────────
    daily_labels  = [str(d['order_date']) for d in daily_orders]
    daily_totals  = [float(d['total']) for d in daily_orders]
    channel_labels = [d['customer_category'] or 'Unknown' for d in channel_breakdown]
    channel_vals   = [float(d['total_value']) for d in channel_breakdown]

    context = {
        'is_admin': is_admin,
        'rep_stats': rep_stats,
        'salespersons': (
            _gum().objects.filter(is_active=True).exclude(is_superuser=True).order_by('first_name')
            if is_admin else None
        ),
        'sp_filter': sp_filter,
        'date_from': date_from_str or (today - timedelta(days=30)).strftime('%Y-%m-%d'),
        'date_to': date_to_str or today.strftime('%Y-%m-%d'),
        'period_days': period_days,
        'ORDERS_TARGET': ORDERS_TARGET,
        'PAYMENTS_TARGET': PAYMENTS_TARGET,
        'NEW_OUTLETS_TARGET': NEW_OUTLETS_TARGET,
        'repeat_orders': list(repeat_orders),
        'wholesaler_reorders': list(wholesaler_reorders),
        'daily_orders': list(daily_orders),
        'daily_labels_json': json.dumps(daily_labels),
        'daily_totals_json': json.dumps(daily_totals),
        'active_30': active_30,
        'active_60': active_60,
        'active_90': active_90,
        'inactive_90': inactive_90,
        'top_products': list(top_products),
        'channel_breakdown': list(channel_breakdown),
        'channel_labels_json': json.dumps(channel_labels),
        'channel_vals_json': json.dumps(channel_vals),
        'new_outlet_trend': list(new_outlet_trend),
        'overdue_orders': overdue_orders,
    }
    return render(request, 'analytics/sales_analytics.html', context)
